# assume objective function to be zero
# reports all feasible solutions to output through pool search methods

import pandas as pd
import numpy as np
import gurobipy as gp
from gurobipy import GRB, nlfunc
from gurobipy import quicksum
from openpyxl import load_workbook
import os
import argparse
from typing import Optional

# Generate versioned input tables (V1/V2/V3/V4/D2)
try:
    # Local import; both files live in the same folder
    from generate_outputs import main as _generate_outputs_main
except Exception:
    _generate_outputs_main = None  # Fallback if generator isn't importable

# Get the directory where this script is located
script_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(script_dir)

# Command-line argument parsing
parser = argparse.ArgumentParser(description='Run Aviation Intensity Standard model')
parser.add_argument('--input-folder', type=str, default='data_input',
                    help='Input data folder name (default: data_input)')
parser.add_argument('--output-folder', type=str, default='results',
                    help='Output results folder name (default: results)')
parser.add_argument('--k-range', type=int, default=1,
                    help='Number of robustness iterations (default: 1, max: 5)')
args = parser.parse_args()

# Set up paths
DATA_INPUT_FOLDER = args.input_folder
OUTPUT_FOLDER = args.output_folder
K_RANGE = args.k_range

data_input_dir = os.path.join(parent_dir, DATA_INPUT_FOLDER)
results_dir = os.path.join(parent_dir, OUTPUT_FOLDER)
intermediate_dir = os.path.join(parent_dir, 'intermediate')

# Create results directory if it doesn't exist
os.makedirs(results_dir, exist_ok=True)

# Create the solution DataFrame
solution = pd.DataFrame(
    np.zeros((80, 6)),  # 70 rows and 6 columns filled with zeros
    columns=['Name', 'Mean', 'Q_10th', 'Q_33th', 'Q_67th', 'Q_90th']
)
# Assign the 'Name' column
solution['Name'] = [
    'Biofuel quantity', 'Q_B100-soyoil-CA', 'Q_B100-animalfat-CA', 'Q_RD-soyoil-CA', 'Q_RD-animalfat-CA', 'Q_SAF-animalfat-CA', 'Q_gasE100-corn-CA',
                 'Q_B100-soyoil-NC', 'Q_B100-animalfat-NC', 'Q_RD-soyoil-NC', 'Q_RD-animalfat-NC', 'Q_SAF-animalfat-NC', 'Q_gasE100-corn-NC',
                 'Q_gasE100-sugarcane-CA', 'Q_gasE100-sugarcane-NC', 'Q_SAF-soyoil-CA', 'Q_SAF-soyoil-NC', 'Q_SAF-sugarcane-ETJ-CA', 'Q_SAF-sugarcane-ETJCCS-CA',
                 'Q_SAF-corn-ETJ-CA', 'Q_SAF-corn-ETJCCS-CA', 'Q_SAF-sugarcane-ETJ-NC', 'Q_SAF-sugarcane-ETJCCS-NC', 'Q_SAF-corn-ETJ-NC', 'Q_SAF-corn-ETJCCS-NC',
                 'Prices', 'State_tax_credit', 'Additional_tax_credit','P_D2', 'P_D4', 'P_D6', 'P_LCFS', 'P_SAF', 'P_B100_CA', 'P_RD_CA', 'P_SAF_CA', 'P_gasE100_CA', 'P_B100_NC', 'P_RD_NC', 'P_SAF_NC', 'P_gasE100_NC',
                 'P_B0_CA', 'P_E0_CA', 'P_J0_CA', 'P_B0_NC', 'P_E0_NC', 'P_J0_NC', 'Feedstock', 'P_soyoil', 'P_animal fat', 'P_corn', 'P_sugarcane', 'Q_soyoil', 'Q_animal fat', 'Q_corn', 'Q_sugarcane',
                 'Conventional fuel quantity', 'Q_B0_CA', 'Q_E0_CA', 'Q_J0_CA', 'Q_B0_NC', 'Q_E0_NC', 'Q_J0_NC',
                 'Blended-Fuel', 'D_CA', 'G_CA', 'J_CA', 'D_NC', 'G_NC', 'J_NC','Aviation_intensity_standard',
                 'DGE_price',
                 'P_B100_CA_DGE','P_RD_CA_DGE', 'P_SAF_CA_DGE', 'P_gasE100_CA_DGE', 'P_B100_NC_DGE', 'P_RD_NC_DGE', 'P_SAF_NC_DGE', 'P_gasE100_NC_DGE'
]

# Fitted_quantity DataFrame
Fitted_quantity = pd.DataFrame(
    np.zeros((40, 6)),
    columns=['Name', 'Mean', 'Q_10th', 'Q_33th', 'Q_67th', 'Q_90th']
)
Fitted_quantity['Name'] = [
    'RVO_D2', 'RVO_BBD', 'RVO_RF', 'RV_BBD', 'RV_RF', 'Fitted_BBD', 'Fitted_RF', 'Fitted_LCFS_deficit-credits','Fitted_SAF_LCFS_deficit-credits',
    'Total', 'BD_all', 'RD_all', 'BBD_all', 'B0_all', 'E100_all', 'E0_all', 'SAF_HEFA_all', 'SAF_ETJ_all', 'SAF_all', 'J0_all',
    'CA', 'BD_CA', 'RD_CA', 'B0_CA', 'E100_CA', 'E0_CA', 'SAF_CA', 'J0_CA',
    'Feedstock', 'soyoil', 'animal fat', 'corn', 'sugarcane', 'Demand_all', 'D_CA', 'G_CA', 'J_CA', 'D_NC', 'G_NC', 'J_NC'
]

# Fuel_price DataFrame
Fuel_price = pd.DataFrame(
    np.zeros((34, 6)),
    columns=['Name', 'Mean', 'Q_10th', 'Q_33th', 'Q_67th', 'Q_90th']
)
Fuel_price['Name'] = [
    'Blended fuel', 'Blended-diesel-CA', 'Blended-diesel-NC', 'Blended-gasoline-CA', 'Blended-gasoline-NC','Aviation-CA', 'Aviation-NC','E100-CA', 'E100-NC', 'E10-CA', 'E10-NC', 
    'Finished biofuel volume per gallon', 'B100', 'RD', 'E100', 'SAF',
    'Conventional fuel', 'B0-CA', 'E0-CA','J0-CA','B0-NC', 'E0-NC', 'J0-NC',
    'Implicit tax', 'RFS_RVO_road','RFS_RVO_jet', 'LCFS_CA_diesel', 'LCFS_CA_gasoline','Aviation_intensity_standard',
    'Finished biofuel DGE per gallon', 'B100', 'RD', 'E100', 'SAF'
]

# Fuel_CI DataFrame
Fuel_CI = pd.DataFrame(
    np.zeros((15, 6)),
    columns=['Name', 'Mean', 'Q_10th', 'Q_33th', 'Q_67th', 'Q_90th']
)
Fuel_CI['Name'] = [
    'Total', 'Blended fuel', 'Blended-diesel', 'Blended-gasoline', 'Blended-jet', 'CA', 'Blended-fuel-CA', 'Blended-diesel-CA',
    'Blended-gasoline-CA', 'Blended-jet-CA', 'NC', 'Blended-fuel-NC', 'Blended-diesel-NC', 'Blended-gasoline-NC', 'Blended-jet-NC'
]

# Total DataFrame
Total = pd.DataFrame(
    np.zeros((53, 6)),
    columns=['Name', 'Mean', 'Q_10th', 'Q_33th', 'Q_67th', 'Q_90th']
)
Total['Name'] = [
    'Total', 'Emissions_D', 'Emissions_G', 'Emissions_J', 'Emissions_total', 'Costs_D', 'Costs_G', 'Costs_J', 'Costs_total',
    'Subsidy', 'Taxpayer', 'IRA', '45Z', '45Q', 'State_tax_credit', 'Additional_SAF_tax',
    'RFS_credits_road_CA', 'RFS_credits_road_NC', 'RFS_credits_jet_CA', 'RFS_credits_jet_NC',
    'RFS_deficits_road_CA', 'RFS_deficits_road_NC', 'RFS_deficits_jet_CA', 'RFS_deficits_jet_NC',
    'LCFS_CA_biofuel_credits', 'LCFS_CA_other_credits', 'LCFS_CA_all_credits', 'LCFS_all_jet', 'LCFS_all_credits_jet_CA', 'LCFS_all_credits_jet_NC',
    'LCFS_all_deficits_jet_CA', 'LCFS_all_deficits_jet_NC', 'Subsidy_total', 
    'Taxpayer_D', 'Taxpayer_G', 'Taxpayer_J','LCFS_CA_biofuel_credits_D','LCFS_CA_biofuel_credits_G', 'LCFS_CA_biofuel_credits_jet','LCFS_CA_deficits_D','LCFS_CA_deficits_G',
    'RFS_credits_D_CA', 'RFS_credits_D_NC','RFS_credits_G_CA', 'RFS_credits_G_NC', 'RFS_credits_J_CA', 'RFS_credits_J_NC',
    'RFS_deficits_D_CA', 'RFS_deficits_D_NC','RFS_deficits_G_CA', 'RFS_deficits_G_NC', 'RFS_deficits_J_CA', 'RFS_deficits_J_NC'
]

# Feedstock_source DataFrame
Feedstock_source = pd.DataFrame(
    np.zeros((12, 6)),
    columns=['Name', 'Mean', 'Q_10th', 'Q_33th', 'Q_67th', 'Q_90th']
)
Feedstock_source['Name'] = [
    'B100-soyoil', 'B100-animalfat', 'RD-soyoil', 'RD-animalfat', 'SAF-soyoil', 'SAF-animalfat', 'SAF-corn', 'SAF-sugarcane',
    'SAF-corn-CCS', 'SAF-sugarcane-CCS', 'gasE100-corn', 'gasE100-sugarcane'
]

for k in range(K_RANGE):    
    # Read input_biofuel from generated V4 (OLD rules, NC-zeroed CI columns)
    intermediate_dir = os.path.join(parent_dir, 'intermediate')
    v4_path = os.path.join(intermediate_dir, 'input_biofuel_V4.xlsx')
    if not os.path.exists(v4_path):
        # Attempt to generate all outputs if missing
        if _generate_outputs_main is not None:
            try:
                _generate_outputs_main()
            except Exception as e:
                raise RuntimeError(f"Failed to generate input_biofuel_V4.xlsx via generate_outputs.py: {e}")
        else:
            raise FileNotFoundError(
                "input_biofuel_V4.xlsx not found and generate_outputs.py unavailable to build it."
            )
    if not os.path.exists(v4_path):
        raise FileNotFoundError(f"Expected generated file is missing: {v4_path}")

    input_biofuel = pd.read_excel(v4_path, index_col=0)
    
    # Drop old CI_std_LCFS column and rename CI_std_LCFS_new to CI_std_LCFS (using 30% reduction values)
    if 'CI_std_LCFS' in input_biofuel.columns:
        input_biofuel = input_biofuel.drop(columns=['CI_std_LCFS'])
    if 'CI_std_LCFS_CA' in input_biofuel.columns:
        input_biofuel = input_biofuel.drop(columns=['CI_std_LCFS_CA'])
    if 'CI_std_LCFS_CA_new' in input_biofuel.columns:
        input_biofuel = input_biofuel.rename(columns={'CI_std_LCFS_CA_new': 'CI_std_LCFS_CA'})
    
    pd.set_option('display.max_columns', None)
    print(input_biofuel.head(10))
    
    # Ensure results folder exists for outputs later
    results_dir = OUTPUT_FOLDER  # Already has full path from parent_dir
    os.makedirs(results_dir, exist_ok=True)
    
    # Read feedstock supply parameters from data_input folder
    data_input_dir = DATA_INPUT_FOLDER  # Already has full path from parent_dir
    feedstock_path = os.path.join(data_input_dir, 'feedstock_supply.xlsx')
    input_feedstock = pd.read_excel(feedstock_path, index_col=0)
    
    demand_path = os.path.join(data_input_dir, 'fuel_demand.xlsx')
    input_demand = pd.read_excel(demand_path, sheet_name='V1', index_col=0)
    input_demand = input_demand.iloc[:, [0, 1, 2, k+3, 8, 9]]
    input_demand.columns = ['F', 'j', 'PC_F0', 'alpha', 'beta', 'ETJ_percent']

    compliance_path = os.path.join(data_input_dir, 'Policy_constraints.xlsx')
    input_compliance = pd.read_excel(compliance_path, sheet_name='V2', index_col=0)
    input_compliance = input_compliance.iloc[:, [0, k+1, 6, 7]]
    input_compliance.columns = ['mandate', 'Q', 'CI_LCFS', 'ED_LCFS']
    print(input_compliance)

    state_tax_path = os.path.join(data_input_dir, 'state_tax.xlsx')
    state_level_tax = pd.read_excel(state_tax_path)

    # Read jet fuel percentage from the new spreadsheet
    jet_volume = state_level_tax.iloc[:, k+1]
    print(jet_volume)
    state_tax_credit = state_level_tax['State tax credit'].values
    
    # # # Assign to breaks, add 0 at front and 1 at end
    breaks = list(jet_volume)
    vals = list(state_tax_credit) + [0.0]
    print("Breaks:", breaks)
    print("Values:", vals)

    # Define constants
    alpha_eth_d = 1.5
    beta_eth_d = 0.2
    P_cap = 250
    duty_e100sugarcane = 0.125  # 12.5% import tariff on ethanol from sugarcane
    P_cap_SAF = 800
    epsilon = 1e-4
    Sigma = 1e-6 #Should be ≳ feasibility tol
    iteration = 0
    max_total_profit = -float('inf')
    best_df = None
    best_break_idx = None
    state_tax_profit = 0.0
    
    # Initialize list to track objective values for each break
    objective_values = []

    for break_idx, break_val in enumerate(breaks):
        val = vals[break_idx]
        val_pre = vals[break_idx - 1] if break_idx > 0 else 0.0
        break_val_pre = breaks[break_idx - 1] if break_idx > 0 else 0.0
        break_val_step = break_val_pre - breaks[break_idx-2] if break_idx > 1 else break_val_pre
        print(f"Break {break_idx}: break = {break_val}, val = {val},val_pre = {val_pre},break_val_pre = {break_val_pre},break_val_step = {break_val_step}")
        state_tax_profit_step = (-val + val_pre) * break_val_step if break_idx > 0 else 0.0
        state_tax_profit = state_tax_profit + state_tax_profit_step

        print(f"State tax profit step: {state_tax_profit_step}, Cumulative state tax profit: {state_tax_profit}")
        #print(f"Break {break_idx}: break = {break_val}, val = {val}")
        # Set up your model here, using break_val as needed
        # Create a new model
        m = gp.Model("Aviation_Intensity_Standard_Model")   

        # Create variables for biofuel quantity
        biofuel_quantity_Q = m.addVars(24, vtype=GRB.CONTINUOUS, name="Q")
        biofuel_quantity_Q_equiv = m.addVars(24, vtype=GRB.CONTINUOUS, name="Q_equiv")
        # Set row names for biofuel_quantity
        biofuel_quantity = pd.DataFrame({
            'Q': [biofuel_quantity_Q[i] for i in range(24)],
            'Q_equiv': [biofuel_quantity_Q_equiv[i] for i in range(24)]
        }, index=input_biofuel.index)

        # Create multipliers mu
        mu_value = m.addVars(24, vtype=GRB.CONTINUOUS, name="mu_value")
        # Set row names for mu
        mu = pd.DataFrame({
            'value': [mu_value[i] for i in range(24)]
        }, index=input_biofuel.index)

        muzeta_value = m.addVars(24, vtype=GRB.BINARY, name="muzeta_value")
        muzeta = pd.DataFrame({
            'value': [muzeta_value[i] for i in range(24)]
        }, index=input_biofuel.index)

        # Create multipliers nu
        nu_value = m.addVars(8, vtype=GRB.CONTINUOUS, name="nu_value")
        nuzeta_value = m.addVars(8, vtype=GRB.BINARY, name="nuzeta_value")
        # Set row names for nu
        nu_index = ['B100_CA', 'RD_CA', 'SAF_CA', 'gasE100_CA', 'B100_NC', 'RD_NC', 'gasE100_NC', 'SAF_NC']
        nu = pd.DataFrame({
            'value': [nu_value[i] for i in range(8)],
            'f': ['B100', 'RD', 'SAF', 'gasE100', 'B100', 'RD', 'gasE100', 'SAF'],
            'j': ['CA', 'CA', 'CA', 'CA', 'NC', 'NC', 'NC', 'NC']
        }, index=nu_index)
        nuzeta = pd.DataFrame({
            'value': [nuzeta_value[i] for i in range(8)],
            'f': ['B100', 'RD', 'SAF', 'gasE100', 'B100', 'RD', 'gasE100', 'SAF'],
            'j': ['CA', 'CA', 'CA', 'CA', 'NC', 'NC', 'NC', 'NC']
        }, index=nu_index)

        # Create multipliers for HEFA constraints
        hefa_value = m.addVars(2, vtype=GRB.CONTINUOUS, name="hefa_value")
        hefazeta_value = m.addVars(2, vtype=GRB.BINARY, name="hefazeta_value")
                        
        # RIN prices
        RIN_prices_P = m.addVars(2, vtype=GRB.CONTINUOUS, name="RIN_prices_P")
        RIN_prices = pd.DataFrame({
            'P': [RIN_prices_P[i] for i in range(2)],
            'fueltype': ['B100,RD,SAF', 'gasE100'],
            'code': ['D4', 'D6']
        }, index=['D4', 'D6'])

        # RVO percent
        RVO_percent_P = m.addVars(2, vtype=GRB.CONTINUOUS, name="RVO_percent_P")
        RVO_percent = pd.DataFrame({
            'P': [RVO_percent_P[i] for i in range(2)],
            'mandate': ['BBD', 'RF']
        }, index=['BBD', 'RF'])

        # LCFS prices
        LCFS_prices_P = m.addVar(vtype=GRB.CONTINUOUS, name="LCFS_prices_P")
        LCFS_prices = pd.DataFrame({
            'P': [LCFS_prices_P],
            'type': ['LCFS']
        }, index=['LCFS'])

        # Create the LCFS_prices_SAF DataFrame
        LCFS_prices_SAF_P = m.addVar(vtype=GRB.CONTINUOUS, name="LCFS_prices_SAF_P")
        LCFS_prices_SAF = pd.DataFrame({
            'P': [LCFS_prices_SAF_P],
            'type': ['LCFS_SAF']
        }, index=['LCFS_SAF'])

        # # SAF credit prices
        # SAFcredit_prices_P = m.addVar(vtype=GRB.CONTINUOUS, name="SAFcredit_prices_P")
        # SAFcredit_prices = pd.DataFrame({
        #     'P': [SAFcredit_prices_P],
        #     'type': ['SAF']
        # }, index=['SAF'])

        # SAF credit prices
        SAF_intensity_standard_P = m.addVar(vtype=GRB.CONTINUOUS, name="SAF_intensity_standard")
        SAF_intensity_standard = pd.DataFrame({
            'P': [SAF_intensity_standard_P],
            'type': ['SAF']
        }, index=['SAF'])

        # Create a new column 'ci_std_LCFS' in input_biofuel
        input_biofuel['CI_std_LCFS'] = 0  # Initialize with 0 for all non-SAF types

        # Set 'ci_std_LCFS' to SAF_intensity_standard_P for SAF type
        input_biofuel.loc[input_biofuel['f'] == 'SAF', 'CI_std_LCFS'] = SAF_intensity_standard_P

        print(input_biofuel['CI_std_LCFS'])

        # Feedstock prices
        feedstock_prices_P = m.addVars(4, vtype=GRB.CONTINUOUS, name="feedstock_prices_P")
        feedstock_prices = pd.DataFrame({
            'P': [feedstock_prices_P[i] for i in range(4)],
            's': ['soyoil', 'animalfat', 'corn', 'sugarcane']
        }, index=['soyoil', 'animalfat', 'corn', 'sugarcane'])

        # Biofuel prices
        biofuel_prices_P = m.addVars(8, vtype=GRB.CONTINUOUS, name="biofuel_prices_P")
        biofuel_prices = pd.DataFrame({
            'P': [biofuel_prices_P[i] for i in range(8)],
            'f': ['B100', 'RD', 'SAF', 'gasE100', 'B100', 'RD', 'SAF', 'gasE100'],
            'j': ['CA', 'CA', 'CA', 'CA', 'NC', 'NC', 'NC', 'NC']
        }, index=['B100_CA', 'RD_CA', 'SAF_CA', 'gasE100_CA', 'B100_NC', 'RD_NC', 'SAF_NC', 'gasE100_NC'])

        # Conventional fuel prices
        conventionalfuel_prices_P = m.addVars(6, vtype=GRB.CONTINUOUS, name="conventionalfuel_prices_P")
        conventionalfuel_prices = pd.DataFrame({
            'P': [conventionalfuel_prices_P[i] for i in range(6)],
            'F': ['D', 'G', 'J', 'D', 'G', 'J'],
            'j': ['CA', 'CA', 'CA', 'NC', 'NC', 'NC']
        }, index=['D_CA', 'G_CA', 'J_CA', 'D_NC', 'G_NC', 'J_NC'])

        # Feedstock quantity
        feedstock_quantity_Q = m.addVars(4, vtype=GRB.CONTINUOUS, name="feedstock_quantity_Q")
        feedstock_quantity = pd.DataFrame({
            'Q': [feedstock_quantity_Q[i] for i in range(4)],
            's': ['soyoil', 'animalfat', 'corn', 'sugarcane']
        }, index=['soyoil', 'animalfat', 'corn', 'sugarcane'])

        # Total demand
        demand_quantity_Q = m.addVars(6, vtype=GRB.CONTINUOUS, name="demand_quantity_Q")
        demand_quantity_Q_actual = m.addVars(6, vtype=GRB.CONTINUOUS, name="demand_quantity_Q_actual")
        demand_quantity = pd.DataFrame({
            'F': ['D', 'G', 'J', 'D', 'G', 'J'],
            'j': ['CA', 'CA', 'CA', 'NC', 'NC', 'NC'],
            'Q': [demand_quantity_Q[i] for i in range(6)],
            'Q_actual': [demand_quantity_Q_actual[i] for i in range(6)]
        }, index=['D_CA', 'G_CA', 'J_CA', 'D_NC', 'G_NC', 'J_NC'])

        # Blended fuel price
        demand_price_P = m.addVars(6, vtype=GRB.CONTINUOUS, name="demand_price_P")
        demand_price = pd.DataFrame({
            'F': ['D', 'G', 'J', 'D', 'G', 'J'],
            'j': ['CA', 'CA', 'CA', 'NC', 'NC', 'NC'],
            'P': [demand_price_P[i] for i in range(6)]
        }, index=['D_CA', 'G_CA', 'J_CA', 'D_NC', 'G_NC', 'J_NC'])

        # Conventional fuel quantity
        conventionalfuel_quantity_Q = m.addVars(6, vtype=GRB.CONTINUOUS, name="conventionalfuel_quantity_Q")
        conventionalfuel_quantity = pd.DataFrame({
            'Q': [conventionalfuel_quantity_Q[i] for i in range(6)],
            'j': ['CA', 'CA', 'CA', 'NC', 'NC', 'NC'],
            'F': ['D', 'G', 'J', 'D', 'G', 'J'],
            'CI_LCFS': [0, 0, 0, 0, 0, 0],
            'CI_std_LCFS': [0, 0, 0, 0, 0, 0],
            'ED_LCFS': [0, 0, 0, 0, 0, 0]
        }, index=['B0_CA', 'E0_CA', 'J0_CA', 'B0_NC', 'E0_NC', 'J0_NC'])

        conventionalfuel_quantity.loc['B0_CA':'J0_CA', 'CI_LCFS'] = input_compliance.loc[input_compliance['mandate'] == 'LCFS', 'CI_LCFS'].values
        conventionalfuel_quantity.loc['B0_NC':'J0_NC', 'CI_LCFS'] = input_compliance.loc[input_compliance['mandate'] == 'LCFS', 'CI_LCFS'].values
        

        conventionalfuel_quantity.loc['B0_CA', 'CI_std_LCFS'] = input_biofuel.loc[(input_biofuel['F'] == 'D') & (input_biofuel['j'] == 'CA'), 'CI_std_LCFS'].unique()[0]
        conventionalfuel_quantity.loc['E0_CA', 'CI_std_LCFS'] = input_biofuel.loc[(input_biofuel['F'] == 'G') & (input_biofuel['j'] == 'CA'), 'CI_std_LCFS'].unique()[0]
        conventionalfuel_quantity.loc['J0_CA', 'CI_std_LCFS'] = SAF_intensity_standard_P
        conventionalfuel_quantity.loc['B0_NC', 'CI_std_LCFS'] = input_biofuel.loc[(input_biofuel['F'] == 'D') & (input_biofuel['j'] == 'CA'), 'CI_std_LCFS'].unique()[0]
        conventionalfuel_quantity.loc['E0_NC', 'CI_std_LCFS'] = input_biofuel.loc[(input_biofuel['F'] == 'G') & (input_biofuel['j'] == 'CA'), 'CI_std_LCFS'].unique()[0]
        conventionalfuel_quantity.loc['J0_NC', 'CI_std_LCFS'] = SAF_intensity_standard_P

        conventionalfuel_quantity.loc['B0_CA', 'CI_std_LCFS_CA'] = input_biofuel.loc[(input_biofuel['F'] == 'D') & (input_biofuel['j'] == 'CA'), 'CI_std_LCFS_CA'].unique()[0]
        conventionalfuel_quantity.loc['E0_CA', 'CI_std_LCFS_CA'] = input_biofuel.loc[(input_biofuel['F'] == 'G') & (input_biofuel['j'] == 'CA'), 'CI_std_LCFS_CA'].unique()[0]
        conventionalfuel_quantity.loc['J0_CA', 'CI_std_LCFS_CA'] = input_biofuel.loc[(input_biofuel['F'] == 'J') & (input_biofuel['j'] == 'CA'), 'CI_std_LCFS_CA'].unique()[0]
        conventionalfuel_quantity.loc['B0_NC', 'CI_std_LCFS_CA'] = input_biofuel.loc[(input_biofuel['F'] == 'D') & (input_biofuel['j'] == 'CA'), 'CI_std_LCFS_CA'].unique()[0]
        conventionalfuel_quantity.loc['E0_NC', 'CI_std_LCFS_CA'] = input_biofuel.loc[(input_biofuel['F'] == 'G') & (input_biofuel['j'] == 'CA'), 'CI_std_LCFS_CA'].unique()[0]
        conventionalfuel_quantity.loc['J0_NC', 'CI_std_LCFS_CA'] = input_biofuel.loc[(input_biofuel['F'] == 'J') & (input_biofuel['j'] == 'CA'), 'CI_std_LCFS_CA'].unique()[0]

        conventionalfuel_quantity.loc['B0_CA':'J0_CA', 'ED_LCFS'] = input_compliance.loc[input_compliance['mandate'] == 'LCFS', 'ED_LCFS'].values
        conventionalfuel_quantity.loc['B0_NC':'J0_NC', 'ED_LCFS'] = input_compliance.loc[input_compliance['mandate'] == 'LCFS', 'ED_LCFS'].values

        print(conventionalfuel_quantity)

        # Price break up: RIN/LCFS obligation
        Price_bio = pd.DataFrame({
            'f': input_biofuel['f'],
            's': input_biofuel['s'],
            'j': input_biofuel['j'],
            'F': input_biofuel['F'],
            'RIN_obligation': [m.addVar(vtype=GRB.CONTINUOUS, name=f"RIN_obligation_{i}") for i in range(len(input_biofuel))],
            'LCFS_CA_obligation': [m.addVar(vtype=GRB.CONTINUOUS, name=f"LCFS_CA_obligation_{i}") for i in range(len(input_biofuel))],
            'LCFS_SAF_obligation': [m.addVar(vtype=GRB.CONTINUOUS, name=f"LCFS_SAF_obligation_{i}") for i in range(len(input_biofuel))],
            'Tax': [m.addVar(vtype=GRB.CONTINUOUS, name=f"Tax_{i}") for i in range(len(input_biofuel))],
            'P_all': [m.addVar(vtype=GRB.CONTINUOUS, name=f"P_all_{i}") for i in range(len(input_biofuel))],
            'PC_F0': [m.addVar(vtype=GRB.CONTINUOUS, name=f"PC_F0_{i}") for i in range(len(input_biofuel))],
            'P_RIN_SAF_LCFS_detach': [m.addVar(vtype=GRB.CONTINUOUS, name=f"P_RIN_SAF_LCFS_detach_{i}") for i in range(len(input_biofuel))],
            'P_RIN_SAF_detach': [m.addVar(vtype=GRB.CONTINUOUS, name=f"P_RIN_SAF_detach_{i}") for i in range(len(input_biofuel))],
            'Q_actual': [m.addVar(vtype=GRB.CONTINUOUS, name=f"Q_actual_{i}") for i in range(len(input_biofuel))],
            'Q_equiv': [m.addVar(vtype=GRB.CONTINUOUS, name=f"Q_equiv_{i}") for i in range(len(input_biofuel))]
        }, index=input_biofuel.index)

        Price_conventional = pd.DataFrame({
            'f': ['B0', 'E0', 'J0', 'B0', 'E0', 'J0'],
            's': ['', '', '', '', '', ''],
            'j': ['CA', 'CA', 'CA', 'NC', 'NC', 'NC'],
            'F': ['D', 'G', 'J', 'D', 'G', 'J'],
            'RIN_obligation': [m.addVar(vtype=GRB.CONTINUOUS, name=f"RIN_obligation_{i+len(input_biofuel)}") for i in range(6)],
            'LCFS_CA_obligation': [m.addVar(vtype=GRB.CONTINUOUS, name=f"LCFS_CA_obligation_{i+len(input_biofuel)}") for i in range(6)],
            'LCFS_SAF_obligation': [m.addVar(vtype=GRB.CONTINUOUS, name=f"LCFS_SAF_obligation_{i+len(input_biofuel)}") for i in range(6)],
            'Tax': [m.addVar(vtype=GRB.CONTINUOUS, name=f"Tax_{i+len(input_biofuel)}") for i in range(6)],
            'P_all': [m.addVar(vtype=GRB.CONTINUOUS, name=f"P_all_{i+len(input_biofuel)}") for i in range(6)],
            'PC_F0': [m.addVar(vtype=GRB.CONTINUOUS, name=f"PC_F0_{i+len(input_biofuel)}") for i in range(6)],
            'P_RIN_SAF_LCFS_detach': [m.addVar(vtype=GRB.CONTINUOUS, name=f"P_RIN_SAF_LCFS_detach_{i+len(input_biofuel)}") for i in range(6)],
            'P_RIN_SAF_detach': [m.addVar(vtype=GRB.CONTINUOUS, name=f"P_RIN_SAF_detach_{i+len(input_biofuel)}") for i in range(6)],
            'Q_actual': [m.addVar(vtype=GRB.CONTINUOUS, name=f"Q_actual_{i+len(input_biofuel)}") for i in range(6)],
            'Q_equiv': [m.addVar(vtype=GRB.CONTINUOUS, name=f"Q_equiv_{i+len(input_biofuel)}") for i in range(6)]
        }, index=conventionalfuel_quantity.index)

        # Combine the biofuel and conventional fuel price DataFrames
        Price = pd.concat([Price_bio, Price_conventional])

            # Variables
        # Replace these lines:
        # statetaxzeta_value = m.addVars(len(vals), vtype=GRB.BINARY, name="statetaxzeta_value")
        # Create multipliers nu
        # diracdelta_value = m.addVar(vtype=GRB.CONTINUOUS, name="diracdelta_value")
        # diraczeta__value = m.addVar(vtype=GRB.BINARY, name="diraczeta_value")
        diracdeltaub_value = m.addVar(vtype=GRB.CONTINUOUS, name="diracdeltaub_value")
        diraczetaub_value = m.addVar(vtype=GRB.BINARY, name="diraczetaub_value")
        # diracdeltalb_value = m.addVar(vtype=GRB.CONTINUOUS, name="diracdeltalb_value")
        # diraczetalb_value = m.addVar(vtype=GRB.BINARY, name="diraczetalb_value")
        
        # Biofuel supply curves
        output = biofuel_quantity.join(input_biofuel)
        biofuel_quantity = biofuel_quantity.join(input_biofuel)

        # Define an objective
        # We do not actually need an objective function

        # # Define the objective function as the sum of profit times quantity
        # objectivefun = gp.quicksum(
        #     (
        #         biofuel_prices.loc[(biofuel_prices['f'] == output['f'][i]) & (biofuel_prices['j'] == output['j'][i]), 'P'].item()
        #         - (output['PC'][i] + output['Conversion'][i] * feedstock_prices.loc[feedstock_prices['s'] == output['s'][i], 'P'].item())
        #         + output['kappa_RFS'][i] * RIN_prices.loc[RIN_prices['fueltype'].str.contains(output['f'][i]), 'P'].item()
        #         + 1 * ((output['f'][i] == 'B100') | (output['f'][i] == 'RD') | (output['f'][i] == 'gasE100')) * ((47.39 - output['CI_Tax'][i]) / 47.39) * output['Taxheavyside'][i]
        #         + max([
        #             1 * (output['f'][i] == 'SAF') * ((47.39 - output['CI_Tax'][i]) / 47.39) * output['Taxheavyside'][i],
        #             85 * 30 * output['ED_LCFS'][i] / (10**6) * output['CCS_tech'][i]
        #         ])
        #         #+ (output['f'][i] == 'SAF') * SAFcredit_prices['P'].item()
        #         + LCFS_prices['P'][0] * (output['CI_std_LCFS'][i] - output['CI_LCFS'][i]) * output['ED_LCFS'][i] * (output['j'][i] == 'CA') / 1000000
        #         + ((output['f'][i] == 'SAF') & (output['j'][i] == 'NC')) * output['Taxheavyside'][i] * val
        #     ) * biofuel_quantity_Q[i]
        #     for i in range(24)
        # )

        # m.setObjective(objectivefun, GRB.MAXIMIZE)
        m.setObjective(0, GRB.MAXIMIZE)

        # Define constraints
        Constraints = []

        for i in range(24):
            Constraints.append(
                biofuel_prices.loc[(biofuel_prices['f'] == output['f'][i]) & (biofuel_prices['j'] == output['j'][i]), 'P'].item() -
                (output['PC'][i] + output['Conversion'][i] * feedstock_prices.loc[feedstock_prices['s'] == output['s'][i], 'P'].item()*
                 (1+duty_e100sugarcane*(output['s'][i] == 'sugarcane'))) +
                output['kappa_RFS'][i] * RIN_prices.loc[RIN_prices['code'] == 'D4', 'P'].item() * output['D4'][i] +
                output['kappa_RFS'][i] * RIN_prices.loc[RIN_prices['code'] == 'D6', 'P'].item() * output['D6'][i] +
                1 * ((output['f'][i] == 'B100') | (output['f'][i] == 'RD') | (output['f'][i] == 'gasE100')) *
                ((47.39 - output['CI_Tax'][i]) / 47.39) * output['Taxheavyside'][i] * (output['s'][i] != 'sugarcane')
                #+ (output['f'][i] == 'SAF') * SAFcredit_prices['P'].item() 
                +max([1 * (output['f'][i] == 'SAF') * ((47.39 - output['CI_Tax'][i]) / 47.39) * output['Taxheavyside'][i],
                    85 * 30 * output['ED_LCFS'][i] / (10**6) * output['CCS_tech'][i]])* (output['s'][i] != 'sugarcane') +
                    (LCFS_prices.loc['LCFS', 'P'] * output.iloc[i]['CALCFSSide'] * (output.iloc[i]['CI_std_LCFS_CA'] - output.iloc[i]['CI_LCFS']) * output.iloc[i]['ED_LCFS'] / 1000000) +
                    (LCFS_prices_SAF.loc['LCFS_SAF', 'P'] * output.iloc[i]['SAFLCFSside'] * (output.iloc[i]['CI_std_LCFS'] - output.iloc[i]['CI_LCFS']) * output.iloc[i]['ED_LCFS'] / 1000000) +
                ((output['f'][i] == 'SAF') & (output['j'][i] == 'NC')) * output['Statetaxheaviside'][i]* val + 
                mu['value'][i] -
                nu.loc[(nu['f'] == output['f'][i]) & (nu['j'] == output['j'][i]), 'value'].item()   - 
                    ((output['f'][i] == 'SAF') & (output['j'][i] == 'NC')) * output['Statetaxheaviside'][i]* diracdeltaub_value  -
                    hefa_value[0] * ((output['f'][i] == 'SAF') & (output['s'][i] == 'soyoil')) +
                    hefa_value[0] * 5.5/2.5 * ((output['f'][i] == 'RD') & (output['s'][i] == 'soyoil')) 
                    - hefa_value[1] * ((output['f'][i] == 'SAF') & (output['s'][i] == 'animalfat')) +
                    hefa_value[1] * 5.5/2.5 * ((output['f'][i] == 'RD') & (output['s'][i] == 'animalfat'))
                   == 0           
                ) 
            
        # Biofuel supply curves - multipliers mu
        M4 = 15
        for j in range(24):
            Constraints.append(mu_value[j] >= 0)
            Constraints.append(mu_value[j] <= muzeta_value[j] * M4)
            Constraints.append(output['Q'][j] >= 0)
            Constraints.append(output['Q'][j] <= (1 - muzeta_value[j]) * M4)
            Constraints.append(output['Q'][j] >= epsilon * (1 - muzeta_value[j]))
            Constraints.append(mu_value[j] >= epsilon *muzeta_value[j])

        # Biofuel supply curves - multipliers nu
        # Blending wall constraints
        M5 = 100
        fuel_types = ['B100', 'RD']
        states = ['CA', 'NC']
        for fuel in fuel_types:
            for state in states:
                Constraints.append(nu_value[nu_index.index(f'{fuel}_{state}')] >= 0)
                Constraints.append(nu_value[nu_index.index(f'{fuel}_{state}')] <= (1 - nuzeta_value[nu_index.index(f'{fuel}_{state}')]) * M5)
                Constraints.append(-(biofuel_quantity.loc[(biofuel_quantity['f'] == fuel) & (biofuel_quantity['j'] == state), 'Q'].sum() -
                                        biofuel_quantity.loc[(biofuel_quantity['f'] == fuel) & (biofuel_quantity['j'] == state), 'BW_constraints'].unique()[0] *
                                        demand_quantity.loc[(demand_quantity['F'] == 'D') & (demand_quantity['j'] == state), 'Q_actual'].item()) >= 0)
                Constraints.append(-(biofuel_quantity.loc[(biofuel_quantity['f'] == fuel) & (biofuel_quantity['j'] == state), 'Q'].sum() -
                                    biofuel_quantity.loc[(biofuel_quantity['f'] == fuel) & (biofuel_quantity['j'] == state), 'BW_constraints'].unique()[0] *
                                    demand_quantity.loc[(demand_quantity['F'] == 'D') & (demand_quantity['j'] == state), 'Q_actual'].item()) <=
                                    nuzeta_value[nu_index.index(f'{fuel}_{state}')] * M5)

        M6 = 200        
        fuel_types = ['gasE100']
        states = ['CA', 'NC']
        for fuel in fuel_types:
            for state in states:
                Constraints.append(nu_value[nu_index.index(f'{fuel}_{state}')] >= 0)
                Constraints.append(nu_value[nu_index.index(f'{fuel}_{state}')] <= (1 - nuzeta_value[nu_index.index(f'{fuel}_{state}')]) * M6)
                Constraints.append(-(biofuel_quantity.loc[(biofuel_quantity['f'] == fuel) & (biofuel_quantity['j'] == state), 'Q'].sum() -
                                        biofuel_quantity.loc[(biofuel_quantity['f'] == fuel) & (biofuel_quantity['j'] == state), 'BW_constraints'].unique()[0] *
                                        demand_quantity.loc[(demand_quantity['F'] == 'G') & (demand_quantity['j'] == state), 'Q_actual'].item()) >= 0)
                Constraints.append(-(biofuel_quantity.loc[(biofuel_quantity['f'] == fuel) & (biofuel_quantity['j'] == state), 'Q'].sum() -
                                    biofuel_quantity.loc[(biofuel_quantity['f'] == fuel) & (biofuel_quantity['j'] == state), 'BW_constraints'].unique()[0] *
                                    demand_quantity.loc[(demand_quantity['F'] == 'G') & (demand_quantity['j'] == state), 'Q_actual'].item()) <=
                                    nuzeta_value[nu_index.index(f'{fuel}_{state}')] * M6)
                
        M7 = 25
        fuel_types = ['SAF']
        states = ['CA', 'NC']
        for fuel in fuel_types:
            for state in states:
                Constraints.append(nu_value[nu_index.index(f'{fuel}_{state}')] >= 0)
                Constraints.append(nu_value[nu_index.index(f'{fuel}_{state}')] <= (1 - nuzeta_value[nu_index.index(f'{fuel}_{state}')]) * M7)
                Constraints.append(-(biofuel_quantity.loc[(biofuel_quantity['f'] == fuel) & (biofuel_quantity['j'] == state), 'Q'].sum() -
                                        biofuel_quantity.loc[(biofuel_quantity['f'] == fuel) & (biofuel_quantity['j'] == state), 'BW_constraints'].unique()[0] *
                                        demand_quantity.loc[(demand_quantity['F'] == 'J') & (demand_quantity['j'] == state), 'Q_actual'].item()) >= 0)
                Constraints.append(-(biofuel_quantity.loc[(biofuel_quantity['f'] == fuel) & (biofuel_quantity['j'] == state), 'Q'].sum() -
                                    biofuel_quantity.loc[(biofuel_quantity['f'] == fuel) & (biofuel_quantity['j'] == state), 'BW_constraints'].unique()[0] *
                                    demand_quantity.loc[(demand_quantity['F'] == 'J') & (demand_quantity['j'] == state), 'Q_actual'].item()) <=
                                    nuzeta_value[nu_index.index(f'{fuel}_{state}')] * M7)       
                
        # # # HEFA constraints
        M1_0 = 10
        Constraints.append(hefa_value[0] >= 0)
        Constraints.append(hefa_value[0] <= hefazeta_value[0] * M1_0)
        Constraints.append(biofuel_quantity.loc[(biofuel_quantity['f'] == 'SAF') & (biofuel_quantity['s'] == 'soyoil'), 'Q'].sum() -
               5.5/2.5*biofuel_quantity.loc[(biofuel_quantity['f'] == 'RD') & (biofuel_quantity['s'] == 'soyoil'), 'Q'].sum() <=0)
        Constraints.append(biofuel_quantity.loc[(biofuel_quantity['f'] == 'SAF') & (biofuel_quantity['s'] == 'soyoil'), 'Q'].sum() -
               5.5/2.5*biofuel_quantity.loc[(biofuel_quantity['f'] == 'RD') & (biofuel_quantity['s'] == 'soyoil'), 'Q'].sum() >= - (1 - hefazeta_value[0]) * M1_0)
        Constraints.append(biofuel_quantity.loc[(biofuel_quantity['f'] == 'SAF') & (biofuel_quantity['s'] == 'soyoil'), 'Q'].sum() -
               5.5/2.5*biofuel_quantity.loc[(biofuel_quantity['f'] == 'RD') & (biofuel_quantity['s'] == 'soyoil'), 'Q'].sum() <= (1 - hefazeta_value[0]) * M1_0)
        
        M1_1 = 10
        Constraints.append(hefa_value[1] >= 0)
        Constraints.append(hefa_value[1] <= hefazeta_value[1] * M1_1)
        Constraints.append(biofuel_quantity.loc[(biofuel_quantity['f'] == 'SAF') & (biofuel_quantity['s'] == 'animalfat'), 'Q'].sum() -
               5.5/2.5*biofuel_quantity.loc[(biofuel_quantity['f'] == 'RD') & (biofuel_quantity['s'] == 'animalfat'), 'Q'].sum() <= 0)
        Constraints.append(biofuel_quantity.loc[(biofuel_quantity['f'] == 'SAF') & (biofuel_quantity['s'] == 'animalfat'), 'Q'].sum() -
               5.5/2.5*biofuel_quantity.loc[(biofuel_quantity['f'] == 'RD') & (biofuel_quantity['s'] == 'animalfat'), 'Q'].sum() >= - (1 - hefazeta_value[1]) * M1_1)
        Constraints.append(biofuel_quantity.loc[(biofuel_quantity['f'] == 'SAF') & (biofuel_quantity['s'] == 'animalfat'), 'Q'].sum() -
               5.5/2.5*biofuel_quantity.loc[(biofuel_quantity['f'] == 'RD') & (biofuel_quantity['s'] == 'animalfat'), 'Q'].sum() <= (1 - hefazeta_value[1]) * M1_1)

        # M8 = 1 
        # Constraints.append(diracdeltaub_value >= 0)
        # Constraints.append(diracdeltaub_value <= (1-diraczetaub_value) * M8)       
        # Constraints.append(-(biofuel_quantity.loc[(biofuel_quantity['f'] == 'SAF') & (biofuel_quantity['j'] == 'NC') & (biofuel_quantity['Statetaxheaviside'] == 1), 'Q'].sum() -
        #         break_val) >= 0)         
        # Constraints.append(-(biofuel_quantity.loc[(biofuel_quantity['f'] == 'SAF') & (biofuel_quantity['j'] == 'NC') & (biofuel_quantity['Statetaxheaviside'] == 1), 'Q'].sum() -
        #         break_val) <= diraczetaub_value * M8)   
        
        M8 = 1 
        if break_idx < len(breaks) - 1:  # Only add constraints for non-last breaks
            Constraints.append(diracdeltaub_value >= 0)
            Constraints.append(diracdeltaub_value <= (1-diraczetaub_value) * M8)       
            Constraints.append(-(biofuel_quantity.loc[(biofuel_quantity['f'] == 'SAF') & (biofuel_quantity['j'] == 'NC') & (biofuel_quantity['Statetaxheaviside'] == 1), 'Q'].sum() -
                    break_val) >= 0)         
            Constraints.append(-(biofuel_quantity.loc[(biofuel_quantity['f'] == 'SAF') & (biofuel_quantity['j'] == 'NC') & (biofuel_quantity['Statetaxheaviside'] == 1), 'Q'].sum() -
                    break_val) <= diraczetaub_value * M8)
        else:  # For the last break, force diracdeltaub_value = 0
            Constraints.append(diracdeltaub_value == 0)   
        
        M8_lb = 2 
        epsilon1=1e-4
        # Constraints.append(diracdeltalb_value >= 0)
        # Constraints.append(diracdeltalb_value <= (1-diraczetalb_value) * M8_lb)
        Constraints.append((biofuel_quantity.loc[(biofuel_quantity['f'] == 'SAF') & (biofuel_quantity['j'] == 'NC') & (biofuel_quantity['Statetaxheaviside'] == 1), 'Q'].sum() - epsilon1 -
                break_val_pre) >= 0)         
        # Constraints.append(biofuel_quantity.loc[(biofuel_quantity['f'] == 'SAF') & (biofuel_quantity['j'] == 'NC'), 'Q'].sum() -
        #         break_val_pre - epsilon1 <= diraczetalb_value * M8_lb)      
 
            
        # Biofuel quantity - get diesel equivalent and gasoline equivalent
        for i in range(24):
            Constraints.append(biofuel_quantity_Q_equiv[i] - biofuel_quantity_Q[i] * output['phi_ed'][i] == 0)
        
        # # Constraint: Disallow sugarcane as feedstock - set all sugarcane-based biofuel quantities to zero
        # for i in range(24):
        #     if output['s'][i] == 'sugarcane':
        #         Constraints.append(biofuel_quantity_Q[i] == 0)
        
        # # Constraint: Set sugarcane feedstock quantity to zero (no sugarcane allowed)
        # Constraints.append(feedstock_quantity.loc[feedstock_quantity['s'] == 'sugarcane', 'Q'].item() == 0)


        # # Calculate denominator variable as specified:
        # # sum of all diesel quantity divided by 0.75 minus biofuel diesel quantity and plus gasoline quantity
        # denominator = m.addVar(vtype=GRB.CONTINUOUS, name="denominator")
        
        # # Define the denominator expression
        # # Total diesel quantities (conventional + biofuel diesel)
        # total_diesel_qty = (
        #     # Conventional diesel quantities
        #     conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'D') & (conventionalfuel_quantity['j'] == 'CA'), 'Q'].item() +
        #     conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'D') & (conventionalfuel_quantity['j'] == 'NC'), 'Q'].item() +
        #     # Biofuel diesel quantities (B100 and RD)
        #     sum(biofuel_quantity.loc[biofuel_quantity['f'].isin(['B100', 'RD']), 'Q'])
        # )
        
        # # Biofuel diesel quantities (B100 and RD)
        # biofuel_diesel_qty = sum(biofuel_quantity.loc[biofuel_quantity['f'].isin(['B100', 'RD']), 'Q'])
        
        # # Total gasoline quantities (conventional + biofuel gasoline)
        # total_gasoline_qty = (
        #     # Conventional gasoline quantities
        #     conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'G') & (conventionalfuel_quantity['j'] == 'CA'), 'Q'].item() +
        #     conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'G') & (conventionalfuel_quantity['j'] == 'NC'), 'Q'].item() +
        #     # Biofuel gasoline quantities (gasE100)
        #     sum(biofuel_quantity.loc[biofuel_quantity['f'] == 'gasE100', 'Q'])
        # )

        # # Biofuel diesel quantities (B100 and RD)
        # biofuel_gasoline_qty = sum(biofuel_quantity.loc[biofuel_quantity['f'].isin(['gasE100']), 'Q'])

        
        # # Set the denominator constraint
        # m.addConstr(denominator == (total_diesel_qty / 0.75) - biofuel_diesel_qty + (total_gasoline_qty/0.96)-biofuel_gasoline_qty, name="denominator_constraint")    

        # RVO percent
        # Constraints.append(
        #     RVO_percent.loc['BBD', 'P'] == 
        #     input_compliance.loc[input_compliance['mandate'] == 'BBD', 'Q'].item() * 
        #     (conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'G') | (conventionalfuel_quantity['F'] == 'D'), 'Q'].sum() / 
        #     conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'G') | (conventionalfuel_quantity['F'] == 'D'), 'Q'].sum())
        # )

        Constraints.append(
            RVO_percent.loc['BBD', 'P'] == 
            biofuel_quantity.loc[biofuel_quantity['D4'] == 1, 'Q'].sum() * 1.6 / 
            conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'G') | (conventionalfuel_quantity['F'] == 'D'), 'Q'].sum()
        )

        # Constraints.append(
        #     RVO_percent.loc['RF', 'P'] == 
        #     input_compliance.loc[input_compliance['mandate'] == 'RF', 'Q'].item() * 
        #     (conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'G') | (conventionalfuel_quantity['F'] == 'D'), 'Q'].sum() / 
        #     conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'G') | (conventionalfuel_quantity['F'] == 'D'), 'Q'].sum())
        # )

        Constraints.append(
            RVO_percent.loc['RF', 'P'] == 
            (sum((biofuel_quantity.loc[biofuel_quantity['f'] == 'B100', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'B100', 'Q'])) + sum(biofuel_quantity.loc[biofuel_quantity['f'] == 'RD', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'RD', 'Q']) + sum(biofuel_quantity.loc[biofuel_quantity['f'] == 'SAF', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'SAF', 'Q']) + sum(biofuel_quantity.loc[biofuel_quantity['f'] == 'gasE100', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'gasE100', 'Q'])) / 
            conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'G') | (conventionalfuel_quantity['F'] == 'D'), 'Q'].sum()
        )


        # Conventional fuel prices
        Fconventionalfuel_pricesD_NC = (conventionalfuel_prices.loc[(conventionalfuel_prices['F'] == 'D') & (conventionalfuel_prices['j'] == 'NC'), 'P'].item() - 
                                        RVO_percent.loc['BBD', 'P'] * RIN_prices.loc['D4', 'P'] - 
                                        (RVO_percent.loc['RF', 'P'] - RVO_percent.loc['BBD', 'P']) * RIN_prices.loc['D6', 'P'] - 
                                        input_demand.loc[input_demand['F'] == 'D', 'PC_F0'].unique()[0])
        Fconventionalfuel_pricesG_NC = (conventionalfuel_prices.loc[(conventionalfuel_prices['F'] == 'G') & (conventionalfuel_prices['j'] == 'NC'), 'P'].item() - 
                                        RVO_percent.loc['BBD', 'P'] * RIN_prices.loc['D4', 'P'] - 
                                        (RVO_percent.loc['RF', 'P'] - RVO_percent.loc['BBD', 'P']) * RIN_prices.loc['D6', 'P'] - 
                                        input_demand.loc[input_demand['F'] == 'G', 'PC_F0'].unique()[0])
        Fconventionalfuel_pricesJ_NC = (conventionalfuel_prices.loc[(conventionalfuel_prices['F'] == 'J') & (conventionalfuel_prices['j'] == 'NC'), 'P'].item() - (LCFS_prices_SAF.loc['LCFS_SAF', 'P'] * 
                (conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'J', 'CI_LCFS'].unique()[0] - 
                SAF_intensity_standard_P) * 
                conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'J', 'ED_LCFS'].unique()[0] / 1000000) - 
                                        input_demand.loc[input_demand['F'] == 'J', 'PC_F0'].unique()[0])

        Fconventionalfuel_pricesD_CA = (conventionalfuel_prices.loc[(conventionalfuel_prices['F'] == 'D') & (conventionalfuel_prices['j'] == 'CA'), 'P'].item() - 
                                        RVO_percent.loc['BBD', 'P'] * RIN_prices.loc['D4', 'P'] - 
                                        (RVO_percent.loc['RF', 'P'] - RVO_percent.loc['BBD', 'P']) * RIN_prices.loc['D6', 'P'] - 
                                        LCFS_prices.loc['LCFS', 'P'] * 
                                        (-conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'D'), 'CI_std_LCFS_CA'].item() + 
                                        conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'D'), 'CI_LCFS'].item()) * 
                                        conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'D'), 'ED_LCFS'].item() / 1000000  - 
                                        input_demand.loc[input_demand['F'] == 'D', 'PC_F0'].unique()[0])
        Fconventionalfuel_pricesG_CA = (conventionalfuel_prices.loc[(conventionalfuel_prices['F'] == 'G') & (conventionalfuel_prices['j'] == 'CA'), 'P'].item() - 
                                        RVO_percent.loc['BBD', 'P'] * RIN_prices.loc['D4', 'P'] - 
                                        (RVO_percent.loc['RF', 'P'] - RVO_percent.loc['BBD', 'P']) * RIN_prices.loc['D6', 'P'] - 
                                        LCFS_prices.loc['LCFS', 'P'] * 
                                        (-conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'G'), 'CI_std_LCFS_CA'].item() + 
                                        conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'G'), 'CI_LCFS'].item()) * 
                                        conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'G'), 'ED_LCFS'].item() / 1000000 - 
                                        input_demand.loc[input_demand['F'] == 'G', 'PC_F0'].unique()[0])
        Fconventionalfuel_pricesJ_CA = (conventionalfuel_prices.loc[(conventionalfuel_prices['F'] == 'J') & (conventionalfuel_prices['j'] == 'CA'), 'P'].item() - (LCFS_prices_SAF.loc['LCFS_SAF', 'P'] * 
                (conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'J', 'CI_LCFS'].unique()[0] - 
                SAF_intensity_standard_P) * 
                conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'J', 'ED_LCFS'].unique()[0] / 1000000) - 
                                        input_demand.loc[input_demand['F'] == 'J', 'PC_F0'].unique()[0])

        Constraints.append(Fconventionalfuel_pricesD_CA == 0)
        Constraints.append(Fconventionalfuel_pricesG_CA == 0)
        Constraints.append(Fconventionalfuel_pricesJ_CA == 0)
        Constraints.append(Fconventionalfuel_pricesD_NC == 0)
        Constraints.append(Fconventionalfuel_pricesG_NC == 0)
        Constraints.append(Fconventionalfuel_pricesJ_NC == 0)


        # # Biofuel demand
        # FbiofueldemandB100CA = (input_biofuel.loc[input_biofuel['f'] == 'B100', 'phi'].unique()[0] * 
        #                         input_demand.loc[input_demand['F'] == 'D', 'PC_F0'].unique()[0] - 
        #                         biofuel_prices.loc[(biofuel_prices['f'] == 'B100') & (biofuel_prices['j'] == 'CA'), 'P'].item())
        # FbiofueldemandRDCA = (input_biofuel.loc[input_biofuel['f'] == 'RD', 'phi'].unique()[0] * 
        #                     input_demand.loc[input_demand['F'] == 'D', 'PC_F0'].unique()[0] - 
        #                     biofuel_prices.loc[(biofuel_prices['f'] == 'RD') & (biofuel_prices['j'] == 'CA'), 'P'].item())
        # FbiofueldemandSAFCA = (input_biofuel.loc[input_biofuel['f'] == 'SAF', 'phi'].unique()[0] * 
        #                     input_demand.loc[input_demand['F'] == 'J', 'PC_F0'].unique()[0] - 
        #                     biofuel_prices.loc[(biofuel_prices['f'] == 'SAF') & (biofuel_prices['j'] == 'CA'), 'P'].item())

        # FbiofueldemandB100NC = (input_biofuel.loc[input_biofuel['f'] == 'B100', 'phi'].unique()[0] * 
        #                         input_demand.loc[input_demand['F'] == 'D', 'PC_F0'].unique()[0] - 
        #                         biofuel_prices.loc[(biofuel_prices['f'] == 'B100') & (biofuel_prices['j'] == 'NC'), 'P'].item())
        # FbiofueldemandRDNC = (input_biofuel.loc[input_biofuel['f'] == 'RD', 'phi'].unique()[0] * 
        #                     input_demand.loc[input_demand['F'] == 'D', 'PC_F0'].unique()[0] - 
        #                     biofuel_prices.loc[(biofuel_prices['f'] == 'RD') & (biofuel_prices['j'] == 'NC'), 'P'].item())
        # FbiofueldemandSAFNC = (input_biofuel.loc[input_biofuel['f'] == 'SAF', 'phi'].unique()[0] * 
        #                     input_demand.loc[input_demand['F'] == 'J', 'PC_F0'].unique()[0] - 
        #                     biofuel_prices.loc[(biofuel_prices['f'] == 'SAF') & (biofuel_prices['j'] == 'NC'), 'P'].item())

        FbiofueldemandB100CA = (input_biofuel.loc[input_biofuel['f'] == 'B100', 'phi'].unique()[0] * 
                                conventionalfuel_prices.loc[(conventionalfuel_prices['F'] == 'D') & (conventionalfuel_prices['j'] == 'CA'), 'P'].item() - 
                                biofuel_prices.loc[(biofuel_prices['f'] == 'B100') & (biofuel_prices['j'] == 'CA'), 'P'].item())
        FbiofueldemandRDCA = (input_biofuel.loc[input_biofuel['f'] == 'RD', 'phi'].unique()[0] * 
                            conventionalfuel_prices.loc[(conventionalfuel_prices['F'] == 'D') & (conventionalfuel_prices['j'] == 'CA'), 'P'].item() - 
                            biofuel_prices.loc[(biofuel_prices['f'] == 'RD') & (biofuel_prices['j'] == 'CA'), 'P'].item())
        FbiofueldemandSAFCA = (input_biofuel.loc[input_biofuel['f'] == 'SAF', 'phi'].unique()[0] * 
                            conventionalfuel_prices.loc[(conventionalfuel_prices['F'] == 'J') & (conventionalfuel_prices['j'] == 'CA'), 'P'].item() - 
                            biofuel_prices.loc[(biofuel_prices['f'] == 'SAF') & (biofuel_prices['j'] == 'CA'), 'P'].item())

        FbiofueldemandB100NC = (input_biofuel.loc[input_biofuel['f'] == 'B100', 'phi'].unique()[0] * 
                                conventionalfuel_prices.loc[(conventionalfuel_prices['F'] == 'D') & (conventionalfuel_prices['j'] == 'NC'), 'P'].item() - 
                                biofuel_prices.loc[(biofuel_prices['f'] == 'B100') & (biofuel_prices['j'] == 'NC'), 'P'].item())
        FbiofueldemandRDNC = (input_biofuel.loc[input_biofuel['f'] == 'RD', 'phi'].unique()[0] * 
                            conventionalfuel_prices.loc[(conventionalfuel_prices['F'] == 'D') & (conventionalfuel_prices['j'] == 'NC'), 'P'].item() - 
                            biofuel_prices.loc[(biofuel_prices['f'] == 'RD') & (biofuel_prices['j'] == 'NC'), 'P'].item())
        FbiofueldemandSAFNC = (input_biofuel.loc[input_biofuel['f'] == 'SAF', 'phi'].unique()[0] * 
                            conventionalfuel_prices.loc[(conventionalfuel_prices['F'] == 'J') & (conventionalfuel_prices['j'] == 'NC'), 'P'].item() - 
                            biofuel_prices.loc[(biofuel_prices['f'] == 'SAF') & (biofuel_prices['j'] == 'NC'), 'P'].item())

        Constraints.append(FbiofueldemandB100CA == 0)
        Constraints.append(FbiofueldemandRDCA == 0)
        Constraints.append(FbiofueldemandSAFCA == 0)
        Constraints.append(FbiofueldemandB100NC == 0)
        Constraints.append(FbiofueldemandRDNC == 0)
        Constraints.append(FbiofueldemandSAFNC == 0)


        # Biofuel demand for gasE100 in NC
        blending_ratio_NC = m.addVar(vtype=GRB.CONTINUOUS, name="blending_ratio_NC")
        m.addConstr(blending_ratio_NC == sum(biofuel_quantity.loc[(biofuel_quantity['f'] == 'gasE100') & (biofuel_quantity['j'] == 'NC'), 'Q']) / demand_quantity.loc[(demand_quantity['F'] == 'G') & (demand_quantity['j'] == 'NC'), 'Q_actual'].item())

        indicator_var_NC = m.addVar(vtype=GRB.BINARY, name="indicator_var_NC")
        M9 = 1
        m.addConstr(blending_ratio_NC >= 0.1 + Sigma - M9 * (1 - indicator_var_NC), name="indicator_var_NC_constr1")
        m.addConstr(blending_ratio_NC <= 0.1 + M9 * indicator_var_NC, name="indicator_var_NC_constr2")

        m.addConstr(blending_ratio_NC >= 0.1)

        # Modify the constraint to include the indicator variable
        Constraints.append((input_biofuel.loc[input_biofuel['f'] == 'gasE100', 'phi'].unique()[0] * 
                            conventionalfuel_prices.loc[(conventionalfuel_prices['F'] == 'G') & (conventionalfuel_prices['j'] == 'NC'), 'P'].item() * 
                            (1 - indicator_var_NC * (nlfunc.exp(((sum(biofuel_quantity.loc[(biofuel_quantity['f'] == 'gasE100') & (biofuel_quantity['j'] == 'NC'), 'Q']) / 
                                                demand_quantity.loc[(demand_quantity['F'] == 'G') & (demand_quantity['j'] == 'NC'), 'Q_actual'].item() - 0.1) / 0.02) * 
                                                np.log(2)) - 1)) == 
                            biofuel_prices.loc[(biofuel_prices['f'] == 'gasE100') & (biofuel_prices['j'] == 'NC'), 'P'].item()))

        # Biofuel demand for gasE100 in CA
        blending_ratio_CA = m.addVar(vtype=GRB.CONTINUOUS, name="blending_ratio_CA")
        m.addConstr(blending_ratio_CA == sum(biofuel_quantity.loc[(biofuel_quantity['f'] == 'gasE100') & (biofuel_quantity['j'] == 'CA'), 'Q']) / demand_quantity.loc[(demand_quantity['F'] == 'G') & (demand_quantity['j'] == 'CA'), 'Q_actual'].item())

        indicator_var_CA = m.addVar(vtype=GRB.BINARY, name="indicator_var_CA")
        M10 = 1
        m.addConstr(blending_ratio_CA >= 0.1 + Sigma - M10 * (1 - indicator_var_CA), name="indicator_var_CA_constr1")
        m.addConstr(blending_ratio_CA <= 0.1 + M10 * indicator_var_CA, name="indicator_var_CA_constr2")

        m.addConstr(blending_ratio_CA >= 0.1)

        # Modify the constraint to include the indicator variable
        Constraints.append((input_biofuel.loc[input_biofuel['f'] == 'gasE100', 'phi'].unique()[0] * 
                            conventionalfuel_prices.loc[(conventionalfuel_prices['F'] == 'G') & (conventionalfuel_prices['j'] == 'CA'), 'P'].item() * 
                            (1 - indicator_var_CA * (nlfunc.exp(((sum(biofuel_quantity.loc[(biofuel_quantity['f'] == 'gasE100') & (biofuel_quantity['j'] == 'CA'), 'Q']) / 
                                                demand_quantity.loc[(demand_quantity['F'] == 'G') & (demand_quantity['j'] == 'CA'), 'Q_actual'].item() - 0.1) / 0.02) * 
                                                np.log(2)) - 1)) == 
                            biofuel_prices.loc[(biofuel_prices['f'] == 'gasE100') & (biofuel_prices['j'] == 'CA'), 'P'].item()))

        # Add constraints for blend prices
        Constraints.append(sum(Price.loc[(Price['F'] == 'D') & (Price['j'] == 'CA'), 'P_RIN_SAF_LCFS_detach'] * 
                            Price.loc[(Price['F'] == 'D') & (Price['j'] == 'CA'), 'Q_actual']) / 
                            (demand_quantity.loc[(demand_quantity['F'] == 'D') & (demand_quantity['j'] == 'CA'), 'Q'].item()) == 
                            demand_price.loc[(demand_price['F'] == 'D') & (demand_price['j'] == 'CA'), 'P'].item())
        Constraints.append(sum(Price.loc[(Price['F'] == 'D') & (Price['j'] == 'NC'), 'P_RIN_SAF_LCFS_detach'] * 
                            Price.loc[(Price['F'] == 'D') & (Price['j'] == 'NC'), 'Q_actual'])/
                            (demand_quantity.loc[(demand_quantity['F'] == 'D') & (demand_quantity['j'] == 'NC'), 'Q'].item()) == 
                            demand_price.loc[(demand_price['F'] == 'D') & (demand_price['j'] == 'NC'), 'P'].item())
        Constraints.append(sum(Price.loc[(Price['F'] == 'G') & (Price['j'] == 'CA'), 'P_RIN_SAF_LCFS_detach'] * 
                            Price.loc[(Price['F'] == 'G') & (Price['j'] == 'CA'), 'Q_actual'])/
                            (demand_quantity.loc[(demand_quantity['F'] == 'G') & (demand_quantity['j'] == 'CA'), 'Q'].item()) == 
                            demand_price.loc[(demand_price['F'] == 'G') & (demand_price['j'] == 'CA'), 'P'].item())
        Constraints.append(sum(Price.loc[(Price['F'] == 'G') & (Price['j'] == 'NC'), 'P_RIN_SAF_LCFS_detach'] * 
                            Price.loc[(Price['F'] == 'G') & (Price['j'] == 'NC'), 'Q_actual'])/
                            (demand_quantity.loc[(demand_quantity['F'] == 'G') & (demand_quantity['j'] == 'NC'), 'Q'].item()) == 
                            demand_price.loc[(demand_price['F'] == 'G') & (demand_price['j'] == 'NC'), 'P'].item())
        Constraints.append(sum(Price.loc[(Price['F'] == 'J') & (Price['j'] == 'CA'), 'P_RIN_SAF_LCFS_detach'] * 
                            Price.loc[(Price['F'] == 'J') & (Price['j'] == 'CA'), 'Q_actual'])/
                            (demand_quantity.loc[(demand_quantity['F'] == 'J') & (demand_quantity['j'] == 'CA'), 'Q'].item()) == 
                            demand_price.loc[(demand_price['F'] == 'J') & (demand_price['j'] == 'CA'), 'P'].item())
        Constraints.append(sum(Price.loc[(Price['F'] == 'J') & (Price['j'] == 'NC'), 'P_RIN_SAF_LCFS_detach'] * 
                            Price.loc[(Price['F'] == 'J') & (Price['j'] == 'NC'), 'Q_actual'])/
                            (demand_quantity.loc[(demand_quantity['F'] == 'J') & (demand_quantity['j'] == 'NC'), 'Q'].item()) == 
                            demand_price.loc[(demand_price['F'] == 'J') & (demand_price['j'] == 'NC'), 'P'].item())

        # Total demand elasticity
        for region in ['CA','NC']:
            for fuel in ['D','G','J']:
                alpha = input_demand.loc[(input_demand['F'] == fuel) & (input_demand['j'] == region), 'alpha'].item()
                beta = input_demand.loc[(input_demand['F'] == fuel) & (input_demand['j'] == region), 'beta'].item()
                price = demand_price.loc[(demand_price['F'] == fuel) & (demand_price['j'] == region), 'P'].item()
                quantity = demand_quantity.loc[(demand_quantity['F'] == fuel) & (demand_quantity['j'] == region), 'Q'].item()
                
                # Create an auxiliary variable for the power term
                power_term = m.addVar(vtype=GRB.CONTINUOUS, name=f"power_term_{fuel}_{region}")
                
                # Add the power constraint
                m.addGenConstrPow(price, power_term, beta, name=f"power_constraint_{fuel}_{region}")
                
                # Add the elasticity constraint
                Constraints.append(alpha * power_term == quantity)

        # Market clearing - demand equation
        Ddemand_CA = (
            sum(biofuel_quantity.loc[(biofuel_quantity['j'] == 'CA') & (biofuel_quantity['F'] == 'D'), 'Q_equiv']) +
            conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'D'), 'Q'].item() -
            demand_quantity.loc[(demand_quantity['F'] == 'D') & (demand_quantity['j'] == 'CA'), 'Q'].item()
        )
        Gdemand_CA = (
            sum(biofuel_quantity.loc[(biofuel_quantity['j'] == 'CA') & (biofuel_quantity['F'] == 'G'), 'Q_equiv']) +
            conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'G'), 'Q'].item() -
            demand_quantity.loc[(demand_quantity['F'] == 'G') & (demand_quantity['j'] == 'CA'), 'Q'].item()
        )
        Jdemand_CA = (
            sum(biofuel_quantity.loc[(biofuel_quantity['j'] == 'CA') & (biofuel_quantity['F'] == 'J'), 'Q_equiv']) +
            conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'J'), 'Q'].item() -
            demand_quantity.loc[(demand_quantity['F'] == 'J') & (demand_quantity['j'] == 'CA'), 'Q'].item()
        )
        Ddemand_NC = (
            sum(biofuel_quantity.loc[(biofuel_quantity['j'] == 'NC') & (biofuel_quantity['F'] == 'D'), 'Q_equiv']) +
            conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'NC') & (conventionalfuel_quantity['F'] == 'D'), 'Q'].item() -
            demand_quantity.loc[(demand_quantity['F'] == 'D') & (demand_quantity['j'] == 'NC'), 'Q'].item()
        )
        Gdemand_NC = (
            sum(biofuel_quantity.loc[(biofuel_quantity['j'] == 'NC') & (biofuel_quantity['F'] == 'G'), 'Q_equiv']) +
            conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'NC') & (conventionalfuel_quantity['F'] == 'G'), 'Q'].item() -
            demand_quantity.loc[(demand_quantity['F'] == 'G') & (demand_quantity['j'] == 'NC'), 'Q'].item()
        )
        Jdemand_NC = (
            sum(biofuel_quantity.loc[(biofuel_quantity['j'] == 'NC') & (biofuel_quantity['F'] == 'J'), 'Q_equiv']) +
            conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'NC') & (conventionalfuel_quantity['F'] == 'J'), 'Q'].item() -
            demand_quantity.loc[(demand_quantity['F'] == 'J') & (demand_quantity['j'] == 'NC'), 'Q'].item()
        )

        Constraints.append(Ddemand_CA == 0)
        Constraints.append(Gdemand_CA == 0)
        Constraints.append(Jdemand_CA == 0)
        Constraints.append(Ddemand_NC == 0)
        Constraints.append(Gdemand_NC == 0)
        Constraints.append(Jdemand_NC == 0)

        Constraints.append(
            demand_quantity.loc[(demand_quantity['F'] == 'D') & (demand_quantity['j'] == 'CA'), 'Q_actual'].item() -
            (biofuel_quantity.loc[(biofuel_quantity['F'] == 'D') & (biofuel_quantity['j'] == 'CA'), 'Q'].sum() +
                conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'D') & (conventionalfuel_quantity['j'] == 'CA'), 'Q'].item()
            ) == 0
        )
        Constraints.append(
            demand_quantity.loc[(demand_quantity['F'] == 'D') & (demand_quantity['j'] == 'NC'), 'Q_actual'].item() -
            (biofuel_quantity.loc[(biofuel_quantity['F'] == 'D') & (biofuel_quantity['j'] == 'NC'), 'Q'].sum() +
                conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'D') & (conventionalfuel_quantity['j'] == 'NC'), 'Q'].item()
            ) == 0
        )
        Constraints.append(
            demand_quantity.loc[(demand_quantity['F'] == 'G') & (demand_quantity['j'] == 'CA'), 'Q_actual'].item() -
            (biofuel_quantity.loc[(biofuel_quantity['F'] == 'G') & (biofuel_quantity['j'] == 'CA'), 'Q'].sum() +
                conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'G') & (conventionalfuel_quantity['j'] == 'CA'), 'Q'].item()
            ) == 0
        )
        Constraints.append(
            demand_quantity.loc[(demand_quantity['F'] == 'G') & (demand_quantity['j'] == 'NC'), 'Q_actual'].item() -
            (biofuel_quantity.loc[(biofuel_quantity['F'] == 'G') & (biofuel_quantity['j'] == 'NC'), 'Q'].sum() +
                conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'G') & (conventionalfuel_quantity['j'] == 'NC'), 'Q'].item()
            ) == 0
        )
        Constraints.append(
            demand_quantity.loc[(demand_quantity['F'] == 'J') & (demand_quantity['j'] == 'CA'), 'Q_actual'].item() -
            (biofuel_quantity.loc[(biofuel_quantity['F'] == 'J') & (biofuel_quantity['j'] == 'CA'), 'Q'].sum() +
                conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'J') & (conventionalfuel_quantity['j'] == 'CA'), 'Q'].item()
            ) == 0
        )
        Constraints.append(
            demand_quantity.loc[(demand_quantity['F'] == 'J') & (demand_quantity['j'] == 'NC'), 'Q_actual'].item() -
            (biofuel_quantity.loc[(biofuel_quantity['F'] == 'J') & (biofuel_quantity['j'] == 'NC'), 'Q'].sum() +
                conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'J') & (conventionalfuel_quantity['j'] == 'NC'), 'Q'].item()
            ) == 0
        )

        # Price breakup - get obligations
        for i in range(24):
            Constraints.append(
                Price.iloc[i]['RIN_obligation'] - (output['kappa_RFS'][i] * RIN_prices.loc[RIN_prices['code'] == 'D4', 'P'].item() * output['D4'][i] +
                output['kappa_RFS'][i] * RIN_prices.loc[RIN_prices['code'] == 'D6', 'P'].item() * output['D6'][i]) == 0
            )

        Constraints.append(
            Price.loc[(Price['f'] == 'B0') & (Price['j'] == 'CA'), 'RIN_obligation'].item() ==
            RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].item() * RIN_prices.loc[RIN_prices['code'] == 'D4', 'P'].item() +
            (RVO_percent.loc[RVO_percent['mandate'] == 'RF', 'P'].item() - RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].item()) * RIN_prices.loc[RIN_prices['code'] == 'D6', 'P'].item()
        )

        Constraints.append(
            Price.loc[(Price['f'] == 'E0') & (Price['j'] == 'CA'), 'RIN_obligation'].item() ==
            RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].item() * RIN_prices.loc[RIN_prices['code'] == 'D4', 'P'].item() +
            (RVO_percent.loc[RVO_percent['mandate'] == 'RF', 'P'].item() - RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].item()) * RIN_prices.loc[RIN_prices['code'] == 'D6', 'P'].item()
        )

        Constraints.append(
            Price.loc[(Price['f'] == 'B0') & (Price['j'] == 'NC'), 'RIN_obligation'].item() ==
            RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].item() * RIN_prices.loc[RIN_prices['code'] == 'D4', 'P'].item() +
            (RVO_percent.loc[RVO_percent['mandate'] == 'RF', 'P'].item() - RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].item()) * RIN_prices.loc[RIN_prices['code'] == 'D6', 'P'].item()
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'E0') & (Price['j'] == 'NC'), 'RIN_obligation'].item() ==
            RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].item() * RIN_prices.loc[RIN_prices['code'] == 'D4', 'P'].item() +
            (RVO_percent.loc[RVO_percent['mandate'] == 'RF', 'P'].item() - RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].item()) * RIN_prices.loc[RIN_prices['code'] == 'D6', 'P'].item()
        )

        for i in range(24):
            Constraints.append(
                Price.iloc[i]['LCFS_CA_obligation'] - (LCFS_prices.loc['LCFS', 'P'] * output.iloc[i]['CALCFSSide'] * (output.iloc[i]['CI_std_LCFS_CA'] - output.iloc[i]['CI_LCFS']) *
                (output.iloc[i]['j'] == 'CA') * output.iloc[i]['ED_LCFS'] / 1000000) == 0
            )

        Constraints.append(
            Price.loc[(Price['f'] == 'B0') & (Price['j'] == 'CA'), 'LCFS_CA_obligation'].item() -
            LCFS_prices.loc['LCFS', 'P'] * (
                -conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'D'), 'CI_std_LCFS_CA'].item() +
                conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'D'), 'CI_LCFS'].item()
            ) * conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'D'), 'ED_LCFS'].item() / 1000000 == 0
        )

        Constraints.append(
            Price.loc[(Price['f'] == 'E0') & (Price['j'] == 'CA'), 'LCFS_CA_obligation'].item() -
            LCFS_prices.loc['LCFS', 'P'] * (
                -conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'G'), 'CI_std_LCFS_CA'].item() +
                conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'G'), 'CI_LCFS'].item()
            ) * conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'G'), 'ED_LCFS'].item() / 1000000 == 0
        )

        # Add the new constraints for LCFS_SAF_obligation
        for i in range(24):
            Constraints.append(
                Price.iloc[i]['LCFS_SAF_obligation'] - 
                (LCFS_prices_SAF.loc['LCFS_SAF', 'P'] * output.iloc[i]['SAFLCFSside'] * 
                (output.iloc[i]['CI_std_LCFS'] - output.iloc[i]['CI_LCFS']) * output.iloc[i]['ED_LCFS'] / 1000000) == 0
            )

        # Add the constraints for LCFS_SAF_obligation for 'J0' in CA
        Constraints.append(
            Price.loc[(Price['f'] == 'J0') & (Price['j'] == 'CA'), 'LCFS_SAF_obligation'].item() - 
            (LCFS_prices_SAF.loc['LCFS_SAF', 'P'] * 
            (-SAF_intensity_standard_P + 
            conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'J', 'CI_LCFS'].unique()[0]) * 
            conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'J', 'ED_LCFS'].unique()[0] / 1000000) == 0
        )

        # Add the constraints for LCFS_SAF_obligation for 'J0' in NC
        Constraints.append(
            Price.loc[(Price['f'] == 'J0') & (Price['j'] == 'NC'), 'LCFS_SAF_obligation'].item() - 
            (LCFS_prices_SAF.loc['LCFS_SAF', 'P'] * 
            (-SAF_intensity_standard_P + 
            conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'J', 'CI_LCFS'].unique()[0]) * 
            conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'J', 'ED_LCFS'].unique()[0] / 1000000) == 0
        )

        for i in range(24):
            Constraints.append(
                Price.iloc[i]['Tax'] - (
                    val * ((output.iloc[i]['f'] == 'SAF') & (output.iloc[i]['j'] == 'NC')) * output.iloc[i]['Statetaxheaviside'] +
                    1 * ((output.iloc[i]['f'] == 'B100') | (output.iloc[i]['f'] == 'RD') | (output.iloc[i]['f'] == 'gasE100')) * ((47.39 - output.iloc[i]['CI_Tax']) / 47.39) * output.iloc[i]['Taxheavyside']* (output['s'][i] != 'sugarcane') +
                    max([
                        1 * (output.iloc[i]['f'] == 'SAF') * ((47.39 - output.iloc[i]['CI_Tax']) / 47.39) * output.iloc[i]['Taxheavyside'],
                        85 * 30 * output.iloc[i]['ED_LCFS'] / (10**6) * output.iloc[i]['CCS_tech']
                    ])* (output['s'][i] != 'sugarcane')
                )  == 0
            )

        for i in range(24):
            Constraints.append(
                Price.iloc[i]['P_all'] - (
                    biofuel_prices.loc[(biofuel_prices['f'] == output.iloc[i]['f']) & (biofuel_prices['j'] == output.iloc[i]['j']), 'P'].item() +
                    output['kappa_RFS'][i] * RIN_prices.loc[RIN_prices['code'] == 'D4', 'P'].item() * output['D4'][i] +
                    output['kappa_RFS'][i] * RIN_prices.loc[RIN_prices['code'] == 'D6', 'P'].item() * output['D6'][i] +
                    val * ((output.iloc[i]['f'] == 'SAF') & (output.iloc[i]['j'] == 'NC')) * output.iloc[i]['Statetaxheaviside'] +
                    1 * ((output.iloc[i]['f'] == 'B100') | (output.iloc[i]['f'] == 'RD') | (output.iloc[i]['f'] == 'gasE100')) * ((47.39 - output.iloc[i]['CI_Tax']) / 47.39) * output.iloc[i]['Taxheavyside']* (output['s'][i] != 'sugarcane') +
                    max([
                        1 * (output.iloc[i]['f'] == 'SAF') * ((47.39 - output.iloc[i]['CI_Tax']) / 47.39) * output.iloc[i]['Taxheavyside'],
                        85 * 30 * output.iloc[i]['ED_LCFS'] / (10**6) * output.iloc[i]['CCS_tech']
                    ])* (output['s'][i] != 'sugarcane') + (LCFS_prices.loc['LCFS', 'P'] * output.iloc[i]['CALCFSSide'] * (output.iloc[i]['CI_std_LCFS_CA'] - output.iloc[i]['CI_LCFS']) * output.iloc[i]['ED_LCFS'] / 1000000) +
                    (LCFS_prices_SAF.loc['LCFS_SAF', 'P'] * output.iloc[i]['SAFLCFSside'] * (output.iloc[i]['CI_std_LCFS'] - output.iloc[i]['CI_LCFS']) * output.iloc[i]['ED_LCFS'] / 1000000) 
                ) == 0
            )

        # Add constraints for Price DataFrame
        Constraints.append(
            Price.loc[(Price['f'] == 'B0') & (Price['j'] == 'CA'), 'P_all'].values[0] -
            input_demand.loc[input_demand['F'] == 'D', 'PC_F0'].values[0] == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'E0') & (Price['j'] == 'CA'), 'P_all'].values[0] -
            input_demand.loc[input_demand['F'] == 'G', 'PC_F0'].values[0] == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'J0') & (Price['j'] == 'CA'), 'P_all'].values[0] -
            input_demand.loc[input_demand['F'] == 'J', 'PC_F0'].values[0] == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'B0') & (Price['j'] == 'NC'), 'P_all'].values[0] -
            input_demand.loc[input_demand['F'] == 'D', 'PC_F0'].values[0] == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'E0') & (Price['j'] == 'NC'), 'P_all'].values[0] -
            input_demand.loc[input_demand['F'] == 'G', 'PC_F0'].values[0] == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'J0') & (Price['j'] == 'NC'), 'P_all'].values[0] -
            input_demand.loc[input_demand['F'] == 'J', 'PC_F0'].values[0] == 0
        )

        for i in range(24):
            Constraints.append(
                Price.iloc[i]['PC_F0'] - biofuel_prices.loc[(biofuel_prices['f'] == output.iloc[i]['f']) & (biofuel_prices['j'] == output.iloc[i]['j']), 'P'].item() == 0
            )

        Constraints.append(
            Price.loc[(Price['f'] == 'B0') & (Price['j'] == 'CA'), 'PC_F0'].values[0] -
            input_demand.loc[input_demand['F'] == 'D', 'PC_F0'].values[0] == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'E0') & (Price['j'] == 'CA'), 'PC_F0'].values[0] -
            input_demand.loc[input_demand['F'] == 'G', 'PC_F0'].values[0] == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'J0') & (Price['j'] == 'CA'), 'PC_F0'].values[0] -
            input_demand.loc[input_demand['F'] == 'J', 'PC_F0'].values[0] == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'B0') & (Price['j'] == 'NC'), 'PC_F0'].values[0] -
            input_demand.loc[input_demand['F'] == 'D', 'PC_F0'].values[0] == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'E0') & (Price['j'] == 'NC'), 'PC_F0'].values[0] -
            input_demand.loc[input_demand['F'] == 'G', 'PC_F0'].values[0] == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'J0') & (Price['j'] == 'NC'), 'PC_F0'].values[0] -
            input_demand.loc[input_demand['F'] == 'J', 'PC_F0'].values[0] == 0
        )


        for i in range(24):
            Constraints.append(
                Price.iloc[i]['P_RIN_SAF_LCFS_detach'] - biofuel_prices.loc[(biofuel_prices['f'] == output.iloc[i]['f']) & (biofuel_prices['j'] == output.iloc[i]['j']), 'P'].item() == 0
            )

        Constraints.append(
            Price.loc[(Price['f'] == 'B0') & (Price['j'] == 'NC'), 'P_RIN_SAF_LCFS_detach'].item() -
            input_demand.loc[input_demand['F'] == 'D', 'PC_F0'].unique()[0] -
            RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].item() * RIN_prices.loc[RIN_prices['code'] == 'D4', 'P'].item() -
            (RVO_percent.loc[RVO_percent['mandate'] == 'RF', 'P'].item() - RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].item()) * RIN_prices.loc[RIN_prices['code'] == 'D6', 'P'].item() == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'E0') & (Price['j'] == 'NC'), 'P_RIN_SAF_LCFS_detach'].item() -
            input_demand.loc[input_demand['F'] == 'G', 'PC_F0'].unique()[0] -
            RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].item() * RIN_prices.loc[RIN_prices['code'] == 'D4', 'P'].item() -
            (RVO_percent.loc[RVO_percent['mandate'] == 'RF', 'P'].item() - RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].item()) * RIN_prices.loc[RIN_prices['code'] == 'D6', 'P'].item() == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'J0') & (Price['j'] == 'NC'), 'P_RIN_SAF_LCFS_detach'].item() - 
            (LCFS_prices_SAF.loc['LCFS_SAF', 'P'] * 
                (conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'J', 'CI_LCFS'].unique()[0] - 
                SAF_intensity_standard_P) * 
                conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'J', 'ED_LCFS'].unique()[0] / 1000000) -
            input_demand.loc[input_demand['F'] == 'J', 'PC_F0'].unique()[0] == 0
        )

        Constraints.append(
            Price.loc[(Price['f'] == 'B0') & (Price['j'] == 'CA'), 'P_RIN_SAF_LCFS_detach'].item() -
            input_demand.loc[input_demand['F'] == 'D', 'PC_F0'].unique()[0] -
            RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].item() * RIN_prices.loc[RIN_prices['code'] == 'D4', 'P'].item() -
            (RVO_percent.loc[RVO_percent['mandate'] == 'RF', 'P'].item() - RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].item()) * RIN_prices.loc[RIN_prices['code'] == 'D6', 'P'].item() -
            LCFS_prices.loc['LCFS', 'P'] * (
                -conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'D'), 'CI_std_LCFS_CA'].item() +
                conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'D'), 'CI_LCFS'].item()
            ) * conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'D'), 'ED_LCFS'].item() / 1000000 == 0
        )

        Constraints.append(
            Price.loc[(Price['f'] == 'E0') & (Price['j'] == 'CA'), 'P_RIN_SAF_LCFS_detach'].item() -
            input_demand.loc[input_demand['F'] == 'G', 'PC_F0'].unique()[0] -
            RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].item() * RIN_prices.loc[RIN_prices['code'] == 'D4', 'P'].item() -
            (RVO_percent.loc[RVO_percent['mandate'] == 'RF', 'P'].item() - RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].item()) * RIN_prices.loc[RIN_prices['code'] == 'D6', 'P'].item() -
            LCFS_prices.loc['LCFS', 'P'] * (
                -conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'G'), 'CI_std_LCFS_CA'].item() +
                conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'G'), 'CI_LCFS'].item()
            ) * conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'G'), 'ED_LCFS'].item() / 1000000 == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'J0') & (Price['j'] == 'CA'), 'P_RIN_SAF_LCFS_detach'].item() - (LCFS_prices_SAF.loc['LCFS_SAF', 'P'] * 
                (conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'J', 'CI_LCFS'].unique()[0] - 
                SAF_intensity_standard_P) * 
                conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'J', 'ED_LCFS'].unique()[0] / 1000000) -
            input_demand.loc[input_demand['F'] == 'J', 'PC_F0'].unique()[0] == 0
        )

        for i in range(24):
            Constraints.append(
                Price.iloc[i]['P_RIN_SAF_detach'] - (
                    biofuel_prices.loc[(biofuel_prices['f'] == output.iloc[i]['f']) & (biofuel_prices['j'] == output.iloc[i]['j']), 'P'].item() + (LCFS_prices.loc['LCFS', 'P'] * output.iloc[i]['CALCFSSide'] * (output.iloc[i]['CI_std_LCFS_CA'] - output.iloc[i]['CI_LCFS']) * (output.iloc[i]['j'] == 'CA') * output.iloc[i]['ED_LCFS'] / 1000000)
                ) == 0
            )

        Constraints.append(
            Price.loc[(Price['f'] == 'B0') & (Price['j'] == 'CA'), 'P_RIN_SAF_detach'].item() -
            input_demand.loc[input_demand['F'] == 'D', 'PC_F0'].unique()[0] -
            RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].item() * RIN_prices.loc[RIN_prices['code'] == 'D4', 'P'].item() -
            (RVO_percent.loc[RVO_percent['mandate'] == 'RF', 'P'].item() - RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].item()) * RIN_prices.loc[RIN_prices['code'] == 'D6', 'P'].item() == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'E0') & (Price['j'] == 'CA'), 'P_RIN_SAF_detach'].item() -
            input_demand.loc[input_demand['F'] == 'G', 'PC_F0'].unique()[0] -
            RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].item() * RIN_prices.loc[RIN_prices['code'] == 'D4', 'P'].item() -
            (RVO_percent.loc[RVO_percent['mandate'] == 'RF', 'P'].item() - RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].item()) * RIN_prices.loc[RIN_prices['code'] == 'D6', 'P'].item() == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'J0') & (Price['j'] == 'CA'), 'P_RIN_SAF_detach'].item() - (LCFS_prices_SAF.loc['LCFS_SAF', 'P'] * 
                (conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'J', 'CI_LCFS'].unique()[0] - 
                SAF_intensity_standard_P) * 
                conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'J', 'ED_LCFS'].unique()[0] / 1000000) -
            input_demand.loc[input_demand['F'] == 'J', 'PC_F0'].unique()[0] == 0
        )

        Constraints.append(
            Price.loc[(Price['f'] == 'B0') & (Price['j'] == 'NC'), 'P_RIN_SAF_detach'].item() -
            input_demand.loc[input_demand['F'] == 'D', 'PC_F0'].unique()[0] -
            RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].item() * RIN_prices.loc[RIN_prices['code'] == 'D4', 'P'].item() -
            (RVO_percent.loc[RVO_percent['mandate'] == 'RF', 'P'].item() - RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].item()) * RIN_prices.loc[RIN_prices['code'] == 'D6', 'P'].item() == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'E0') & (Price['j'] == 'NC'), 'P_RIN_SAF_detach'].item() -
            input_demand.loc[input_demand['F'] == 'G', 'PC_F0'].unique()[0] -
            RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].item() * RIN_prices.loc[RIN_prices['code'] == 'D4', 'P'].item() -
            (RVO_percent.loc[RVO_percent['mandate'] == 'RF', 'P'].item() - RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].item()) * RIN_prices.loc[RIN_prices['code'] == 'D6', 'P'].item() == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'J0') & (Price['j'] == 'NC'), 'P_RIN_SAF_detach'].item() - (LCFS_prices_SAF.loc['LCFS_SAF', 'P'] * 
                (conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'J', 'CI_LCFS'].unique()[0] - 
                SAF_intensity_standard_P) * 
                conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'J', 'ED_LCFS'].unique()[0] / 1000000) -
            input_demand.loc[input_demand['F'] == 'J', 'PC_F0'].unique()[0] == 0
        )

        for i in range(24):
            Constraints.append(
                Price.iloc[i]['Q_actual'] - biofuel_quantity.iloc[i]['Q'] == 0
            )

        Constraints.append(
            Price.loc[(Price['f'] == 'B0') & (Price['j'] == 'CA'), 'Q_actual'].item() -
            conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'D') & (conventionalfuel_quantity['j'] == 'CA'), 'Q'].item() == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'E0') & (Price['j'] == 'CA'), 'Q_actual'].item() -
            conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'G') & (conventionalfuel_quantity['j'] == 'CA'), 'Q'].item() == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'J0') & (Price['j'] == 'CA'), 'Q_actual'].item() -
            conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'J') & (conventionalfuel_quantity['j'] == 'CA'), 'Q'].item() == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'B0') & (Price['j'] == 'NC'), 'Q_actual'].item() -
            conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'D') & (conventionalfuel_quantity['j'] == 'NC'), 'Q'].item() == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'E0') & (Price['j'] == 'NC'), 'Q_actual'].item() -
            conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'G') & (conventionalfuel_quantity['j'] == 'NC'), 'Q'].item() == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'J0') & (Price['j'] == 'NC'), 'Q_actual'].item() -
            conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'J') & (conventionalfuel_quantity['j'] == 'NC'), 'Q'].item() == 0
        )

        # Price breakup - get obligations
        for i in range(24):
            Constraints.append(
                Price.iloc[i]['Q_equiv'] - biofuel_quantity.iloc[i]['Q_equiv'] == 0
            )

        Constraints.append(
            Price.loc[(Price['f'] == 'B0') & (Price['j'] == 'CA'), 'Q_equiv'].item() -
            conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'D') & (conventionalfuel_quantity['j'] == 'CA'), 'Q'].item() == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'E0') & (Price['j'] == 'CA'), 'Q_equiv'].item() -
            conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'G') & (conventionalfuel_quantity['j'] == 'CA'), 'Q'].item() == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'J0') & (Price['j'] == 'CA'), 'Q_equiv'].item() -
            conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'J') & (conventionalfuel_quantity['j'] == 'CA'), 'Q'].item() == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'B0') & (Price['j'] == 'NC'), 'Q_equiv'].item() -
            conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'D') & (conventionalfuel_quantity['j'] == 'NC'), 'Q'].item() == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'E0') & (Price['j'] == 'NC'), 'Q_equiv'].item() -
            conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'G') & (conventionalfuel_quantity['j'] == 'NC'), 'Q'].item() == 0
        )
        Constraints.append(
            Price.loc[(Price['f'] == 'J0') & (Price['j'] == 'NC'), 'Q_equiv'].item() -
            conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'J') & (conventionalfuel_quantity['j'] == 'NC'), 'Q'].item() == 0
        )

        # Define the feedstocks
        feedstocks = ['soyoil', 'animalfat', 'corn', 'sugarcane']
        #feedstocks = ['soyoil', 'animalfat', 'corn']


        # Loop through each feedstock and add the corresponding constraint
        for feedstock in feedstocks:
            Constraints.append(
                feedstock_quantity.loc[feedstock_quantity['s'] == feedstock, 'Q'].item() ==
                input_feedstock.loc[input_feedstock['s'] == feedstock, 'alpha'].item() *
                feedstock_prices.loc[feedstock_prices['s'] == feedstock, 'P'].item() **
                input_feedstock.loc[input_feedstock['s'] == feedstock, 'beta'].item()
            )

        # Calculate the demand for each feedstock
        for feedstock in feedstocks:
            demand = (
                sum(biofuel_quantity.loc[biofuel_quantity['s'] == feedstock, 'Conversion_demand'] * 
                    biofuel_quantity.loc[biofuel_quantity['s'] == feedstock, 'Q']) - 
                feedstock_quantity.loc[feedstock_quantity['s'] == feedstock, 'Q'].item()
            )
            Constraints.append(demand == 0)


        # Create binary variables for PD4zeta and PD6zeta
        PD4zeta = m.addVar(vtype=GRB.BINARY, name="PD4zeta")
        PD6zeta = m.addVar(vtype=GRB.BINARY, name="PD6zeta")

        # Add constraints for RIN prices and biofuel quantities
        M11 = 10
        Constraints.append((RIN_prices.loc['D4', 'P'] - RIN_prices.loc['D6', 'P']) >= 0)
        Constraints.append((RIN_prices.loc['D4', 'P'] - RIN_prices.loc['D6', 'P']) <= (1 - PD4zeta) * M11)

        #here the conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'G') | (conventionalfuel_quantity['F'] == 'D'), 'Q'].sum() include not only transportation fuels, but also other sectors
        # Constraints.append((biofuel_quantity.loc[biofuel_quantity['f'].isin(['B100', 'RD', 'SAF']), 'Q'].sum() -
        #     RVO_percent.loc['BBD', 'P'] / 1.6 * conventionalfuel_quantity.loc[conventionalfuel_quantity['F'].isin(['G', 'D']), 'Q'].sum()) >= 0)
        # Constraints.append((biofuel_quantity.loc[biofuel_quantity['f'].isin(['B100', 'RD', 'SAF']), 'Q'].sum() -
        #     RVO_percent.loc['BBD', 'P'] / 1.6 * conventionalfuel_quantity.loc[conventionalfuel_quantity['F'].isin(['G', 'D']), 'Q'].sum()) <= PD4zeta * M11)

        # Constraints.append((biofuel_quantity.loc[biofuel_quantity['f'].isin(['B100', 'RD', 'SAF']), 'Q'].sum() -
        #     RVO_percent.loc['BBD', 'P'] / 1.6 * conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'G') | (conventionalfuel_quantity['F'] == 'D'), 'Q'].sum()) >= 0)
        # Constraints.append((biofuel_quantity.loc[biofuel_quantity['f'].isin(['B100', 'RD', 'SAF']), 'Q'].sum() -
        #     RVO_percent.loc['BBD', 'P'] / 1.6 * conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'G') | (conventionalfuel_quantity['F'] == 'D'), 'Q'].sum()) <= PD4zeta * M11)

        #SAF ATJ without CCS only get D6 RINs
        # Constraints.append((biofuel_quantity.loc[biofuel_quantity['D4'] == 1, 'Q'].sum() -
        #     RVO_percent.loc['BBD', 'P'] / 1.6 * conventionalfuel_quantity.loc[conventionalfuel_quantity['F'].isin(['G', 'D']), 'Q'].sum()) >= 0)
        # Constraints.append((biofuel_quantity.loc[biofuel_quantity['D4'] == 1, 'Q'].sum() -
        #     RVO_percent.loc['BBD', 'P'] / 1.6 * conventionalfuel_quantity.loc[conventionalfuel_quantity['F'].isin(['G', 'D']), 'Q'].sum()) <= PD4zeta * M11)

        Constraints.append((biofuel_quantity.loc[biofuel_quantity['D4'] == 1, 'Q'].sum() -
            input_compliance.loc[input_compliance['mandate'] == 'BBD', 'Q'].item() / 1.6 * conventionalfuel_quantity.loc[conventionalfuel_quantity['F'].isin(['G', 'D']), 'Q'].sum()) >= 0)
        Constraints.append((biofuel_quantity.loc[biofuel_quantity['D4'] == 1, 'Q'].sum() -
            input_compliance.loc[input_compliance['mandate'] == 'BBD', 'Q'].item() / 1.6 * conventionalfuel_quantity.loc[conventionalfuel_quantity['F'].isin(['G', 'D']), 'Q'].sum()) <= PD4zeta * M11)


        M12 = 50
        Constraints.append(RIN_prices.loc['D6', 'P'] >= 0)
        Constraints.append(RIN_prices.loc['D6', 'P'] <= (1 - PD6zeta) * M12)

        # Constraints.append((sum((biofuel_quantity.loc[biofuel_quantity['f'] == 'B100', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'B100', 'Q'])) + sum(biofuel_quantity.loc[biofuel_quantity['f'] == 'RD', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'RD', 'Q']) + sum(biofuel_quantity.loc[biofuel_quantity['f'] == 'SAF', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'SAF', 'Q']) + sum(biofuel_quantity.loc[biofuel_quantity['f'] == 'gasE100', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'gasE100', 'Q']) - RVO_percent.loc['RF', 'P'] * conventionalfuel_quantity.loc[conventionalfuel_quantity['F'].isin(['G', 'D']), 'Q'].sum()) >= 0)
        # Constraints.append((sum((biofuel_quantity.loc[biofuel_quantity['f'] == 'B100', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'B100', 'Q'])) + sum(biofuel_quantity.loc[biofuel_quantity['f'] == 'RD', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'RD', 'Q']) + sum(biofuel_quantity.loc[biofuel_quantity['f'] == 'SAF', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'SAF', 'Q']) +sum(biofuel_quantity.loc[biofuel_quantity['f'] == 'gasE100', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'gasE100', 'Q']) - RVO_percent.loc['RF', 'P'] * conventionalfuel_quantity.loc[conventionalfuel_quantity['F'].isin(['G', 'D']), 'Q'].sum()) <= PD6zeta * M12)

        # Constraints.append((sum((biofuel_quantity.loc[biofuel_quantity['f'] == 'B100', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'B100', 'Q'])) + sum(biofuel_quantity.loc[biofuel_quantity['f'] == 'RD', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'RD', 'Q']) + sum(biofuel_quantity.loc[biofuel_quantity['f'] == 'SAF', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'SAF', 'Q']) + sum(biofuel_quantity.loc[biofuel_quantity['f'] == 'gasE100', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'gasE100', 'Q']) - RVO_percent.loc['RF', 'P'] * conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'G') | (conventionalfuel_quantity['F'] == 'D'), 'Q'].sum()) >= 0)

        # Constraints.append((sum((biofuel_quantity.loc[biofuel_quantity['f'] == 'B100', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'B100', 'Q'])) + sum(biofuel_quantity.loc[biofuel_quantity['f'] == 'RD', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'RD', 'Q']) + sum(biofuel_quantity.loc[biofuel_quantity['f'] == 'SAF', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'SAF', 'Q']) +sum(biofuel_quantity.loc[biofuel_quantity['f'] == 'gasE100', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'gasE100', 'Q']) - RVO_percent.loc['RF', 'P'] * conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'] == 'G') | (conventionalfuel_quantity['F'] == 'D'), 'Q'].sum()) <= PD6zeta * M12)


        Constraints.append((sum((biofuel_quantity.loc[biofuel_quantity['f'] == 'B100', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'B100', 'Q'])) + sum(biofuel_quantity.loc[biofuel_quantity['f'] == 'RD', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'RD', 'Q']) + sum(biofuel_quantity.loc[biofuel_quantity['f'] == 'SAF', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'SAF', 'Q']) + sum(biofuel_quantity.loc[biofuel_quantity['f'] == 'gasE100', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'gasE100', 'Q']) - input_compliance.loc[input_compliance['mandate'] == 'RF', 'Q'].item() * conventionalfuel_quantity.loc[conventionalfuel_quantity['F'].isin(['G', 'D']), 'Q'].sum()) >= 0)

        Constraints.append((sum((biofuel_quantity.loc[biofuel_quantity['f'] == 'B100', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'B100', 'Q'])) + sum(biofuel_quantity.loc[biofuel_quantity['f'] == 'RD', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'RD', 'Q']) + sum(biofuel_quantity.loc[biofuel_quantity['f'] == 'SAF', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'SAF', 'Q']) +sum(biofuel_quantity.loc[biofuel_quantity['f'] == 'gasE100', 'kappa_RFS'] * biofuel_quantity.loc[biofuel_quantity['f'] == 'gasE100', 'Q']) - input_compliance.loc[input_compliance['mandate'] == 'RF', 'Q'].item() * conventionalfuel_quantity.loc[conventionalfuel_quantity['F'].isin(['G', 'D']), 'Q'].sum()) <= PD6zeta * M12)

        
        # SAF goal

        # # Create binary variable for SAFCreditzeta
        # SAFCreditzeta = m.addVar(vtype=GRB.BINARY, name="SAFCreditzeta")

        # M13 = 3
        # # Add constraints for SAF credit prices and biofuel quantities
        # Constraints.append(SAFcredit_prices.loc['SAF', 'P'] >= 0)
        # Constraints.append(SAFcredit_prices.loc['SAF', 'P'] <= (1 - SAFCreditzeta) * M13)

        # # Calculate the sum of SAF biofuel quantities
        # sum_SAF_biofuel_quantity = biofuel_quantity.loc[biofuel_quantity['f'] == 'SAF', 'Q'].sum()

        # # Add constraints for SAF biofuel quantities
        # Constraints.append(sum_SAF_biofuel_quantity - input_compliance.loc[input_compliance['mandate'] == 'SAF', 'Q'].item() >= 0)
        # Constraints.append(sum_SAF_biofuel_quantity - input_compliance.loc[input_compliance['mandate'] == 'SAF', 'Q'].item() <= SAFCreditzeta * M13)



        # Calculate DeltaCI
        DeltaCI_expr = (
            sum(biofuel_quantity.loc[biofuel_quantity['j'] == 'CA', 'Q'] *
                (biofuel_quantity.loc[biofuel_quantity['j'] == 'CA', 'CI_std_LCFS_CA'] - biofuel_quantity.loc[biofuel_quantity['j'] == 'CA', 'CI_LCFS']) *
                biofuel_quantity.loc[biofuel_quantity['j'] == 'CA', 'ED_LCFS'])/10**6 +
            sum(input_compliance.loc[input_compliance['mandate'] == 'LCFS_Q', 'Q']) -
            sum(conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'D'), 'Q'] *
                (-conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'D'), 'CI_std_LCFS_CA'] +
                conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'D'), 'CI_LCFS']) *
                conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'D'), 'ED_LCFS'])/10**6 -
            sum(conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'G'), 'Q'] *
                (-conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'G'), 'CI_std_LCFS_CA'] +
                conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'G'), 'CI_LCFS']) *
                conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'G'), 'ED_LCFS'])/10**6
        )

        # Define variables
        #y1 = m.addVar(vtype=GRB.BINARY, name='y1')
        #y2 = m.addVar(vtype=GRB.BINARY, name='y2')

        #m.addConstr(y1 + y2 <= 1, "only_one_case")
        # --- Regime 1: q < 0, y2 = 1, p = a ---
        #M14 = 100
        # m.addConstr(DeltaCI_expr <= -Sigma + M14*(1 - y2), "q_neg_active")
        # m.addConstr(LCFS_prices_P >= P_cap * y2, "p_y1_lb")
        # m.addConstr(LCFS_prices_P <= P_cap * y2 + M14 * (1 - y2), "p_y1_ub")

        # # --- Regime 2: q = 0, y1=1, p=lambda*a ---
        # m.addConstr(DeltaCI_expr >= -Sigma - M14*(1 - y1), "q_zero_active_lb")
        # m.addConstr(DeltaCI_expr <= Sigma + M14*(1 - y1), "q_zero_active_ub")
        # m.addConstr(LCFS_prices_P >= Sigma * y1, "p_y2_lb")
        # m.addConstr(LCFS_prices_P <= P_cap * y1, "p_y2_ub")

        # # --- Regime 3: q > 0, neither y1 nor y2, p = 0 ---
        # m.addConstr(DeltaCI_expr >= Sigma - M14*(y1 + y2), "q_pos_active")
        # m.addConstr(LCFS_prices_P <= M14 * (y1 + y2), "p_y1y2_zero")  # p=0 if y1=0 and y2=0
        
        epsilon = 1e-4
        M14 = 300  # Large constant for the constraints
        yneg = m.addVar(vtype=GRB.BINARY, name='yneg')
        ypos = m.addVar(vtype=GRB.BINARY, name='ypos')

        m.addConstr(yneg + ypos <= 1, "only_one_case")
        
        m.addConstr(DeltaCI_expr >= epsilon - M14*(1 - ypos), "q_pos_active_epsilon")
        m.addConstr(DeltaCI_expr <= M14*ypos, "q_pos_active_epsilon2")
        m.addConstr(DeltaCI_expr <= -epsilon + M14*(1 - yneg), "q_neg_active_epsilon")
        m.addConstr(DeltaCI_expr >= -M14*yneg, "q_neg_active_epsilon2")

        m.addConstr(LCFS_prices_P <= 0 + M14*(1-ypos))
        m.addConstr(LCFS_prices_P >= 0 - M14*(1-ypos))
        m.addConstr(LCFS_prices_P >= P_cap*yneg)
        m.addConstr(LCFS_prices_P <= P_cap*yneg + M14* (1-yneg))
        m.addConstr(LCFS_prices_P >= epsilon - M14 * (yneg + ypos), name='p_strict_lb')
        m.addConstr(LCFS_prices_P <= P_cap - epsilon + M14 * (yneg + ypos), name='p_strict_ub')

        # constraints for SAF LCFS certificate
        Constraints.append((sum(biofuel_quantity.loc[biofuel_quantity['f'] == 'SAF', 'Q']) - input_compliance.loc[input_compliance['mandate'] == 'SAF', 'Q'].item()) == 0)

        # Create the AlphaLCFS_SAF variable
        AlphaLCFS_SAF = m.addVar(vtype=GRB.CONTINUOUS, name="AlphaLCFS_SAF")

        # Add the constraints for AlphaLCFS_SAF
        Constraints.append(LCFS_prices_SAF.loc['LCFS_SAF', 'P'] - AlphaLCFS_SAF * P_cap_SAF == 0)
        Constraints.append(AlphaLCFS_SAF >= 0)
        Constraints.append(AlphaLCFS_SAF <= 1)

        # Define the DeltaCI_SAF expression
        DeltaCI_SAF_expr = (
            sum((biofuel_quantity['Q'] * (biofuel_quantity['CI_std_LCFS'] - biofuel_quantity['CI_LCFS']) * biofuel_quantity['ED_LCFS'] * biofuel_quantity['SAFLCFSside'])) -
            sum((conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'J', 'Q'] *
                (-conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'J', 'CI_std_LCFS'] + 
                conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'J', 'CI_LCFS']) * 
                conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'J', 'ED_LCFS']))
        )

        Constraints.append(DeltaCI_SAF_expr == 0)

        # Add constraints to the model
        for constraint in Constraints:
            m.addConstr(constraint)

        #m.Params.FeasibilityTol = 2e-7
        m.Params.FeasibilityTol = 4e-8
        m.Params.OptimalityTol = 1e-7
        m.Params.NumericFocus = 1       
        # # Add these parameters before m.optimize()
        # m.Params.NumericFocus = 3  # Maximum focus on numerical stability
        # m.Params.FeasibilityTol = 2e-7  # Relax feasibility tolerance
        # m.Params.OptimalityTol = 1e-7   # Relax optimality tolerance
        # m.Params.ScaleFlag = 2          # Aggressive scaling
        # m.Params.Quad = 1               # Use quad precision
        #       
        m.optimize()

        if m.status == GRB.OPTIMAL:
            print(f"Solution found for break {break_idx}")
            print(f"Break {break_idx}: break = {break_val}, val = {val}")
            
            # Store objective value for this break
            objective_values.append({
                'Break': break_idx,
                'Break_Value': break_val,
                'Val': val,
                'Objective_Value': m.ObjVal,
                'Total_Profit': 0,  # Will be updated after total_profit calculation
                'State_tax': state_tax_profit
                #'Denominator': denominator.X  # Add denominator value
            })

            # with open(f'results/variables_break{break_idx}.txt', 'w') as f:
            #     f.write(f"Objective value: {m.ObjVal}\n")  # <-- Add this line
            #     DeltaCI_value = DeltaCI_expr.getValue()
            #     f.write(f"DeltaCI_expr value: {DeltaCI_value:g}\n")
            #     for v in m.getVars():
            #         f.write(f"{v.VarName} {v.X:g}\n")

            # # After m.optimize() and if m.status == GRB.OPTIMAL:
            # with open(f'results/hefa_value_break{break_idx}1.txt', 'w') as f:
            #     for i in range(2):
            #         f.write(f"hefa_value[{i}] = {hefa_value[i].X}\n")

            names = [
                'B100-soyoil-CA', 'B100-animalfat-CA', 'RD-soyoil-CA', 'RD-animalfat-CA', 'SAF-animalfat-CA', 'gasE100-corn-CA',
                'B100-soyoil-NC', 'B100-animalfat-NC', 'RD-soyoil-NC', 'RD-animalfat-NC', 'SAF-animalfat-NC', 'gasE100-corn-NC',
                'gasE100-sugarcane-CA', 'gasE100-sugarcane-NC', 'SAF-soyoil-CA', 'SAF-soyoil-NC', 'SAF-sugarcane-ETJ-CA', 'SAF-sugarcane-ETJCCS-CA',
                'SAF-corn-ETJ-CA', 'SAF-corn-ETJCCS-CA', 'SAF-sugarcane-ETJ-NC', 'SAF-sugarcane-ETJCCS-NC', 'SAF-corn-ETJ-NC', 'SAF-corn-ETJCCS-NC'
            ]
            
            # Collect results for each term in a list of dicts
            results = []
            for i in range(24):
                v1 = biofuel_prices.loc[(biofuel_prices['f'] == output['f'][i]) & (biofuel_prices['j'] == output['j'][i]), 'P'].item().X
                #v2 = output['PC'][i] + output['Conversion'][i] * feedstock_prices.loc[feedstock_prices['s'] == output['s'][i], 'P'].item().X
                v2 = (output['PC'][i] + output['Conversion'][i] * feedstock_prices.loc[feedstock_prices['s'] == output['s'][i], 'P'].item().X*(1+duty_e100sugarcane*(output['s'][i] == 'sugarcane')))
                v3 = output['kappa_RFS'][i] * RIN_prices.loc[RIN_prices['code'] == 'D4', 'P'].item().X * output['D4'][i] + output['kappa_RFS'][i] * RIN_prices.loc[RIN_prices['code'] == 'D6', 'P'].item().X * output['D6'][i]
                v4 = 1 * ((output['f'][i] == 'B100') | (output['f'][i] == 'RD') | (output['f'][i] == 'gasE100')) * ((47.39 - output['CI_Tax'][i]) / 47.39) * output['Taxheavyside'][i]* (output['s'][i] != 'sugarcane') +  max([
                    1 * (output['f'][i] == 'SAF') * ((47.39 - output['CI_Tax'][i]) / 47.39) * output['Taxheavyside'][i],
                    85 * 30 * output['ED_LCFS'][i] / (10**6) * output['CCS_tech'][i]
                ])* (output['s'][i] != 'sugarcane')
                #v5 = 0
                #v5 = (output['f'][i] == 'SAF') * SAFcredit_prices['P'].item().X
                v5 = LCFS_prices_SAF['P'][0].X * (SAF_intensity_standard_P.X - output['CI_LCFS'][i]) * output['ED_LCFS'][i] * output['SAFLCFSside'][i] / 1000000
                v6 = LCFS_prices['P'][0].X * (output['CI_std_LCFS_CA'][i] - output['CI_LCFS'][i]) * output['ED_LCFS'][i] * output['CALCFSSide'][i]  / 1000000
                v7 = ((output['f'][i] == 'SAF') & (output['j'][i] == 'NC')) * output['Statetaxheaviside'][i] * val
                v8 = mu['value'][i].X
                v9 = nu.loc[(nu['f'] == output['f'][i]) & (nu['j'] == output['j'][i]), 'value'].item().X
                v10 = ((output['f'][i] == 'SAF') & (output['j'][i] == 'NC')) * output['Statetaxheaviside'][i]* diracdeltaub_value.X
                v11 = biofuel_quantity_Q[i].X
                v12 = v1 - v2 + v3 + v4 + v5 + v6 + v7
                v13 =  - hefa_value[0].X * ((output['f'][i] == 'SAF') & (output['s'][i] == 'soyoil')) + hefa_value[0].X * 5.5/2.5 * ((output['f'][i] == 'RD') & (output['s'][i] == 'soyoil')) - hefa_value[1].X * ((output['f'][i] == 'SAF') & (output['s'][i] == 'animalfat')) + hefa_value[1].X * 5.5/2.5 * ((output['f'][i] == 'RD') & (output['s'][i] == 'animalfat'))
                total = v1 - v2 + v3 + v4 + v5 + v6 + v7 + v8  - v9 - v10 + v13
                # ... calculate v1, v2, ..., v11 ...
                results.append({
                    'i': i,
                    'Name': names[i],
                    'fuel': output['f'][i],
                    'state': output['j'][i],
                    'biofuel_price': v1,
                    'production_cost': v2,
                    'RIN': v3,
                    '45Z_tax': v4,
                    'SAF_certicate': v5,
                    'LCFS_CA': v6,
                    'state_tax': v7,
                    'mu': v8,
                    'nu': v9,
                    'dirac_delta': v10,
                    'hefa':v13,
                    'FOC': total,
                    'quantity': v11,
                    'break_idx': break_idx,
                    'break_val': break_val,
                    'val': val,
                    'profit':v12
                })
        
            # Save to spreadsheet for this break
            df_results = pd.DataFrame(results)
            #df_results = df_results.sort_values(by=['state','fuel'])
            #total_profit = (df_results['profit'] * df_results['quantity']).sum()
            total_profit = (df_results['profit'] * df_results['quantity']).sum() + state_tax_profit

            # for step function - state tax, total profit is not just marginal, but also need to include the previous fixed amount part
            
            # Update objective values to include total profit
            objective_values[-1]['Total_Profit'] = total_profit

            
            if total_profit > max_total_profit:
                max_total_profit = total_profit
                best_df = df_results.copy()
                best_break_idx = break_idx
                best_demand_quantity = demand_quantity.copy()
                best_model = m  # Save the best model
                best_val = val  # <-- Add this line
            
            # # Instead of comparing total_profit, update best result if solution exists
            # max_total_profit = total_profit
            # best_df = df_results.copy()
            # best_break_idx = break_idx
            # best_demand_quantity = demand_quantity.copy()
            # best_model = m  # Save the best model
            # best_val = val  # <-- Add this line

            #df_results.to_excel(f'results/terms_breakdown_break{break_idx}.xlsx', index=False)
            
            # Save demand_quantity DataFrame to spreadsheet for this break
            demand_quantity_out = demand_quantity.copy()
            demand_quantity_out['Q'] = [q.X if hasattr(q, 'X') else q for q in demand_quantity_out['Q']]
            demand_quantity_out['Q_actual'] = [q.X if hasattr(q, 'X') else q for q in demand_quantity_out['Q_actual']]
            #demand_quantity_out.to_excel(f'results/demand_quantity_break{break_idx}.xlsx', index=False)

        else:
            print(f"No solution for break {break_idx}, status: {m.status}")
            # Store result even for infeasible solutions
            objective_values.append({
                'Break': break_idx,
                'Break_Value': break_val,
                'Val': val,
                'Objective_Value': None if m.status != GRB.OPTIMAL else m.ObjVal,
                'Total_Profit': None
            })
            continue  # Skip to next break

    # After the loop
    if best_df is not None:
        best_df.to_excel(os.path.join(results_dir, 'terms_breakdown_best.xlsx'), index=False)
        print(f"Best break_idx: {best_break_idx}, max total profit: {max_total_profit}")

        # Save demand_quantity DataFrame for the best solution
        demand_quantity_out = best_demand_quantity.copy()
        demand_quantity_out['Q'] = [q.X if hasattr(q, 'X') else q for q in demand_quantity_out['Q']]
        demand_quantity_out['Q_actual'] = [q.X if hasattr(q, 'X') else q for q in demand_quantity_out['Q_actual']]
        demand_quantity_out.to_excel(os.path.join(results_dir, 'demand_quantity_best.xlsx'), index=False)
        
    # Save objective values to Excel file (regardless of whether best solution exists)
    if objective_values:
        objective_df = pd.DataFrame(objective_values)
        # Add a column to indicate which is the best solution
        if best_break_idx is not None:
            objective_df['Is_Best'] = objective_df['Break'] == best_break_idx
        else:
            objective_df['Is_Best'] = False
        
        objective_df.to_excel(os.path.join(results_dir, 'objective_values.xlsx'), index=False)
        print(f"Objective values saved to {os.path.join(results_dir, 'objective_values.xlsx')}")
        
        # Display summary
        print("\nObjective Values Summary:")
        for _, row in objective_df.iterrows():
            status = " (BEST)" if row['Is_Best'] else ""
            if row['Objective_Value'] is not None:
                profit_str = f", Total Profit: {row['Total_Profit']:.6f}" if row['Total_Profit'] is not None else ""
                print(f"Break {row['Break']}: Objective: {row['Objective_Value']:.6f}{profit_str}{status}")
            else:
                print(f"Break {row['Break']}: No solution{status}")
    
    if best_df is not None:
        # Print and save variable values for the best model
        variables_data = []
        
        with open(os.path.join(results_dir, 'variables_best.txt'), 'w') as f:
            for v in best_model.getVars():
                var_name = v.VarName
                var_value = v.X
                
                print(f"{var_name} {var_value:g}")
                f.write(f"{var_name} {var_value:g}\n")
                
                # Collect data for Excel file
                variables_data.append({
                    'Variable_Name': var_name,
                    'Value': var_value
                })
        
        # Create and save Excel file for variables
        variables_df = pd.DataFrame(variables_data)
        variables_df.to_excel(os.path.join(results_dir, 'variables_best.xlsx'), index=False)
        
        print(f"Best solution variables saved to {os.path.join(results_dir, 'variables_best.xlsx')}")

        # # Define the Intermediate_var DataFrame
        Intermediate_var = pd.DataFrame({
            'Name': ['B100-soyoil-CA', 'B100-animalfat-CA', 'RD-soyoil-CA', 'RD-animalfat-CA', 'SAF-animalfat-CA', 'gasE100-corn-CA',
                    'B100-soyoil-NC', 'B100-animalfat-NC', 'RD-soyoil-NC', 'RD-animalfat-NC', 'SAF-animalfat-NC', 'gasE100-corn-NC',
                    'gasE100-sugarcane-CA', 'gasE100-sugarcane-NC', 'SAF-soyoil-CA', 'SAF-soyoil-NC', 'SAF-sugarcane-ETJ-CA', 'SAF-sugarcane-ETJCCS-CA',
                    'SAF-corn-ETJ-CA', 'SAF-corn-ETJCCS-CA', 'SAF-sugarcane-ETJ-NC', 'SAF-sugarcane-ETJCCS-NC', 'SAF-corn-ETJ-NC', 'SAF-corn-ETJCCS-NC',
                    'B0_CA', 'E0_CA', 'J0_CA', 'B0_NC', 'E0_NC', 'J0_NC'],
            'RIN_obligation': [0]*30,
            'LCFS_SAF_obligation': [0]*30,
            'LCFS_CA_obligation': [0]*30,
            'Tax': [0]*30,
            'P_all': [0]*30,
            'PC_F0': [0]*30,
            'P_RIN_SAF_LCFS_detach': [0]*30,
            'P_RIN_SAF_detach': [0]*30,
            'Q_actual': [0]*30,
            'Q_equiv': [0]*30,
            'RFS_road': [0]*30,
            'RFS_jet': [0]*30,
            'LCFS_CA': [0]*30,
            'LCFS_SAF': [0]*30,
            'TC': [0]*30,
            'TC_IRA': [0]*30,
            'TC_45Z': [0]*30,
            'TC_45Q': [0]*30,
            'TC_state': [0]*30,
            'TC_SAF': [0]*30,
            'Subsidy': [0]*30,
            'Production_cost': [0]*30,
            'Fixed_production_cost': [0]*30,
            'Feedstock_cost': [0]*30,
            'mu_value': [0]*30,
            'nu_value': [0]*30,
            'etj_value': [0]*30,
            'foc': [0]*30
        })

        # Calculate the values for each row using best_model
        for i in range(24):
            # Intermediate_var.at[i, 'RFS_road'] = (
            #     (output['f'][i] in ['B100', 'RD', 'gasE100'])
            #     * output['kappa_RFS'][i]
            #     * best_model.getVarByName(RIN_prices.loc[RIN_prices['fueltype'].str.contains(output['f'][i]), 'P'].item().VarName).X
            # )
            # Intermediate_var.at[i, 'RFS_jet'] = (
            #     (output['f'][i] == 'SAF')
            #     * output['kappa_RFS'][i]
            #     * best_model.getVarByName(RIN_prices.loc[RIN_prices['fueltype'].str.contains(output['f'][i]), 'P'].item().VarName).X
            # )
            Intermediate_var.at[i, 'RFS_road'] = (
                (output['f'][i] in ['B100', 'RD', 'gasE100'])
                * (
                    output['kappa_RFS'][i] * best_model.getVarByName(RIN_prices.loc[RIN_prices['code'] == 'D4', 'P'].item().VarName).X * output['D4'][i] +
                    output['kappa_RFS'][i] * best_model.getVarByName(RIN_prices.loc[RIN_prices['code'] == 'D6', 'P'].item().VarName).X * output['D6'][i]
                )
            )
            Intermediate_var.at[i, 'RFS_jet'] = (
                (output['f'][i] == 'SAF')
                * (
                    output['kappa_RFS'][i] * best_model.getVarByName(RIN_prices.loc[RIN_prices['code'] == 'D4', 'P'].item().VarName).X * output['D4'][i] +
                    output['kappa_RFS'][i] * best_model.getVarByName(RIN_prices.loc[RIN_prices['code'] == 'D6', 'P'].item().VarName).X * output['D6'][i]
                )
                                            )
            Intermediate_var.at[i, 'LCFS_CA'] = (
                best_model.getVarByName(LCFS_prices['P'][0].VarName).X * output.iloc[i]['CALCFSSide']
                * (output['CI_std_LCFS_CA'][i] - output['CI_LCFS'][i])
                * (output['j'][i] == 'CA')
                * output['ED_LCFS'][i] / 1000000
            )
            Intermediate_var.at[i, 'LCFS_SAF'] = (
                best_model.getVarByName(LCFS_prices_SAF['P'][0].VarName).X * output.iloc[i]['SAFLCFSside']
                * (best_model.getVarByName(SAF_intensity_standard['P'][0].VarName).X - output['CI_LCFS'][i])
                * output['ED_LCFS'][i] / 1000000
            )
            # Intermediate_var.at[i, 'LCFS_SAF'] = (
            #     best_model.getVarByName(LCFS_prices_SAF['P'][0].VarName).X * output.iloc[i]['SAFLCFSside']
            #     * (best_model.getVarByName(output['CI_std_LCFS'][i].VarName).X if hasattr(output['CI_std_LCFS'][i], 'VarName') else output['CI_std_LCFS'][i] - output['CI_Tax'][i])
            #     * (output['j'][i] == 'CA')
            #     * output['ED_LCFS'][i] / 1000000
            # )
            Intermediate_var.at[i, 'TC'] = (
                 1 * (output['f'][i] in ['B100', 'RD', 'gasE100'])
                * ((47.39 - output['CI_Tax'][i]) / 47.39)
                * output['Taxheavyside'][i]* (output['s'][i] != 'sugarcane')
                + max(
                    1 * (output['f'][i] == 'SAF')
                    * ((47.39 - output['CI_Tax'][i]) / 47.39)
                    * output['Taxheavyside'][i],
                    85 * 30 * output['ED_LCFS'][i] / (10**6) * output['CCS_tech'][i]
                )* (output['s'][i] != 'sugarcane')
            ) + best_val * (output['f'][i] == 'SAF') * (output['j'][i] == 'NC') * output['Statetaxheaviside'][i] 
            #+ (output['f'][i] == 'SAF') * best_model.getVarByName(SAFcredit_prices['P'].item().VarName).X
            Intermediate_var.at[i, 'TC_IRA'] = (
                1 * (output['f'][i] in ['B100', 'RD', 'gasE100'])
                * ((47.39 - output['CI_Tax'][i]) / 47.39)
                * output['Taxheavyside'][i]* (output['s'][i] != 'sugarcane')
                + max(
                    1 * (output['f'][i] == 'SAF')
                    * ((47.39 - output['CI_Tax'][i]) / 47.39)
                    * output['Taxheavyside'][i],
                    85 * 30 * output['ED_LCFS'][i] / (10**6) * output['CCS_tech'][i]
                )* (output['s'][i] != 'sugarcane')
            )
            Intermediate_var.at[i, 'TC_45Z'] = (
                1 * (output['f'][i] in ['B100', 'RD', 'gasE100'])
                * ((47.39 - output['CI_Tax'][i]) / 47.39)
                * output['Taxheavyside'][i]* (output['s'][i] != 'sugarcane')
                + (
                    1 * (output['f'][i] == 'SAF')
                    * ((47.39 - output['CI_Tax'][i]) / 47.39)
                    * output['Taxheavyside'][i]
                    > 85 * 30 * output['ED_LCFS'][i] / (10**6) * output['CCS_tech'][i]
                )
                * 1 * (output['f'][i] == 'SAF')
                * ((47.39 - output['CI_Tax'][i]) / 47.39)
                * output['Taxheavyside'][i]* (output['s'][i] != 'sugarcane')
            )
            Intermediate_var.at[i, 'TC_45Q'] = (
                (
                    85 * 30 * output['ED_LCFS'][i] / (10**6) * output['CCS_tech'][i]
                    > 1 * (output['f'][i] == 'SAF')
                    * ((47.39 - output['CI_Tax'][i]) / 47.39)
                    * output['Taxheavyside'][i]
                )
                * 85 * 30 * output['ED_LCFS'][i] / (10**6) * output['CCS_tech'][i]* (output['s'][i] != 'sugarcane')
            )
            Intermediate_var.at[i, 'TC_state'] = best_val * (output['f'][i] == 'SAF') * (output['j'][i] == 'NC') * output['Statetaxheaviside'][i]
            # Intermediate_var.at[i, 'TC_SAF'] = (output['f'][i] == 'SAF') * best_model.getVarByName(SAFcredit_prices['P'].item().VarName).X
            Intermediate_var.at[i, 'TC_SAF'] = 0
            Intermediate_var.at[i, 'Subsidy'] = (
                Intermediate_var.at[i, 'RFS_road']
                + Intermediate_var.at[i, 'RFS_jet']
                + Intermediate_var.at[i, 'LCFS_CA']  + Intermediate_var.at[i, 'LCFS_SAF']
                + Intermediate_var.at[i, 'TC']
            )
            Intermediate_var.at[i, 'Production_cost'] = (
                output['PC'][i]
                + output['Conversion'][i]
                * best_model.getVarByName(feedstock_prices.loc[feedstock_prices['s'] == output['s'][i], 'P'].item().VarName).X*(1+duty_e100sugarcane*(output['s'][i] == 'sugarcane'))
            )
            Intermediate_var.at[i, 'Fixed_production_cost'] = output['PC'][i]
            Intermediate_var.at[i, 'Feedstock_cost'] = (
                output['Conversion'][i]
                * best_model.getVarByName(feedstock_prices.loc[feedstock_prices['s'] == output['s'][i], 'P'].item().VarName).X*(1+duty_e100sugarcane*(output['s'][i] == 'sugarcane'))
            )
            Intermediate_var.at[i, 'mu_value'] = best_model.getVarByName(mu['value'][i].VarName).X
            Intermediate_var.at[i, 'nu_value'] = best_model.getVarByName(
                nu.loc[(nu['f'] == output['f'][i]) & (nu['j'] == output['j'][i]), 'value'].item().VarName
            ).X
            Intermediate_var.at[i, 'etj_value'] =  ((output['f'][i] == 'SAF') & (output['j'][i] == 'NC')) * output['Statetaxheaviside'][i]* best_model.getVarByName('diracdeltaub_value').X
            Intermediate_var.at[i, 'foc'] = (
                best_model.getVarByName(biofuel_prices.loc[(biofuel_prices['f'] == output['f'][i]) & (biofuel_prices['j'] == output['j'][i]), 'P'].item().VarName).X
                - (output['PC'][i]
                + output['Conversion'][i]
                * best_model.getVarByName(feedstock_prices.loc[feedstock_prices['s'] == output['s'][i], 'P'].item().VarName).X*(1+duty_e100sugarcane*(output['s'][i] == 'sugarcane')))
                # + output['kappa_RFS'][i] * best_model.getVarByName(RIN_prices.loc[RIN_prices['fueltype'].str.contains(output['f'][i]), 'P'].item().VarName).X
                +  (
                    output['kappa_RFS'][i] * best_model.getVarByName(RIN_prices.loc[RIN_prices['code'] == 'D4', 'P'].item().VarName).X * output['D4'][i] +
                    output['kappa_RFS'][i] * best_model.getVarByName(RIN_prices.loc[RIN_prices['code'] == 'D6', 'P'].item().VarName).X * output['D6'][i]
                )
                + 1 * ((output['f'][i] == 'B100') | (output['f'][i] == 'RD') | (output['f'][i] == 'gasE100'))
                * ((47.39 - output['CI_Tax'][i]) / 47.39)
                * output['Taxheavyside'][i]* (output['s'][i] != 'sugarcane')
                + max([
                    1 * (output['f'][i] == 'SAF')
                    * ((47.39 - output['CI_Tax'][i]) / 47.39)
                    * output['Taxheavyside'][i],
                    85 * 30 * output['ED_LCFS'][i] / (10**6) * output['CCS_tech'][i]
                ])* (output['s'][i] != 'sugarcane')
                + best_model.getVarByName(LCFS_prices_SAF['P'][0].VarName).X * output.iloc[i]['SAFLCFSside']
                * (best_model.getVarByName(SAF_intensity_standard['P'][0].VarName).X - output['CI_LCFS'][i])
                * output['ED_LCFS'][i] / 1000000
                + best_model.getVarByName(LCFS_prices['P'][0].VarName).X * output.iloc[i]['CALCFSSide']
                * (output['CI_std_LCFS_CA'][i] - output['CI_LCFS'][i])
                * (output['j'][i] == 'CA')
                * output['ED_LCFS'][i] / 1000000
                + best_model.getVarByName(mu['value'][i].VarName).X
                - best_model.getVarByName(
                    nu.loc[(nu['f'] == output['f'][i]) & (nu['j'] == output['j'][i]), 'value'].item().VarName
                ).X
                + ((output['f'][i] == 'SAF') & (output['j'][i] == 'NC')) * output['Statetaxheaviside'][i] * best_val
                 - ((output['f'][i] == 'SAF') & (output['j'][i] == 'NC')) * output['Statetaxheaviside'][i]* best_model.getVarByName('diracdeltaub_value').X - best_model.getVarByName('hefa_value[0]').X * ((output['f'][i] == 'SAF') & (output['s'][i] == 'soyoil')) 
                + best_model.getVarByName('hefa_value[0]').X * 5.5/2.5 * ((output['f'][i] == 'RD') & (output['s'][i] == 'soyoil')) 
                - best_model.getVarByName('hefa_value[1]').X * ((output['f'][i] == 'SAF') & (output['s'][i] == 'animalfat')) 
                + best_model.getVarByName('hefa_value[1]').X * 5.5/2.5 * ((output['f'][i] == 'RD') & (output['s'][i] == 'animalfat'))
            )
        
        # # Assign values from the Price DataFrame to Intermediate_var
        # for col in ['RIN_obligation', 'LCFS_CA_obligation', 'Tax', 'P_all', 'PC_F0', 'P_RIN_SAF_LCFS_detach', 'P_RIN_detach', 'Q_actual', 'Q_equiv']:
        #     Intermediate_var[col] = [Price.at[i, col].X for i in Price.index]

        # # Assign values from the Price DataFrame to Intermediate_var
        # for col in ['RIN_obligation', 'LCFS_CA_obligation','LCFS_SAF_obligation']:
        #     Intermediate_var.loc[:23, col] = [Price.at[i, col].X for i in Price.index[:24]]
        #     Intermediate_var.loc[24:, col] = [-Price.at[i, col].X for i in Price.index[24:]]


        # # Assign values from the Price DataFrame to Intermediate_var
        # for col in ['Tax', 'P_all', 'PC_F0', 'P_RIN_SAF_LCFS_detach', 'P_RIN_SAF_detach', 'Q_actual', 'Q_equiv']:
        #     Intermediate_var[col] = [Price.at[i, col].X for i in Price.index] 
        # 
        # Assign values from the Price DataFrame to Intermediate_var using best_model
        for col in ['RIN_obligation', 'LCFS_CA_obligation','LCFS_SAF_obligation']:
            Intermediate_var.loc[:23, col] = [
                best_model.getVarByName(Price.at[i, col].VarName).X for i in Price.index[:24]
            ]
            Intermediate_var.loc[24:, col] = [
                -best_model.getVarByName(Price.at[i, col].VarName).X for i in Price.index[24:]
            ]

        for col in ['Tax', 'P_all', 'PC_F0', 'P_RIN_SAF_LCFS_detach', 'P_RIN_SAF_detach', 'Q_actual', 'Q_equiv']:
            Intermediate_var[col] = [
                best_model.getVarByName(Price.at[i, col].VarName).X for i in Price.index
            ]   

        from openpyxl import load_workbook

        output_dir1 = os.path.join(results_dir, 'Intermediate_results')

        # Ensure the output directory exists
        os.makedirs(output_dir1, exist_ok=True)

        # # Function to append data to an existing Excel file
        # def append_to_excel(file_path, df, sheet_name):
        #     try:
        #         # Load the existing workbook
        #         with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        #             # Write the DataFrame to the specified sheet
        #             df.to_excel(writer, index=False, sheet_name=sheet_name)
        #     except FileNotFoundError:
        #         # If the file does not exist, create a new one
        #         with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        #             df.to_excel(writer, index=False, sheet_name=sheet_name)

        # Function to append data to an existing Excel file or create new one
        def append_to_excel(file_path, df, sheet_name):
            """
            Append DataFrame to existing Excel file or create new one if it doesn't exist
            """
            if os.path.exists(file_path):
                # File exists, check if sheet exists and append/replace accordingly
                try:
                    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df.to_excel(writer, index=False, sheet_name=sheet_name)
                except Exception as e:
                    print(f"Error appending to {file_path}: {e}")
                    # If there's an issue, create a new file
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name=sheet_name)
            else:
                # File doesn't exist, create new one
                try:
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name=sheet_name)
                except Exception as e:
                    print(f"Error creating {file_path}: {e}")

        if k == 0:
            append_to_excel(os.path.join(output_dir1, 'Mean.xlsx'), Intermediate_var, 'Aviation_intensity_standard')

        if k == 1:
            append_to_excel(os.path.join(output_dir1, '10th.xlsx'), Intermediate_var, 'Aviation_intensity_standard')

        if k == 2:
            append_to_excel(os.path.join(output_dir1, '33th.xlsx'), Intermediate_var, 'Aviation_intensity_standard')

        if k == 3:
            append_to_excel(os.path.join(output_dir1, '67th.xlsx'), Intermediate_var, 'Aviation_intensity_standard') 

        if k == 4:
            append_to_excel(os.path.join(output_dir1, '90th.xlsx'), Intermediate_var, 'Aviation_intensity_standard')
                               
                                    

        # # Read the external file into a DataFrame
        # Intermediate_input = pd.read_excel(
        #     '/Users/Mandywu/Dropbox/Aviation-SI project/Spreadsheets/Policy_projection/code/main_paper/code/Input_intermediate.xlsx'
        # )

        # New automatic path resolution - look in intermediate folder
        # Pure_quantity uses V1 version (not V2)
        intermediate_input_path = os.path.join(parent_dir, 'intermediate', 'Input_intermediate_V1.xlsx')
        
        # Verify the file exists with fallback options
        if not os.path.exists(intermediate_input_path):
            # Try V2 version as backup
            intermediate_input_backup = os.path.join(parent_dir, 'intermediate', 'Input_intermediate_V1.xlsx')
            if os.path.exists(intermediate_input_backup):
                intermediate_input_path = intermediate_input_backup
                print(f"Using Input_intermediate_V1.xlsx from intermediate directory: {intermediate_input_path}")
            else:
                raise FileNotFoundError(f"Input_intermediate.xlsx not found. Tried:\n1. {intermediate_input_path}\n2. {intermediate_input_backup}")
        else:
            print(f"Using Input_intermediate_V1.xlsx: {intermediate_input_path}")
        
        # Read the external file into a DataFrame
        Intermediate_input = pd.read_excel(intermediate_input_path)

        
        # Join the existing Intermediate_var with the new data
        Intermediate_var = Intermediate_var.join(Intermediate_input)

        # Update the solution DataFrame using best_model
        solution.iloc[:, k+1] = (
            [0]
            + [best_model.getVarByName(biofuel_quantity_Q[i].VarName).X for i in range(24)]
            + [0]
            + [best_val]
            + [0]
            #+ [best_model.getVarByName(SAFcredit_prices_P.VarName).X]
            + [0]
            + [best_model.getVarByName(RIN_prices_P[i].VarName).X for i in range(2)]
            + [best_model.getVarByName(LCFS_prices_P.VarName).X]
            + [best_model.getVarByName(LCFS_prices_SAF_P.VarName).X]
            + [best_model.getVarByName(biofuel_prices_P[i].VarName).X for i in range(8)]
            + [best_model.getVarByName(conventionalfuel_prices_P[i].VarName).X for i in range(6)]
            + [0]
            + [best_model.getVarByName(feedstock_prices_P[i].VarName).X for i in range(4)]
            + [best_model.getVarByName(feedstock_quantity_Q[i].VarName).X for i in range(4)]
            + [0]
            + [best_model.getVarByName(conventionalfuel_quantity_Q[i].VarName).X for i in range(6)]
            + [0]
            + [best_model.getVarByName(demand_price_P[i].VarName).X for i in range(6)]
            + [best_model.getVarByName(SAF_intensity_standard_P.VarName).X]
            + [0]
            + [best_model.getVarByName(biofuel_prices_P[i].VarName).X/input_biofuel.loc[input_biofuel['f'] == ['B100', 'RD', 'SAF', 'gasE100', 'B100', 'RD', 'SAF', 'gasE100'][i], 'phi'].unique()[0] for i in range(8)]            
        )

        # Helper function to get .X from a Gurobi variable or a pandas Series of variables
        def get_x(val):
            if hasattr(val, 'VarName'):
                return best_model.getVarByName(val.VarName).X
            elif hasattr(val, 'X'):
                return val.X
            elif hasattr(val, '__iter__'):
                return sum(get_x(v) for v in val)
            else:
                return val
            
        biofuel_part = 0
        for idx, row in biofuel_quantity.iterrows():
            q_val = best_model.getVarByName(row['Q'].VarName).X
            ci_std_val = best_model.getVarByName(row['CI_std_LCFS'].VarName).X if hasattr(row['CI_std_LCFS'], 'VarName') else row['CI_std_LCFS']
            ci_tax_val = row['CI_LCFS']
            ed_val = row['ED_LCFS']  
            saf_side_val = row['SAFLCFSside']
            
            biofuel_part += q_val * (ci_std_val - ci_tax_val) * ed_val * saf_side_val
        
        # Second part: conventional jet fuel contribution
        conventional_part = 0
        jet_fuel_rows = conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'J']
        for idx, row in jet_fuel_rows.iterrows():
            q_val = best_model.getVarByName(row['Q'].VarName).X
            ci_std_val = best_model.getVarByName(row['CI_std_LCFS'].VarName).X  # This should be SAF_intensity_standard_P
            ci_lcfs_val = row['CI_LCFS']
            ed_val = row['ED_LCFS']
            
            conventional_part += q_val * (-ci_std_val + ci_lcfs_val) * ed_val
        
        DeltaCI_SAF_value = biofuel_part - conventional_part
        
        # print(f"DeltaCI_SAF_value manual calculation: {DeltaCI_SAF_value}")
        # print(f"DeltaCI_SAF_expr.getValue(): {DeltaCI_SAF_expr.getValue()}")
        # print(f"Biofuel part: {biofuel_part}")
        # print(f"Conventional part: {conventional_part}")
    

        Fitted_quantity.iloc[1:20, k+1] = [
            best_model.getVarByName(RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].values[0].VarName).X,
            best_model.getVarByName(RVO_percent.loc[RVO_percent['mandate'] == 'RF', 'P'].values[0].VarName).X,
            best_model.getVarByName(RVO_percent.loc[RVO_percent['mandate'] == 'BBD', 'P'].values[0].VarName).X *
                get_x(conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'].isin(['G', 'D'])), 'Q']) / 1.6,
            best_model.getVarByName(RVO_percent.loc[RVO_percent['mandate'] == 'RF', 'P'].values[0].VarName).X *
                get_x(conventionalfuel_quantity.loc[(conventionalfuel_quantity['F'].isin(['G', 'D'])), 'Q']),
            #get_x(biofuel_quantity.loc[biofuel_quantity['f'].isin(['B100', 'RD', 'SAF']), 'Q']),
            get_x(biofuel_quantity.loc[biofuel_quantity['D4'] == 1, 'Q']),
            sum(
                biofuel_quantity.loc[biofuel_quantity['f'] == 'B100', 'kappa_RFS'].values[i] *
                best_model.getVarByName(q.VarName).X
                for i, q in enumerate(biofuel_quantity.loc[biofuel_quantity['f'] == 'B100', 'Q'])
            ) +
            sum(
                biofuel_quantity.loc[biofuel_quantity['f'] == 'RD', 'kappa_RFS'].values[i] *
                best_model.getVarByName(q.VarName).X
                for i, q in enumerate(biofuel_quantity.loc[biofuel_quantity['f'] == 'RD', 'Q'])
            ) +
            sum(
                biofuel_quantity.loc[biofuel_quantity['f'] == 'SAF', 'kappa_RFS'].values[i] *
                best_model.getVarByName(q.VarName).X
                for i, q in enumerate(biofuel_quantity.loc[biofuel_quantity['f'] == 'SAF', 'Q'])
            ) +
            sum(
                biofuel_quantity.loc[biofuel_quantity['f'] == 'gasE100', 'kappa_RFS'].values[i] *
                best_model.getVarByName(q.VarName).X
                for i, q in enumerate(biofuel_quantity.loc[biofuel_quantity['f'] == 'gasE100', 'Q'])
            ),
            -(
                sum(
                    best_model.getVarByName(q.VarName).X *
                    (ci_std - ci) *
                    ed / 1e6
                    for q, ci_std, ci, ed in zip(
                        biofuel_quantity.loc[biofuel_quantity['j'] == 'CA', 'Q'],
                        biofuel_quantity.loc[biofuel_quantity['j'] == 'CA', 'CI_std_LCFS_CA'],
                        biofuel_quantity.loc[biofuel_quantity['j'] == 'CA', 'CI_LCFS'],
                        biofuel_quantity.loc[biofuel_quantity['j'] == 'CA', 'ED_LCFS']
                    )
                )
                + sum(input_compliance.loc[input_compliance['mandate'] == 'LCFS_Q', 'Q'])
                - sum(
                    best_model.getVarByName(q.VarName).X *
                    (-ci_std + ci) *
                    ed / 1e6
                    for q, ci_std, ci, ed in zip(
                        conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'D'), 'Q'],
                        conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'D'), 'CI_std_LCFS_CA'],
                        conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'D'), 'CI_LCFS'],
                        conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'D'), 'ED_LCFS']
                    )
                )
                - sum(
                    best_model.getVarByName(q.VarName).X *
                    (-ci_std + ci) *
                    ed / 1e6
                    for q, ci_std, ci, ed in zip(
                        conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'G'), 'Q'],
                        conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'G'), 'CI_std_LCFS_CA'],
                        conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'G'), 'CI_LCFS'],
                        conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'G'), 'ED_LCFS']
                    )
                )
            ),
            -DeltaCI_SAF_value, 0,
            get_x(biofuel_quantity.loc[biofuel_quantity['f'] == 'B100', 'Q']),
            get_x(biofuel_quantity.loc[biofuel_quantity['f'] == 'RD', 'Q']),
            get_x(biofuel_quantity.loc[biofuel_quantity['f'].isin(['B100', 'RD']), 'Q']),
            get_x(conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'D', 'Q']),
            get_x(biofuel_quantity.loc[biofuel_quantity['f'] == 'gasE100', 'Q']),
            get_x(conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'G', 'Q']),
            get_x(biofuel_quantity.loc[(biofuel_quantity['f'] == 'SAF') & (biofuel_quantity['s'].isin(['soyoil', 'animalfat'])), 'Q']),
            get_x(biofuel_quantity.loc[(biofuel_quantity['f'] == 'SAF') & (biofuel_quantity['s'].isin(['corn', 'sugarcane'])), 'Q']),
            get_x(biofuel_quantity.loc[biofuel_quantity['f'] == 'SAF', 'Q']),
            get_x(conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'J', 'Q'])
        ]

        Fitted_quantity.iloc[20:28, k+1] = [
            0,
            get_x(biofuel_quantity.loc[(biofuel_quantity['j'] == 'CA') & (biofuel_quantity['f'] == 'B100'), 'Q']),
            get_x(biofuel_quantity.loc[(biofuel_quantity['j'] == 'CA') & (biofuel_quantity['f'] == 'RD'), 'Q']),
            get_x(conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'D'), 'Q']),
            get_x(biofuel_quantity.loc[(biofuel_quantity['j'] == 'CA') & (biofuel_quantity['f'] == 'gasE100'), 'Q']),
            get_x(conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'G'), 'Q']),
            get_x(biofuel_quantity.loc[(biofuel_quantity['j'] == 'CA') & (biofuel_quantity['f'] == 'SAF'), 'Q']),
            get_x(conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'J'), 'Q'])
        ]

        Fitted_quantity.iloc[29:33, k+1] = [
            best_model.getVarByName(feedstock_quantity_Q[i].VarName).X for i in range(4)
        ]
        Fitted_quantity.iloc[34:40, k+1] = [
            best_model.getVarByName(demand_quantity_Q[i].VarName).X for i in range(6)
        ]
        
        # Calculate and assign values to Fuel_price DataFrame
        Fuel_price.iloc[1:9, k+1] = [
            sum(Intermediate_var.loc[(Intermediate_var['j'] == 'CA') & (Intermediate_var['F'] == 'D'), 'P_RIN_SAF_LCFS_detach'
                ]*Intermediate_var.loc[(Intermediate_var['j'] == 'CA') & (Intermediate_var['F'] == 'D'), 'Q_actual']) /
            sum(Intermediate_var.loc[(Intermediate_var['j'] == 'CA') & (Intermediate_var['F'] == 'D'), 'Q_equiv']), 
            sum(Intermediate_var.loc[(Intermediate_var['j'] == 'NC') & (Intermediate_var['F'] == 'D'), 'P_RIN_SAF_LCFS_detach'] *
                Intermediate_var.loc[(Intermediate_var['j'] == 'NC') & (Intermediate_var['F'] == 'D'), 'Q_actual']) /
            sum(Intermediate_var.loc[(Intermediate_var['j'] == 'NC') & (Intermediate_var['F'] == 'D'), 'Q_equiv']),
        
            sum(Intermediate_var.loc[(Intermediate_var['j'] == 'CA') & (Intermediate_var['F'] == 'G'), 'P_RIN_SAF_LCFS_detach'
                ] *Intermediate_var.loc[(Intermediate_var['j'] == 'CA') & (Intermediate_var['F'] == 'G'), 'Q_actual']) /
            sum(Intermediate_var.loc[(Intermediate_var['j'] == 'CA') & (Intermediate_var['F'] == 'G'), 'Q_equiv']), 
            sum(Intermediate_var.loc[(Intermediate_var['j'] == 'NC') & (Intermediate_var['F'] == 'G'), 'P_RIN_SAF_LCFS_detach'] *
                Intermediate_var.loc[(Intermediate_var['j'] == 'NC') & (Intermediate_var['F'] == 'G'), 'Q_actual']) /
            sum(Intermediate_var.loc[(Intermediate_var['j'] == 'NC') & (Intermediate_var['F'] == 'G'), 'Q_equiv']),

            sum(Intermediate_var.loc[(Intermediate_var['j'] == 'CA') & (Intermediate_var['F'] == 'J'), 'P_RIN_SAF_LCFS_detach'
                ] *Intermediate_var.loc[(Intermediate_var['j'] == 'CA') & (Intermediate_var['F'] == 'J'), 'Q_actual']) /
            sum(Intermediate_var.loc[(Intermediate_var['j'] == 'CA') & (Intermediate_var['F'] == 'J'), 'Q_equiv']), 
            sum(Intermediate_var.loc[(Intermediate_var['j'] == 'NC') & (Intermediate_var['F'] == 'J'), 'P_RIN_SAF_LCFS_detach'] *
                Intermediate_var.loc[(Intermediate_var['j'] == 'NC') & (Intermediate_var['F'] == 'J'), 'Q_actual']) /
            sum(Intermediate_var.loc[(Intermediate_var['j'] == 'NC') & (Intermediate_var['F'] == 'J'), 'Q_equiv']),
            sum(Intermediate_var.loc[(Intermediate_var['j'] == 'CA') & (Intermediate_var['f'] == 'gasE100'), 'P_RIN_SAF_LCFS_detach'] *
                Intermediate_var.loc[(Intermediate_var['j'] == 'CA') & (Intermediate_var['f'] == 'gasE100'), 'Q_actual']) /
            sum(Intermediate_var.loc[(Intermediate_var['j'] == 'CA') & (Intermediate_var['f'] == 'gasE100'), 'Q_actual']),
            sum(Intermediate_var.loc[(Intermediate_var['j'] == 'NC') & (Intermediate_var['f'] == 'gasE100'), 'P_RIN_SAF_LCFS_detach'] *
                Intermediate_var.loc[(Intermediate_var['j'] == 'NC') & (Intermediate_var['f'] == 'gasE100'), 'Q_actual']) /
            sum(Intermediate_var.loc[(Intermediate_var['j'] == 'NC') & (Intermediate_var['f'] == 'gasE100'), 'Q_actual'])
        ]

        Fuel_price.iloc[9, k+1] = ((Fuel_price.iloc[7, k+1] * 0.1 + Intermediate_var.loc[(Intermediate_var['j'] == 'CA') & (Intermediate_var['f'] == 'E0'), 'P_RIN_SAF_LCFS_detach'].values[0] * 0.9) /(0.1 * input_biofuel.loc[input_biofuel['f'] == 'gasE100', 'phi_ed'].unique()[0] + 0.9))
        
        Fuel_price.iloc[10, k+1] = ((Fuel_price.iloc[8, k+1] * 0.1 + Intermediate_var.loc[(Intermediate_var['j'] == 'NC') & (Intermediate_var['f'] == 'E0'), 'P_RIN_SAF_LCFS_detach'].values[0] * 0.9) /(0.1 * input_biofuel.loc[input_biofuel['f'] == 'gasE100', 'phi_ed'].unique()[0] + 0.9))

        Fuel_price.iloc[12:16, k+1] = [
            sum(Intermediate_var.loc[Intermediate_var['f'] == 'B100', 'P_all'] *Intermediate_var.loc[Intermediate_var['f'] == 'B100', 'Q_actual']) /sum(Intermediate_var.loc[Intermediate_var['f'] == 'B100', 'Q_actual']) if sum(Intermediate_var.loc[Intermediate_var['f'] == 'B100', 'Q_actual']) != 0 else np.nan,   
            sum(Intermediate_var.loc[Intermediate_var['f'] == 'RD', 'P_all'] * Intermediate_var.loc[Intermediate_var['f'] == 'RD', 'Q_actual']) /sum(Intermediate_var.loc[Intermediate_var['f'] == 'RD', 'Q_actual']),  
            sum(Intermediate_var.loc[Intermediate_var['f'] == 'gasE100', 'P_all'] *Intermediate_var.loc[Intermediate_var['f'] == 'gasE100', 'Q_actual'])/sum(Intermediate_var.loc[Intermediate_var['f'] == 'gasE100', 'Q_actual']),
            sum(Intermediate_var.loc[Intermediate_var['f'] == 'SAF', 'P_all'] *Intermediate_var.loc[Intermediate_var['f'] == 'SAF', 'Q_actual']) /sum(Intermediate_var.loc[Intermediate_var['f'] == 'SAF', 'Q_actual'])
        ]

        # Example for conventionalfuel_prices['P']
        Fuel_price.iloc[17:23, k+1] = [
            best_model.getVarByName(conventionalfuel_prices_P[i].VarName).X for i in range(6)
        ]

        Fuel_price.iloc[24, k+1] = [
            best_model.getVarByName(RVO_percent_P[0].VarName).X * best_model.getVarByName(RIN_prices_P[0].VarName).X +
            (best_model.getVarByName(RVO_percent_P[1].VarName).X - best_model.getVarByName(RVO_percent_P[0].VarName).X) * best_model.getVarByName(RIN_prices_P[1].VarName).X
        ]

        Fuel_price.iloc[26, k+1] = [
            best_model.getVarByName(LCFS_prices_P.VarName).X * (
                -conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'D'), 'CI_std_LCFS_CA'].item() +
                conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'D'), 'CI_LCFS'].item()
            ) * conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'D'), 'ED_LCFS'].item() / 1000000
        ]

        Fuel_price.iloc[27, k+1] = [
            best_model.getVarByName(LCFS_prices_P.VarName).X * (
                -conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'G'), 'CI_std_LCFS_CA'].item() +
                conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'G'), 'CI_LCFS'].item()
            ) * conventionalfuel_quantity.loc[(conventionalfuel_quantity['j'] == 'CA') & (conventionalfuel_quantity['F'] == 'G'), 'ED_LCFS'].item() / 1000000
        ]

        Fuel_price.iloc[28, k+1] = [
            best_model.getVarByName(LCFS_prices_SAF.loc['LCFS_SAF', 'P'].VarName).X * 
            (-best_model.getVarByName(SAF_intensity_standard.loc['SAF', 'P'].VarName).X + 
            conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'J', 'CI_LCFS'].unique()[0]) * 
            conventionalfuel_quantity.loc[conventionalfuel_quantity['F'] == 'J', 'ED_LCFS'].unique()[0] / 1000000
        ]

        Fuel_price.iloc[30:34, k+1] = [
            sum(Intermediate_var.loc[Intermediate_var['f'] == 'B100', 'P_all'] *Intermediate_var.loc[Intermediate_var['f'] == 'B100', 'Q_actual']) /sum(Intermediate_var.loc[Intermediate_var['f'] == 'B100', 'Q_equiv']) if sum(Intermediate_var.loc[Intermediate_var['f'] == 'B100', 'Q_equiv']) != 0 else np.nan,   
            sum(Intermediate_var.loc[Intermediate_var['f'] == 'RD', 'P_all'] * Intermediate_var.loc[Intermediate_var['f'] == 'RD', 'Q_actual']) /sum(Intermediate_var.loc[Intermediate_var['f'] == 'RD', 'Q_equiv']),  
            sum(Intermediate_var.loc[Intermediate_var['f'] == 'gasE100', 'P_all'] *Intermediate_var.loc[Intermediate_var['f'] == 'gasE100', 'Q_actual'])/sum(Intermediate_var.loc[Intermediate_var['f'] == 'gasE100', 'Q_equiv']),
            sum(Intermediate_var.loc[Intermediate_var['f'] == 'SAF', 'P_all'] *Intermediate_var.loc[Intermediate_var['f'] == 'SAF', 'Q_actual']) /sum(Intermediate_var.loc[Intermediate_var['f'] == 'SAF', 'Q_equiv'])
        ]


        # # Calculate and assign values to Fuel_CI DataFrame
        # Fuel_CI.iloc[1:5, k+1] = [
        #     sum(Intermediate_var['CI_Tax'] * Intermediate_var['Q_actual']) / sum(Intermediate_var['Q_actual']),
        #     sum(Intermediate_var.loc[Intermediate_var['F'] == 'D', 'CI_Tax'] * Intermediate_var.loc[Intermediate_var['F'] == 'D', 'Q_actual']) /
        #     sum(Intermediate_var.loc[Intermediate_var['F'] == 'D', 'Q_actual']),
        #     sum(Intermediate_var.loc[Intermediate_var['F'] == 'G', 'CI_Tax'] * Intermediate_var.loc[Intermediate_var['F'] == 'G', 'Q_actual']) /
        #     sum(Intermediate_var.loc[Intermediate_var['F'] == 'G', 'Q_actual']),
        #     sum(Intermediate_var.loc[Intermediate_var['F'] == 'J', 'CI_Tax'] * Intermediate_var.loc[Intermediate_var['F'] == 'J', 'Q_actual']) /
        #     sum(Intermediate_var.loc[Intermediate_var['F'] == 'J', 'Q_actual'])
        # ]
        
        # # Rows 7 to 10 (Blended fuel in CA, Blended-diesel in CA, Blended-gasoline in CA, Blended-jet in CA)
        # Fuel_CI.iloc[6:10, k+1] = [
        #     sum(Intermediate_var.loc[Intermediate_var['j'] == 'CA', 'CI_Tax'] * Intermediate_var.loc[Intermediate_var['j'] == 'CA', 'Q_actual']) /
        #     sum(Intermediate_var.loc[Intermediate_var['j'] == 'CA', 'Q_actual']),
        #     sum(Intermediate_var.loc[(Intermediate_var['F'] == 'D') & (Intermediate_var['j'] == 'CA'), 'CI_Tax'] *
        #         Intermediate_var.loc[(Intermediate_var['F'] == 'D') & (Intermediate_var['j'] == 'CA'), 'Q_actual']) /
        #     sum(Intermediate_var.loc[(Intermediate_var['F'] == 'D') & (Intermediate_var['j'] == 'CA'), 'Q_actual']),
        #     sum(Intermediate_var.loc[(Intermediate_var['F'] == 'G') & (Intermediate_var['j'] == 'CA'), 'CI_Tax'] *
        #         Intermediate_var.loc[(Intermediate_var['F'] == 'G') & (Intermediate_var['j'] == 'CA'), 'Q_actual']) /
        #     sum(Intermediate_var.loc[(Intermediate_var['F'] == 'G') & (Intermediate_var['j'] == 'CA'), 'Q_actual']),
        #     sum(Intermediate_var.loc[(Intermediate_var['F'] == 'J') & (Intermediate_var['j'] == 'CA'), 'CI_Tax'] *
        #         Intermediate_var.loc[(Intermediate_var['F'] == 'J') & (Intermediate_var['j'] == 'CA'), 'Q_actual']) /
        #     sum(Intermediate_var.loc[(Intermediate_var['F'] == 'J') & (Intermediate_var['j'] == 'CA'), 'Q_actual'])
        # ]
        
        # # Rows 12 to 15 (Blended fuel in NC, Blended-diesel in NC, Blended-gasoline in NC, Blended-jet in NC)
        # Fuel_CI.iloc[11:15, k+1] = [
        #     sum(Intermediate_var.loc[Intermediate_var['j'] == 'NC', 'CI_Tax'] * Intermediate_var.loc[Intermediate_var['j'] == 'NC', 'Q_actual']) /
        #     sum(Intermediate_var.loc[Intermediate_var['j'] == 'NC', 'Q_actual']),
        #     sum(Intermediate_var.loc[(Intermediate_var['F'] == 'D') & (Intermediate_var['j'] == 'NC'), 'CI_Tax'] *
        #         Intermediate_var.loc[(Intermediate_var['F'] == 'D') & (Intermediate_var['j'] == 'NC'), 'Q_actual']) /
        #     sum(Intermediate_var.loc[(Intermediate_var['F'] == 'D') & (Intermediate_var['j'] == 'NC'), 'Q_actual']),
        #     sum(Intermediate_var.loc[(Intermediate_var['F'] == 'G') & (Intermediate_var['j'] == 'NC'), 'CI_Tax'] *
        #         Intermediate_var.loc[(Intermediate_var['F'] == 'G') & (Intermediate_var['j'] == 'NC'), 'Q_actual']) /
        #     sum(Intermediate_var.loc[(Intermediate_var['F'] == 'G') & (Intermediate_var['j'] == 'NC'), 'Q_actual']),
        #     sum(Intermediate_var.loc[(Intermediate_var['F'] == 'J') & (Intermediate_var['j'] == 'NC'), 'CI_Tax'] *
        #         Intermediate_var.loc[(Intermediate_var['F'] == 'J') & (Intermediate_var['j'] == 'NC'), 'Q_actual']) /
        #     sum(Intermediate_var.loc[(Intermediate_var['F'] == 'J') & (Intermediate_var['j'] == 'NC'), 'Q_actual'])
        # ]

        # Calculate and assign values to Fuel_CI DataFrame
        Fuel_CI.iloc[1:5, k+1] = [
            sum(Intermediate_var['CI_Emissions'] * Intermediate_var['Q_actual']) / sum(Intermediate_var['Q_actual']),
            sum(Intermediate_var.loc[Intermediate_var['F'] == 'D', 'CI_Emissions'] * Intermediate_var.loc[Intermediate_var['F'] == 'D', 'Q_actual']) /
            sum(Intermediate_var.loc[Intermediate_var['F'] == 'D', 'Q_actual']),
            sum(Intermediate_var.loc[Intermediate_var['F'] == 'G', 'CI_Emissions'] * Intermediate_var.loc[Intermediate_var['F'] == 'G', 'Q_actual']) /
            sum(Intermediate_var.loc[Intermediate_var['F'] == 'G', 'Q_actual']),
            sum(Intermediate_var.loc[Intermediate_var['F'] == 'J', 'CI_Emissions'] * Intermediate_var.loc[Intermediate_var['F'] == 'J', 'Q_actual']) /
            sum(Intermediate_var.loc[Intermediate_var['F'] == 'J', 'Q_actual'])
        ]
        
        # Rows 7 to 10 (Blended fuel in CA, Blended-diesel in CA, Blended-gasoline in CA, Blended-jet in CA)
        Fuel_CI.iloc[6:10, k+1] = [
            sum(Intermediate_var.loc[Intermediate_var['j'] == 'CA', 'CI_Emissions'] * Intermediate_var.loc[Intermediate_var['j'] == 'CA', 'Q_actual']) /
            sum(Intermediate_var.loc[Intermediate_var['j'] == 'CA', 'Q_actual']),
            sum(Intermediate_var.loc[(Intermediate_var['F'] == 'D') & (Intermediate_var['j'] == 'CA'), 'CI_Emissions'] *
                Intermediate_var.loc[(Intermediate_var['F'] == 'D') & (Intermediate_var['j'] == 'CA'), 'Q_actual']) /
            sum(Intermediate_var.loc[(Intermediate_var['F'] == 'D') & (Intermediate_var['j'] == 'CA'), 'Q_actual']),
            sum(Intermediate_var.loc[(Intermediate_var['F'] == 'G') & (Intermediate_var['j'] == 'CA'), 'CI_Emissions'] *
                Intermediate_var.loc[(Intermediate_var['F'] == 'G') & (Intermediate_var['j'] == 'CA'), 'Q_actual']) /
            sum(Intermediate_var.loc[(Intermediate_var['F'] == 'G') & (Intermediate_var['j'] == 'CA'), 'Q_actual']),
            sum(Intermediate_var.loc[(Intermediate_var['F'] == 'J') & (Intermediate_var['j'] == 'CA'), 'CI_Emissions'] *
                Intermediate_var.loc[(Intermediate_var['F'] == 'J') & (Intermediate_var['j'] == 'CA'), 'Q_actual']) /
            sum(Intermediate_var.loc[(Intermediate_var['F'] == 'J') & (Intermediate_var['j'] == 'CA'), 'Q_actual'])
        ]
        
        # Rows 12 to 15 (Blended fuel in NC, Blended-diesel in NC, Blended-gasoline in NC, Blended-jet in NC)
        Fuel_CI.iloc[11:15, k+1] = [
            sum(Intermediate_var.loc[Intermediate_var['j'] == 'NC', 'CI_Emissions'] * Intermediate_var.loc[Intermediate_var['j'] == 'NC', 'Q_actual']) /
            sum(Intermediate_var.loc[Intermediate_var['j'] == 'NC', 'Q_actual']),
            sum(Intermediate_var.loc[(Intermediate_var['F'] == 'D') & (Intermediate_var['j'] == 'NC'), 'CI_Emissions'] *
                Intermediate_var.loc[(Intermediate_var['F'] == 'D') & (Intermediate_var['j'] == 'NC'), 'Q_actual']) /
            sum(Intermediate_var.loc[(Intermediate_var['F'] == 'D') & (Intermediate_var['j'] == 'NC'), 'Q_actual']),
            sum(Intermediate_var.loc[(Intermediate_var['F'] == 'G') & (Intermediate_var['j'] == 'NC'), 'CI_Emissions'] *
                Intermediate_var.loc[(Intermediate_var['F'] == 'G') & (Intermediate_var['j'] == 'NC'), 'Q_actual']) /
            sum(Intermediate_var.loc[(Intermediate_var['F'] == 'G') & (Intermediate_var['j'] == 'NC'), 'Q_actual']),
            sum(Intermediate_var.loc[(Intermediate_var['F'] == 'J') & (Intermediate_var['j'] == 'NC'), 'CI_Emissions'] *
                Intermediate_var.loc[(Intermediate_var['F'] == 'J') & (Intermediate_var['j'] == 'NC'), 'Q_actual']) /
            sum(Intermediate_var.loc[(Intermediate_var['F'] == 'J') & (Intermediate_var['j'] == 'NC'), 'Q_actual'])
        ]



        # # Rows 2 to 5 (Emissions for D, G, J, and total)
        # Total.iloc[1:5, k+1] = [
        #     sum(Intermediate_var.loc[Intermediate_var['F'] == 'D', 'Q_actual'] *
        #         Intermediate_var.loc[Intermediate_var['F'] == 'D', 'CI_Tax'] *
        #         Intermediate_var.loc[Intermediate_var['F'] == 'D', 'ED_LCFS']) / 10**6,
        #     sum(Intermediate_var.loc[Intermediate_var['F'] == 'G', 'Q_actual'] *
        #         Intermediate_var.loc[Intermediate_var['F'] == 'G', 'CI_Tax'] *
        #         Intermediate_var.loc[Intermediate_var['F'] == 'G', 'ED_LCFS']) / 10**6,
        #     sum(Intermediate_var.loc[Intermediate_var['F'] == 'J', 'Q_actual'] *
        #         Intermediate_var.loc[Intermediate_var['F'] == 'J', 'CI_Tax'] *
        #         Intermediate_var.loc[Intermediate_var['F'] == 'J', 'ED_LCFS']) / 10**6,
        #     sum(Intermediate_var['Q_actual'] * Intermediate_var['CI_Tax'] * Intermediate_var['ED_LCFS']) / 10**6
        # ]

        Total.iloc[1:5, k+1] = [
            sum(Intermediate_var.loc[Intermediate_var['F'] == 'D', 'Q_actual'] *
                Intermediate_var.loc[Intermediate_var['F'] == 'D', 'CI_Emissions'] *
                Intermediate_var.loc[Intermediate_var['F'] == 'D', 'ED_LCFS']) / 10**6,
            sum(Intermediate_var.loc[Intermediate_var['F'] == 'G', 'Q_actual'] *
                Intermediate_var.loc[Intermediate_var['F'] == 'G', 'CI_Emissions'] *
                Intermediate_var.loc[Intermediate_var['F'] == 'G', 'ED_LCFS']) / 10**6,
            sum(Intermediate_var.loc[Intermediate_var['F'] == 'J', 'Q_actual'] *
                Intermediate_var.loc[Intermediate_var['F'] == 'J', 'CI_Emissions'] *
                Intermediate_var.loc[Intermediate_var['F'] == 'J', 'ED_LCFS']) / 10**6,
            sum(Intermediate_var['Q_actual'] * Intermediate_var['CI_Emissions'] * Intermediate_var['ED_LCFS']) / 10**6
        ]
        
        # Rows 6 to 9 (Costs for D, G, J, and total)
        Total.iloc[5:9, k+1] = [
            sum(Intermediate_var.loc[Intermediate_var['F'] == 'D', 'Q_actual'] *
                Intermediate_var.loc[Intermediate_var['F'] == 'D', 'P_RIN_SAF_LCFS_detach']),
            sum(Intermediate_var.loc[Intermediate_var['F'] == 'G', 'Q_actual'] *
                Intermediate_var.loc[Intermediate_var['F'] == 'G', 'P_RIN_SAF_LCFS_detach']),
            sum(Intermediate_var.loc[Intermediate_var['F'] == 'J', 'Q_actual'] *
                Intermediate_var.loc[Intermediate_var['F'] == 'J', 'P_RIN_SAF_LCFS_detach']),
            sum(Intermediate_var['Q_actual'] * Intermediate_var['P_RIN_SAF_LCFS_detach'])
        ]

        # Filter Intermediate_var for rows 1 to 24
        Intermediate_var_subsidy = Intermediate_var.iloc[:24, :]
        
        # Rows 11 to 20 in the Total DataFrame
        Total.iloc[10, k+1] = sum(Intermediate_var_subsidy['TC'] * Intermediate_var_subsidy['Q_actual'])
        Total.iloc[11, k+1] = sum(Intermediate_var_subsidy['TC_IRA'] * Intermediate_var_subsidy['Q_actual'])
        Total.iloc[12, k+1] = sum(Intermediate_var_subsidy['TC_45Z'] * Intermediate_var_subsidy['Q_actual'])
        Total.iloc[13, k+1] = sum(Intermediate_var_subsidy['TC_45Q'] * Intermediate_var_subsidy['Q_actual'])
        Total.iloc[14, k+1] = sum(Intermediate_var_subsidy['TC_state'] * Intermediate_var_subsidy['Q_actual'])
        Total.iloc[15, k+1] = sum(Intermediate_var_subsidy['TC_SAF'] * Intermediate_var_subsidy['Q_actual'])
        
        # RFS road and jet credits for CA and NC
        Total.iloc[16, k+1] = sum(
            Intermediate_var_subsidy.loc[Intermediate_var_subsidy['j'] == 'CA', 'RFS_road'] *
            Intermediate_var_subsidy.loc[Intermediate_var_subsidy['j'] == 'CA', 'Q_actual']
        )
        Total.iloc[17, k+1] = sum(
            Intermediate_var_subsidy.loc[Intermediate_var_subsidy['j'] == 'NC', 'RFS_road'] *
            Intermediate_var_subsidy.loc[Intermediate_var_subsidy['j'] == 'NC', 'Q_actual']
        )
        Total.iloc[18, k+1] = sum(
            Intermediate_var_subsidy.loc[Intermediate_var_subsidy['j'] == 'CA', 'RFS_jet'] *
            Intermediate_var_subsidy.loc[Intermediate_var_subsidy['j'] == 'CA', 'Q_actual']
        )
        Total.iloc[19, k+1] = sum(
            Intermediate_var_subsidy.loc[Intermediate_var_subsidy['j'] == 'NC', 'RFS_jet'] *
            Intermediate_var_subsidy.loc[Intermediate_var_subsidy['j'] == 'NC', 'Q_actual']
        )

        Total.iloc[20, k+1] = sum(
            Intermediate_var.loc[
                (Intermediate_var['j'] == 'CA') & (Intermediate_var['f'].isin(['B0', 'E0'])),
                'RIN_obligation'
            ] * Intermediate_var.loc[
                (Intermediate_var['j'] == 'CA') & (Intermediate_var['f'].isin(['B0', 'E0'])),
                'Q_actual'
            ]
        )
        
        Total.iloc[21, k+1] = sum(
            Intermediate_var.loc[
                (Intermediate_var['j'] == 'NC') & (Intermediate_var['f'].isin(['B0', 'E0'])),
                'RIN_obligation'
            ] * Intermediate_var.loc[
                (Intermediate_var['j'] == 'NC') & (Intermediate_var['f'].isin(['B0', 'E0'])),
                'Q_actual'
            ]
        )

        Total.iloc[24, k+1] = sum(
            Intermediate_var_subsidy['LCFS_CA'] * Intermediate_var_subsidy['Q_actual']
        )

        Total.iloc[25, k+1] = (
            best_model.getVarByName(LCFS_prices_P.VarName).X *
            sum(input_compliance.loc[input_compliance['mandate'] == 'LCFS_Q', 'Q'])
        )

        Total.iloc[26, k+1] = (
            sum(Intermediate_var_subsidy['LCFS_CA'] * Intermediate_var_subsidy['Q_actual']) +
            best_model.getVarByName(LCFS_prices_P.VarName).X *
            sum(input_compliance.loc[input_compliance['mandate'] == 'LCFS_Q', 'Q'])
        )

        Total.iloc[27, k+1] = sum(
            Intermediate_var_subsidy['LCFS_SAF'] * Intermediate_var_subsidy['Q_actual']
        )
        
        Total.iloc[28, k+1] = sum(
            Intermediate_var_subsidy.loc[Intermediate_var_subsidy['j'] == 'CA', 'LCFS_SAF'] *
            Intermediate_var_subsidy.loc[Intermediate_var_subsidy['j'] == 'CA', 'Q_actual']
        )
        
        Total.iloc[29, k+1] = sum(
            Intermediate_var_subsidy.loc[Intermediate_var_subsidy['j'] == 'NC', 'LCFS_SAF'] *
            Intermediate_var_subsidy.loc[Intermediate_var_subsidy['j'] == 'NC', 'Q_actual']
        )
        
        Total.iloc[30, k+1] = sum(
            Intermediate_var.loc[
                (Intermediate_var['j'] == 'CA') & (Intermediate_var['f'] == 'J0'),
                'LCFS_SAF_obligation'
            ] * Intermediate_var.loc[
                (Intermediate_var['j'] == 'CA') & (Intermediate_var['f'] == 'J0'),
                'Q_actual'
            ]
        )
        
        Total.iloc[31, k+1] = sum(
            Intermediate_var.loc[
                (Intermediate_var['j'] == 'NC') & (Intermediate_var['f'] == 'J0'),
                'LCFS_SAF_obligation'
            ] * Intermediate_var.loc[
                (Intermediate_var['j'] == 'NC') & (Intermediate_var['f'] == 'J0'),
                'Q_actual'
            ]
        )
        
        Total.iloc[32, k+1] = (
            sum(Intermediate_var_subsidy['Subsidy'] * Intermediate_var_subsidy['Q_actual']) +
            best_model.getVarByName(LCFS_prices_P.VarName).X * sum(input_compliance.loc[input_compliance['mandate'] == 'LCFS_Q', 'Q'])
        )
        
            
        Total.iloc[33, k+1] = sum(
            Intermediate_var_subsidy.loc[Intermediate_var_subsidy['F'] == 'D', 'TC'] *
            Intermediate_var_subsidy.loc[Intermediate_var_subsidy['F'] == 'D', 'Q_actual']
        )
        
        Total.iloc[34, k+1] = sum(
            Intermediate_var_subsidy.loc[Intermediate_var_subsidy['F'] == 'G', 'TC'] *
            Intermediate_var_subsidy.loc[Intermediate_var_subsidy['F'] == 'G', 'Q_actual']
        )
        
        Total.iloc[35, k+1] = sum(
            Intermediate_var_subsidy.loc[Intermediate_var_subsidy['F'] == 'J', 'TC'] *
            Intermediate_var_subsidy.loc[Intermediate_var_subsidy['F'] == 'J', 'Q_actual']
        )

        Total.iloc[36, k+1] = sum(
                Intermediate_var_subsidy.loc[Intermediate_var['F'] == 'D','LCFS_CA'] * Intermediate_var_subsidy.loc[Intermediate_var['F'] == 'D','Q_actual']
            )

        Total.iloc[37, k+1] = sum(
                Intermediate_var_subsidy.loc[Intermediate_var['F'] == 'G','LCFS_CA'] * Intermediate_var_subsidy.loc[Intermediate_var['F'] == 'G','Q_actual']
            )
        
        Total.iloc[38, k+1] = sum(
                Intermediate_var_subsidy.loc[Intermediate_var['F'] == 'J','LCFS_CA'] * Intermediate_var_subsidy.loc[Intermediate_var['F'] == 'J','Q_actual']
            )
        
        Total.iloc[39, k+1] = sum(
                Intermediate_var.loc[Intermediate_var['f'] == 'B0','LCFS_CA_obligation'] * Intermediate_var.loc[Intermediate_var['f'] == 'B0','Q_actual']
            )

        Total.iloc[40, k+1] = sum(
                Intermediate_var.loc[Intermediate_var['f'] == 'E0','LCFS_CA_obligation'] * Intermediate_var.loc[Intermediate_var['f'] == 'E0','Q_actual']
            )

        Total.iloc[41, k+1] = sum(
                Intermediate_var_subsidy.loc[(Intermediate_var_subsidy['j'] == 'CA')&(Intermediate_var_subsidy['F'] == 'D'), 'RFS_road'] *
                Intermediate_var_subsidy.loc[(Intermediate_var_subsidy['j'] == 'CA')&(Intermediate_var_subsidy['F'] == 'D'), 'Q_actual']
            )
        Total.iloc[42, k+1] = sum(
                Intermediate_var_subsidy.loc[(Intermediate_var_subsidy['j'] == 'NC')&(Intermediate_var_subsidy['F'] == 'D'), 'RFS_road'] *
                Intermediate_var_subsidy.loc[(Intermediate_var_subsidy['j'] == 'NC')&(Intermediate_var_subsidy['F'] == 'D'), 'Q_actual']
            )
        Total.iloc[43, k+1] = sum(
                Intermediate_var_subsidy.loc[(Intermediate_var_subsidy['j'] == 'CA')&(Intermediate_var_subsidy['F'] == 'G'), 'RFS_road'] *
                Intermediate_var_subsidy.loc[(Intermediate_var_subsidy['j'] == 'CA')&(Intermediate_var_subsidy['F'] == 'G'), 'Q_actual']
            )
        Total.iloc[44, k+1] = sum(
                Intermediate_var_subsidy.loc[(Intermediate_var_subsidy['j'] == 'NC')&(Intermediate_var_subsidy['F'] == 'G'), 'RFS_road'] *
                Intermediate_var_subsidy.loc[(Intermediate_var_subsidy['j'] == 'NC')&(Intermediate_var_subsidy['F'] == 'G'), 'Q_actual']
            )
        Total.iloc[45, k+1] = sum(
                Intermediate_var_subsidy.loc[(Intermediate_var_subsidy['j'] == 'CA')&(Intermediate_var_subsidy['F'] == 'J'), 'RFS_jet'] *
                Intermediate_var_subsidy.loc[(Intermediate_var_subsidy['j'] == 'CA')&(Intermediate_var_subsidy['F'] == 'J'), 'Q_actual']
            )
        Total.iloc[46, k+1] = sum(
                Intermediate_var_subsidy.loc[(Intermediate_var_subsidy['j'] == 'NC')&(Intermediate_var_subsidy['F'] == 'J'), 'RFS_jet'] *
                Intermediate_var_subsidy.loc[(Intermediate_var_subsidy['j'] == 'NC')&(Intermediate_var_subsidy['F'] == 'J'), 'Q_actual']
            )
        Total.iloc[47, k+1] = sum(
                Intermediate_var.loc[(Intermediate_var['j'] == 'CA') & (Intermediate_var['f'].isin(['B0'])),
                    'RIN_obligation'] * Intermediate_var.loc[
                    (Intermediate_var['j'] == 'CA') & (Intermediate_var['f'].isin(['B0'])),
                    'Q_actual']
            )
        Total.iloc[48, k+1] = sum(
                Intermediate_var.loc[(Intermediate_var['j'] == 'NC') & (Intermediate_var['f'].isin(['B0'])),
                    'RIN_obligation'] * Intermediate_var.loc[
                    (Intermediate_var['j'] == 'NC') & (Intermediate_var['f'].isin(['B0'])),
                    'Q_actual']
            )
        Total.iloc[49, k+1] = sum(
                Intermediate_var.loc[(Intermediate_var['j'] == 'CA') & (Intermediate_var['f'].isin(['E0'])),
                    'RIN_obligation'] * Intermediate_var.loc[
                    (Intermediate_var['j'] == 'CA') & (Intermediate_var['f'].isin(['E0'])),
                    'Q_actual']
            )
        Total.iloc[50, k+1] = sum(
                Intermediate_var.loc[(Intermediate_var['j'] == 'NC') & (Intermediate_var['f'].isin(['E0'])),
                    'RIN_obligation'] * Intermediate_var.loc[
                    (Intermediate_var['j'] == 'NC') & (Intermediate_var['f'].isin(['E0'])),
                    'Q_actual']
            )
        Total.iloc[51, k+1] = sum(
                Intermediate_var.loc[(Intermediate_var['j'] == 'CA') & (Intermediate_var['f'].isin(['J0'])),
                    'RIN_obligation'] * Intermediate_var.loc[
                    (Intermediate_var['j'] == 'CA') & (Intermediate_var['f'].isin(['J0'])),
                    'Q_actual']
            )
        Total.iloc[52, k+1] = sum(
                Intermediate_var.loc[(Intermediate_var['j'] == 'NC') & (Intermediate_var['f'].isin(['J0'])),
                    'RIN_obligation'] * Intermediate_var.loc[
                    (Intermediate_var['j'] == 'NC') & (Intermediate_var['f'].isin(['J0'])),
                    'Q_actual']
            )

        # Rows 1 to 4 in the Feedstock_source DataFrame
        Feedstock_source.iloc[0:4, k+1] = [
            get_x(biofuel_quantity.loc[(biofuel_quantity['f'] == 'B100') & (biofuel_quantity['s'] == 'soyoil'), 'Q']),
            get_x(biofuel_quantity.loc[(biofuel_quantity['f'] == 'B100') & (biofuel_quantity['s'] == 'animalfat'), 'Q']),
            get_x(biofuel_quantity.loc[(biofuel_quantity['f'] == 'RD') & (biofuel_quantity['s'] == 'soyoil'), 'Q']),
            get_x(biofuel_quantity.loc[(biofuel_quantity['f'] == 'RD') & (biofuel_quantity['s'] == 'animalfat'), 'Q'])
        ]

        # Rows 5 to 8 in the Feedstock_source DataFrame
        Feedstock_source.iloc[4:8, k+1] = [
            get_x(biofuel_quantity.loc[(biofuel_quantity['f'] == 'SAF') & (biofuel_quantity['s'] == 'soyoil'), 'Q']),
            get_x(biofuel_quantity.loc[(biofuel_quantity['f'] == 'SAF') & (biofuel_quantity['s'] == 'animalfat'), 'Q']),
            get_x(biofuel_quantity.loc[(biofuel_quantity['f'] == 'SAF') & (biofuel_quantity['s'] == 'corn'), 'Q']),
            get_x(biofuel_quantity.loc[(biofuel_quantity['f'] == 'SAF') & (biofuel_quantity['s'] == 'sugarcane'), 'Q'])
        ]

        # Rows 9 to 10 in the Feedstock_source DataFrame
        # SAF-corn-CCS (row 8)
        mask_corn = (biofuel_quantity['f'] == 'SAF') & (biofuel_quantity['s'] == 'corn')
        Feedstock_source.iloc[8, k+1] = sum(
            best_model.getVarByName(q.VarName).X * ccs
            for q, ccs in zip(
                biofuel_quantity.loc[mask_corn, 'Q'],
                biofuel_quantity.loc[mask_corn, 'CCS_tech']
            )
        )
        
        # SAF-sugarcane-CCS (row 9)
        mask_sugarcane = (biofuel_quantity['f'] == 'SAF') & (biofuel_quantity['s'] == 'sugarcane')
        Feedstock_source.iloc[9, k+1] = sum(
            best_model.getVarByName(q.VarName).X * ccs
            for q, ccs in zip(
                biofuel_quantity.loc[mask_sugarcane, 'Q'],
                biofuel_quantity.loc[mask_sugarcane, 'CCS_tech']
            )
        )

        # Rows 11 to 12 in the Feedstock_source DataFrame
        Feedstock_source.iloc[10:12, k+1] = [
            get_x(biofuel_quantity.loc[(biofuel_quantity['f'] == 'gasE100') & (biofuel_quantity['s'] == 'corn'), 'Q']),
            get_x(biofuel_quantity.loc[(biofuel_quantity['f'] == 'gasE100') & (biofuel_quantity['s'] == 'sugarcane'), 'Q'])
        ]

        
# # Function to append data to an existing Excel file
# def append_to_excel(file_path, df, sheet_name):
#     try:
#         # Load the existing workbook
#         with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
#             # Write the DataFrame to the specified sheet
#             df.to_excel(writer, index=False, sheet_name=sheet_name)
#     except FileNotFoundError:
#         # If the file does not exist, create a new one
#         with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
#             df.to_excel(writer, index=False, sheet_name=sheet_name)

# # Append data to the Excel file
# append_to_excel(f'{output_dir}Solution.xlsx', solution, 'Aviation_intensity_standard')
# append_to_excel(f'{output_dir}Fitted_quantity.xlsx', Fitted_quantity, 'Aviation_intensity_standard')
# append_to_excel(f'{output_dir}Fuel_price.xlsx', Fuel_price, 'Aviation_intensity_standard')
# append_to_excel(f'{output_dir}Fuel_CI.xlsx', Fuel_CI, 'Aviation_intensity_standard')
# append_to_excel(f'{output_dir}Total.xlsx', Total, 'Aviation_intensity_standard')
# append_to_excel(f'{output_dir}Feedstock_source.xlsx', Feedstock_source, 'Aviation_intensity_standard')

from openpyxl import load_workbook

def append_or_create_excel(df, file_path, sheet_name):
    """
    Append DataFrame to existing Excel file or create new one if it doesn't exist
    """
    if os.path.exists(file_path):
        # File exists, append to it
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
    else:
        # File doesn't exist, create new one
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)

# Apply to all your DataFrames using results_dir
append_or_create_excel(solution, os.path.join(results_dir, 'Solution.xlsx'), 'Aviation_intensity_standard')
append_or_create_excel(Fitted_quantity, os.path.join(results_dir, 'Fitted_quantity.xlsx'), 'Aviation_intensity_standard') 
append_or_create_excel(Fuel_price, os.path.join(results_dir, 'Fuel_price.xlsx'), 'Aviation_intensity_standard')
append_or_create_excel(Fuel_CI, os.path.join(results_dir, 'Fuel_CI.xlsx'), 'Aviation_intensity_standard')
append_or_create_excel(Total, os.path.join(results_dir, 'Total.xlsx'), 'Aviation_intensity_standard')
append_or_create_excel(Feedstock_source, os.path.join(results_dir, 'Feedstock_source.xlsx'), 'Aviation_intensity_standard')



