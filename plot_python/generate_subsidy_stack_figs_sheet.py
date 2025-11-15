"""
Generate Subsidy_stack_figs Sheet for results_plot.xlsx

This script reads subsidy and cost metrics from intermediate result files
and compiles them into a "Subsidy_stack_figs" sheet with a summary table.

Source File:
From results/Intermediate_results/{col}.xlsx (Mean, Q_10th, Q_33th, Q_67th, Q_90th):
   - P_all, PC_F0, RIN, LCFS_CA, Tax, TC_45Z, TC_45Q, TC_state, Subsidy,
     Production_cost, Fixed_production_cost, Feedstock_cost

Summary Table:
Displays cost breakdown and pricing:
   - Fixed_production_cost, Feedstock_cost, Production_cost (from original data)
   - RIN, LCFS, 45Z, 45Q, State tax credits (negatives of original values)
   - Net Price (sum of Production_cost, RIN, LCFS, 45Z, 45Q, State tax credits)
   - Fossil Price (PC_F0 from original data)

Usage:
    python generate_subsidy_stack_figs_sheet1.py [--results-dir RESULTS_DIR] [--output-file OUTPUT_FILE] [--columns COLUMNS]
"""

import pandas as pd
import os
import sys
from openpyxl import load_workbook

def generate_subsidy_stack_figs_sheet(results_dir='results', output_file='results_plot.xlsx', columns=None):
    """
    Generate the Subsidy_stack_figs sheet by reading metrics from Intermediate_results/{col}.xlsx
    
    Parameters:
    -----------
    results_dir : str
        Path to the results directory (default: 'results')
    output_file : str
        Path to the output Excel file (default: 'results_plot.xlsx')
    columns : str or list
        Comma-separated column names or list of columns to extract (default: ['Mean'])
    """
    
    # Determine which columns to extract (default: Mean only)
    if columns is None:
        columns = ['Mean']
    elif isinstance(columns, str):
        columns = [c.strip() for c in columns.split(',') if c.strip()]
    
    # Use first column for sheet writing
    sel_col = columns[0]
    
    # Define source file - try multiple naming patterns
    # First try exact match, then try without Q_ prefix (10th instead of Q_10th)
    possible_filenames = [
        f'{sel_col}.xlsx',  # Mean.xlsx, Q_10th.xlsx, etc.
        f'{sel_col.replace("Q_", "")}.xlsx'  # 10th.xlsx if input was Q_10th
    ]
    
    source_file = None
    for filename in possible_filenames:
        candidate = os.path.join(results_dir, 'Intermediate_results', filename)
        if os.path.exists(candidate):
            source_file = candidate
            break
    
    if source_file is None:
        # Try to find with Mean.xlsx as fallback
        fallback_file = os.path.join(results_dir, 'Intermediate_results', 'Mean.xlsx')
        if os.path.exists(fallback_file):
            print(f"Warning: Could not find file for column '{sel_col}' in Intermediate_results/")
            print(f"  Tried: {possible_filenames}")
            print(f"  Falling back to Mean.xlsx")
            source_file = fallback_file
        else:
            print(f"Error: Missing result file for column '{sel_col}'")
            print(f"  Tried: {possible_filenames}")
            print(f"  Also checked: {fallback_file}")
            return False
    
    print("Generating Subsidy_stack_figs sheet...")
    print(f"Reading from: {source_file}")
    print(f"Output file: {output_file}")
    print(f"Extracting columns: {columns}")
    
    # Define columns to extract (map display names to actual column names in the file)
    columns_to_extract = {
        'P_all': 'P_all',
        'PC_F0': 'PC_F0',
        'RIN': 'RIN_obligation',
        'LCFS_CA': 'LCFS_CA_obligation',
        'Tax': 'Tax',
        'TC_45Z': 'TC_45Z',
        'TC_45Q': 'TC_45Q',
        'TC_state': 'TC_state',
        'Subsidy': 'Subsidy',
        'Production_cost': 'Production_cost',
        'Fixed_production_cost': 'Fixed_production_cost',
        'Feedstock_cost': 'Feedstock_cost'
    }
    
    # Read data from the source file (Current_policy sheet only)
    print(f"\nReading {os.path.basename(source_file)} (Current_policy sheet)...")
    
    try:
        df = pd.read_excel(source_file, sheet_name='Current_policy')
        print(f"  ✓ Current_policy sheet loaded")
        
        # Extract only the specified columns, using the actual column names from the file
        extracted_data = {}
        actual_columns_used = []
        
        for display_name, actual_name in columns_to_extract.items():
            if actual_name in df.columns:
                extracted_data[display_name] = df[actual_name]
                actual_columns_used.append(display_name)
            else:
                print(f"  ⚠ Column '{actual_name}' not found in sheet")
        
        if not extracted_data:
            print(f"Error: None of the specified columns found in the sheet")
            print(f"Available columns: {list(df.columns)}")
            return False
        
        subsidy_stack_figs_df = pd.DataFrame(extracted_data)
        print(f"  ✓ Extracted {len(extracted_data)} columns")
        
        # Add Name column as the first column (if it exists in the original df)
        if 'Name' in df.columns:
            subsidy_stack_figs_df.insert(0, 'Name', df['Name'])
            print(f"  ✓ Added Name column as row labels")
        
        # Reorder columns in the specified order
        column_order = [
            'Name',
            'P_all',
            'PC_F0',
            'RIN',
            'LCFS_CA',
            'Tax',
            'TC_45Z',
            'TC_45Q',
            'TC_state',
            'Subsidy',
            'Production_cost',
            'Fixed_production_cost',
            'Feedstock_cost'
        ]
        
        # Only include columns that exist in the dataframe
        existing_columns = [col for col in column_order if col in subsidy_stack_figs_df.columns]
        subsidy_stack_figs_df = subsidy_stack_figs_df[existing_columns]
        print(f"  ✓ Reordered columns in specified order")
    except Exception as e:
        print(f"\nError reading file: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    # Write to Excel file
    print(f"\nWriting Subsidy_stack_figs sheet to {output_file}...")
    
    try:
        if os.path.exists(output_file):
            # Load existing workbook and replace/add the sheet
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', 
                              if_sheet_exists='replace') as writer:
                subsidy_stack_figs_df.to_excel(writer, sheet_name='Subsidy_stack_figs', index=False)
            print(f"✓ Updated 'Subsidy_stack_figs' sheet in {output_file}")
        else:
            # Create new workbook
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                subsidy_stack_figs_df.to_excel(writer, sheet_name='Subsidy_stack_figs', index=False)
            print(f"✓ Created {output_file} with 'Subsidy_stack_figs' sheet")
        
        # Now add the summary table using openpyxl
        print(f"\nAdding cost breakdown summary table...")
        wb = load_workbook(output_file)
        ws = wb['Subsidy_stack_figs']
        
        # Summary table parameters
        summary_start_col = 15  # Column O
        summary_start_row = 2   # Row 2
        
        # Define summary table columns
        summary_columns = [
            'Fixed_production_cost',
            'Feedstock_cost',
            'Production_cost',
            'RIN',           # negative
            'LCFS',          # negative
            '45Z',           # negative
            '45Q',           # negative
            'State tax credits',  # negative
            'Net Price',     # calculated
            'Fossil Price'   # PC_F0
        ]
        
        # Column mapping to source data
        source_mapping = {
            'Fixed_production_cost': 'Fixed_production_cost',
            'Feedstock_cost': 'Feedstock_cost',
            'Production_cost': 'Production_cost',
            'RIN': 'RIN_obligation',
            'LCFS': 'LCFS_CA_obligation',
            '45Z': 'TC_45Z',
            '45Q': 'TC_45Q',
            'State tax credits': 'TC_state',
            'Fossil Price': 'PC_F0'
        }
        
        # Write summary table headers
        ws.cell(row=summary_start_row, column=summary_start_col).value = 'Name'
        for col_idx, col_name in enumerate(summary_columns, start=1):
            ws.cell(row=summary_start_row, column=summary_start_col + col_idx).value = col_name
        
        # Write summary table data rows (one row for each fuel type from original data)
        for row_idx in range(len(df)):
            # Write Name in first column
            if 'Name' in df.columns:
                ws.cell(row=summary_start_row + row_idx + 1, column=summary_start_col).value = df['Name'].iloc[row_idx]
            
            # Write data for each column
            for col_idx, col_name in enumerate(summary_columns, start=1):
                if col_name == 'Net Price':
                    # Calculate Net Price = Production_cost + RIN + LCFS + 45Z + 45Q + State tax credits (with negatives)
                    prod_cost = df['Production_cost'].iloc[row_idx] if 'Production_cost' in df.columns else 0
                    rin = -df['RIN_obligation'].iloc[row_idx] if 'RIN_obligation' in df.columns else 0
                    lcfs = -df['LCFS_CA_obligation'].iloc[row_idx] if 'LCFS_CA_obligation' in df.columns else 0
                    tc_45z = -df['TC_45Z'].iloc[row_idx] if 'TC_45Z' in df.columns else 0
                    tc_45q = -df['TC_45Q'].iloc[row_idx] if 'TC_45Q' in df.columns else 0
                    tc_state = -df['TC_state'].iloc[row_idx] if 'TC_state' in df.columns else 0
                    
                    net_price = prod_cost + rin + lcfs + tc_45z + tc_45q + tc_state
                    ws.cell(row=summary_start_row + row_idx + 1, column=summary_start_col + col_idx).value = round(net_price, 6)
                
                elif col_name in ['RIN', 'LCFS', '45Z', '45Q', 'State tax credits']:
                    # These should be negatives
                    source_col = source_mapping.get(col_name)
                    if source_col and source_col in df.columns:
                        value = df[source_col].iloc[row_idx]
                        if value is not None:
                            try:
                                ws.cell(row=summary_start_row + row_idx + 1, column=summary_start_col + col_idx).value = round(-float(value), 6)
                            except (ValueError, TypeError):
                                ws.cell(row=summary_start_row + row_idx + 1, column=summary_start_col + col_idx).value = None
                
                else:
                    # Regular columns
                    source_col = source_mapping.get(col_name)
                    if source_col and source_col in df.columns:
                        value = df[source_col].iloc[row_idx]
                        if value is not None:
                            try:
                                ws.cell(row=summary_start_row + row_idx + 1, column=summary_start_col + col_idx).value = round(float(value), 6)
                            except (ValueError, TypeError):
                                ws.cell(row=summary_start_row + row_idx + 1, column=summary_start_col + col_idx).value = None
        
        # Save the workbook
        wb.save(output_file)
        print(f"  ✓ Summary table added starting at column O")
        
        print(f"\n✓ Subsidy_stack_figs sheet generation complete!")
        print(f"  Main sheet shape: {subsidy_stack_figs_df.shape}")
        print(f"  Summary table: {len(df)} rows × {len(summary_columns)} columns")
        print(f"  Columns: {len(extracted_data)}")
        print(f"  Rows: {len(subsidy_stack_figs_df)}")
        
        return True
        
    except Exception as e:
        print(f"\nError reading or writing file: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """Main function to run the script"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Generate Subsidy_stack_figs sheet in results_plot.xlsx'
    )
    parser.add_argument(
        '--results-dir', 
        type=str, 
        default='results',
        help='Path to results directory (default: results)'
    )
    parser.add_argument(
        '--output-file',
        type=str,
        default='results_plot.xlsx',
        help='Output Excel file name (default: results_plot.xlsx)'
    )
    parser.add_argument(
        '--columns',
        default=None,
        help='Comma-separated list of columns to extract (default: Mean)'
    )
    
    args = parser.parse_args()
    
    # Generate the Subsidy_stack_figs sheet
    success = generate_subsidy_stack_figs_sheet(args.results_dir, args.output_file, args.columns)
    
    if success:
        sys.exit(0)
    else:
        sys.exit(1)


if __name__ == '__main__':
    main()
