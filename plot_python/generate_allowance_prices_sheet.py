"""
Generate Allowance_Prices Sheet for results_plot.xlsx

This script reads allowance and subsidy price data from Solution.xlsx
and creates an "Allowance_Prices" sheet with all policy comparisons.

Source File:
From results/Solution.xlsx, extracts from each policy sheet:
   - State_tax_credit
   - Additional_tax_credit
   - P_D2
   - P_D4
   - P_D6
   - P_LCFS
   - P_SAF
   - P_carbon_tax (Carbon tax policies only)

Summary:
Creates a comparison table across 11 policies with the above 8 metrics per policy.

Usage:
    python generate_allowance_prices_sheet1.py [--results-dir RESULTS_DIR] [--output-file OUTPUT_FILE] [--columns COLUMNS]
"""

import pandas as pd
import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_allowance_prices_sheet(results_dir='results', output_file='results_plot.xlsx', columns=None):
    """
    Generate the Allowance_Prices sheet by reading price data from Solution.xlsx
    
    Parameters:
    -----------
    results_dir : str
        Path to the results directory (default: 'results')
    output_file : str
        Path to the output Excel file (default: 'results_plot.xlsx')
    columns : str or list
        Comma-separated column names or list of columns to extract (default: ['Mean'])
    """
    
    # Determine which columns to extract (default: Mean only)
    if columns is None:
        columns = ['Mean']
    elif isinstance(columns, str):
        columns = [c.strip() for c in columns.split(',') if c.strip()]
    
    # Use first column for sheet writing
    sel_col = columns[0]
    
    # Define source file
    source_file = os.path.join(results_dir, 'Solution.xlsx')
    
    # Check if file exists
    if not os.path.exists(source_file):
        print(f"Error: Missing result file: {source_file}")
        print("\nPlease run the models first to generate all results.")
        return False
    
    print("Generating Allowance_Prices sheet...")
    print(f"Reading from: {source_file}")
    print(f"Extracting columns: {columns}")
    
    # Policy sheet names mapping to display names - ordered as specified
    policy_mapping = {
        'Pure_quantity': 'Q mandate',
        'Nested_D2_RV1': 'Nested D2',
        'Nonnested_D2': 'Non-nested D2',
        'Nested_D2_RV2': 'Nested D2 + stricter RFS',
        'Nonnested_D2_RV2': 'Non-nested + stricter RFS',
        'Carbon_tax': 'Carbon Tax',
        'Aviation_intensity_standard': 'Aviation intensity standard',
        'Current_policy': 'Current Policy',
        'Current_policy_no_LCFS': 'Current Policy + No LCFS',
        'Carbon_tax_noSAF_floor': 'Carbon tax+no SAF',
        'No_policy_floor': 'No policy'
    }
    
    # Rows to extract
    row_labels = [
        'State_tax_credit',
        'Additional_tax_credit',
        'P_D2',
        'P_D4',
        'P_D6',
        'P_LCFS',
        'P_SAF'
    ]
    
    # Create data dictionary to store extracted values
    # Structure: data[policy_name][row_label][col] = value
    data = {}
    
    # Extract data from each policy sheet
    for sheet_name, display_name in policy_mapping.items():
        try:
            df = pd.read_excel(source_file, sheet_name=sheet_name)
            policy_data = {}  # Will store {row_label: {col: value}}
            
            # Extract specified rows for each requested column
            for row_label in row_labels:
                if row_label not in policy_data:
                    policy_data[row_label] = {}
                
                if row_label in df['Name'].values:
                    idx = df[df['Name'] == row_label].index[0]
                    for col in columns:
                        if col in df.columns:
                            policy_data[row_label][col] = df.loc[idx, col]
                        else:
                            policy_data[row_label][col] = None
                else:
                    for col in columns:
                        policy_data[row_label][col] = None
            
            # Extract P_carbon_tax for Carbon tax related policies
            if 'P_carbon_tax' not in policy_data:
                policy_data['P_carbon_tax'] = {}
            
            if 'Carbon' in sheet_name:
                if 'P_carbon_tax' in df['Name'].values:
                    idx = df[df['Name'] == 'P_carbon_tax'].index[0]
                    for col in columns:
                        if col in df.columns:
                            policy_data['P_carbon_tax'][col] = df.loc[idx, col]
                        else:
                            policy_data['P_carbon_tax'][col] = None
                else:
                    for col in columns:
                        policy_data['P_carbon_tax'][col] = None
            else:
                for col in columns:
                    policy_data['P_carbon_tax'][col] = None
            
            data[display_name] = policy_data
            print(f"✓ {display_name} extracted")
            
        except Exception as e:
            print(f"✗ Error reading {sheet_name}: {e}")
            return False
    
    # Create DataFrame for selected column (using sel_col)
    row_names_display = [
        'State_tax_credit ($)',
        'Additional_tax_credit ($)',
        'P_D2 ($)',
        'P_D4 ($)',
        'P_D6 ($)',
        'P_LCFS ($)',
        'P_SAF ($)',
        'Carbon tax (fixed) ($)'
    ]
    
    # Build DataFrame using sel_col values
    df_prices_data = {}
    for policy_name, policy_data in data.items():
        df_prices_data[policy_name] = []
        for row_label in ['State_tax_credit', 'Additional_tax_credit', 'P_D2', 'P_D4', 'P_D6', 'P_LCFS', 'P_SAF', 'P_carbon_tax']:
            value = policy_data.get(row_label, {}).get(sel_col, None)
            df_prices_data[policy_name].append(value)
    
    df_prices = pd.DataFrame(df_prices_data, index=row_names_display)
    
    print(f"\nExtracted {len(df_prices.columns)} policies × {len(df_prices)} rows")
    
    # Create summary table with selected policies
    summary_policy_mapping = {
        'Current_policy': 'Current Policy',
        'Pure_quantity': 'SAF Credit',
        'Nested_D2_RV1': 'Nested D2',
        'Nonnested_D2': 'D2 with aviation obligation',
        'Nested_D2_RV2': 'Nested D2 + stricter RFS',
        'Carbon_tax': 'Carbon Tax + SAF credit',
        'Aviation_intensity_standard': 'Aviation Intensity Std'
    }
    
    summary_row_sources = [
        ('P_D2', 'D2 RIN'),
        ('P_D4', 'D4 RIN'),
        ('P_D6', 'D6 RIN'),
        ('P_LCFS', 'LCFS Allowance Price'),
        ('State_tax_credit', 'State Tax credit'),
        ('Additional_tax_credit', "Addt'l SAF Credit"),
        ('P_carbon_tax', 'Carbon Tax ($/ton)'),
        ('P_SAF', 'Aviation Intensity Std. Credit')
    ]
    
    summary_data_dict = {}
    
    # Extract data for summary table from specific policies
    for sheet_name, display_name in summary_policy_mapping.items():
        try:
            df_temp = pd.read_excel(source_file, sheet_name=sheet_name)
            policy_data = []
            
            for source_row, display_row in summary_row_sources:
                value = None
                
                if source_row == 'P_SAF':
                    if 'P_SAF' in df_temp['Name'].values:
                        idx = df_temp[df_temp['Name'] == 'P_SAF'].index[0]
                        value = df_temp.loc[idx, sel_col] if sel_col in df_temp.columns else None
                        
                elif source_row == 'P_carbon_tax':
                    if 'P_carbon_tax' in df_temp['Name'].values:
                        idx = df_temp[df_temp['Name'] == 'P_carbon_tax'].index[0]
                        value = df_temp.loc[idx, sel_col] if sel_col in df_temp.columns else None
                        
                elif source_row in df_temp['Name'].values:
                    idx = df_temp[df_temp['Name'] == source_row].index[0]
                    value = df_temp.loc[idx, sel_col] if sel_col in df_temp.columns else None
                
                policy_data.append(value)
            
            summary_data_dict[display_name] = policy_data
            
        except Exception as e:
            print(f"Warning: Could not extract summary data from {sheet_name}: {e}")
    
    summary_row_names = [row[1] for row in summary_row_sources]
    df_summary = pd.DataFrame(summary_data_dict, index=summary_row_names)
    
    print(f"✓ Summary table created: {len(df_summary)} metrics × {len(df_summary.columns)} policies")
    
    # Write to Excel
    try:
        # Check if output file exists, load it
        if os.path.exists(output_file):
            wb = load_workbook(output_file)
            if 'Allowance_Prices' in wb.sheetnames:
                del wb['Allowance_Prices']
        else:
            from openpyxl import Workbook
            wb = Workbook()
            # Remove default sheet if present
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
        
        # Create new sheet
        ws = wb.create_sheet('Allowance_Prices')
        
        # ========== MAIN TABLE (11 policies) ==========
        # Write header row (policy names)
        ws.cell(row=1, column=1).value = 'Metric'
        for col_idx, policy_name in enumerate(df_prices.columns, start=2):
            ws.cell(row=1, column=col_idx).value = policy_name
        
        # Write data rows
        for row_idx, row_label in enumerate(df_prices.index, start=2):
            ws.cell(row=row_idx, column=1).value = row_label
            for col_idx, policy_name in enumerate(df_prices.columns, start=2):
                value = df_prices.loc[row_label, policy_name]
                ws.cell(row=row_idx, column=col_idx).value = value
        
        # ========== SUMMARY TABLE (7 policies) ==========
        summary_start_row = len(df_prices) + 3  # Leave blank row
        summary_start_col = 1
        
        # Write summary table header
        ws.cell(row=summary_start_row, column=summary_start_col).value = 'Metric'
        for col_idx, policy_name in enumerate(df_summary.columns, start=summary_start_col + 1):
            ws.cell(row=summary_start_row, column=col_idx).value = policy_name
        
        # Write summary data rows
        for row_idx, row_label in enumerate(df_summary.index, start=summary_start_row + 1):
            ws.cell(row=row_idx, column=summary_start_col).value = row_label
            for col_idx, policy_name in enumerate(df_summary.columns, start=summary_start_col + 1):
                value = df_summary.loc[row_label, policy_name]
                ws.cell(row=row_idx, column=col_idx).value = value
        
        # Apply formatting - MAIN TABLE
        # Header row - bold and light blue fill
        header_fill = PatternFill(start_color='D9E8F5', end_color='D9E8F5', fill_type='solid')
        header_font = Font(bold=True)
        
        for col in range(1, len(df_prices.columns) + 2):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Metric column - bold
        for row in range(2, len(df_prices) + 2):
            cell = ws.cell(row=row, column=1)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Set column widths for main table
        ws.column_dimensions['A'].width = 25
        for col in range(2, len(df_prices.columns) + 2):
            ws.column_dimensions[chr(64 + col)].width = 18
        
        # Center align all data cells in main table
        for row in range(2, len(df_prices) + 2):
            for col in range(2, len(df_prices.columns) + 2):
                ws.cell(row=row, column=col).alignment = Alignment(
                    horizontal='center', 
                    vertical='center'
                )
        
        # Add borders to main table
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(1, len(df_prices) + 2):
            for col in range(1, len(df_prices.columns) + 2):
                ws.cell(row=row, column=col).border = thin_border
        
        # Apply formatting - SUMMARY TABLE
        summary_end_row = summary_start_row + len(df_summary)
        
        # Header row - bold (no fill color)
        for col in range(summary_start_col, summary_start_col + len(df_summary.columns) + 1):
            cell = ws.cell(row=summary_start_row, column=col)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Metric column - bold
        for row in range(summary_start_row + 1, summary_end_row + 1):
            cell = ws.cell(row=row, column=summary_start_col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Center align all data cells in summary table
        for row in range(summary_start_row + 1, summary_end_row + 1):
            for col in range(summary_start_col + 1, summary_start_col + len(df_summary.columns) + 1):
                ws.cell(row=row, column=col).alignment = Alignment(
                    horizontal='center', 
                    vertical='center'
                )
        
        # Add borders to summary table
        for row in range(summary_start_row, summary_end_row + 1):
            for col in range(summary_start_col, summary_start_col + len(df_summary.columns) + 1):
                ws.cell(row=row, column=col).border = thin_border
        # Save workbook
        wb.save(output_file)
        print(f"\n✓ Allowance_Prices sheet added to {output_file}")
        print(f"  Main table: {len(df_prices)} rows × {len(df_prices.columns)} columns")
        print(f"  Summary table: {len(df_summary)} rows × {len(df_summary.columns)} columns (starting at row {summary_start_row})")
        
        return True
        
    except Exception as e:
        print(f"Error writing to Excel: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Generate Allowance_Prices sheet for results_plot.xlsx'
    )
    parser.add_argument(
        '--results-dir',
        default='results',
        help='Path to results directory (default: results)'
    )
    parser.add_argument(
        '--output-file',
        default='results_plot.xlsx',
        help='Path to output Excel file (default: results_plot.xlsx)'
    )
    parser.add_argument(
        '--columns',
        default=None,
        help='Comma-separated list of columns to extract (default: Mean)'
    )
    
    args = parser.parse_args()
    
    success = generate_allowance_prices_sheet(
        results_dir=args.results_dir,
        output_file=args.output_file,
        columns=args.columns
    )
    
    sys.exit(0 if success else 1)
