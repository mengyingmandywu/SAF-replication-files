"""
Generate Feedstock Sheet for results_plot.xlsx

This script reads feedstock quantity metrics from Fitted_quantity.xlsx and compiles them
into a "Feedstock" sheet with two summary tables:
1. Absolute feedstock quantities with unit conversions
2. Change in feedstock demand relative to Current Policy

Source Metrics:
From Fitted_quantity.xlsx:
   - Feedstock quantity (billion)
   - Q_soyoil (pounds)
   - Q_animal fat (pounds)
   - Q_corn (bushels)
   - Q_sugarcane (tons)

Summary Tables:
1. Starting at N2: Feedstock quantities by policy
2. Starting at N8: Change in feedstock demand relative to Current Policy

Usage:
    python generate_feedstock_sheet1.py [--results-dir RESULTS_DIR] [--output-file OUTPUT_FILE] [--columns COLUMNS]
"""

import pandas as pd
import os
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def generate_feedstock_sheet(results_dir='results', output_file='results_plot.xlsx', columns=None):
    """
    Generate the Feedstock sheet by reading metrics from Fitted_quantity.xlsx
    
    Parameters:
    -----------
    results_dir : str
        Path to the results directory (default: 'results')
    output_file : str
        Path to the output Excel file (default: 'results_plot.xlsx')
    columns : str or list
        Comma-separated column names or list of columns to extract (default: ['Mean'])
    """
    
    # Determine which columns to extract (default: Mean only)
    if columns is None:
        columns = ['Mean']
    elif isinstance(columns, str):
        columns = [c.strip() for c in columns.split(',') if c.strip()]
    
    # Use first column for sheet writing
    sel_col = columns[0]
    
    # Define source file
    source_file = os.path.join(results_dir, 'Fitted_quantity.xlsx')
    
    # Check if file exists
    if not os.path.exists(source_file):
        print(f"Error: Missing result file: {source_file}")
        print("\nPlease run the models first to generate all results.")
        return False
    
    print("Generating Feedstock sheet...")
    print(f"Reading from: {results_dir}")
    print(f"Output file: {output_file}")
    print(f"Extracting columns: {columns}")
    
    # Model name mapping (from sheet names to column headers)
    model_mapping = {
        'Current_policy': 'Current Policy',
        'Pure_quantity': 'Q mandate',
        'Nested_D2_RV1': 'Nested D2',
        'Nested_D2_RV2': 'Nested D2 + stricter RFS',
        'Nonnested_D2': 'Non-nested D2',
        'Nonnested_D2_RV2': 'Non-nested + stricter RFS',
        'Carbon_tax': 'Carbon Tax',
        'Aviation_intensity_standard': 'Aviation intensity standard',
        'Current_policy_no_LCFS': 'Current Policy + No LCFS',
        'Carbon_tax_noSAF_floor': 'Carbon tax+no SAF',
        'No_policy_floor': 'No policy'
    }
    
    # Define feedstock metrics to extract, organized by section
    # Maps from display name (for plot sheet) to actual name in Fitted_quantity.xlsx
    metrics_by_section = {
        'Feedstock quantity (billion)': [
            ('Q_soyoil (pounds)', 'soyoil'),
            ('Q_animal fat (pounds)', 'animal fat'),
            ('Q_corn (bushels)', 'corn'),
            ('Q_sugarcane (tons)', 'sugarcane')
        ]
    }
    
    # Read data from Fitted_quantity.xlsx
    print(f"\nReading Fitted_quantity.xlsx...")
    xl = pd.ExcelFile(source_file)
    all_data = {}
    
    for sheet_name in xl.sheet_names:
        if sheet_name in model_mapping:
            model_col_name = model_mapping[sheet_name]
            
            # Read the data
            df = pd.read_excel(xl, sheet_name=sheet_name)
            
            # Initialize the model's data dictionary
            if model_col_name not in all_data:
                all_data[model_col_name] = {}
            
            # Extract all metrics for this model
            for section, metrics in metrics_by_section.items():
                for item in metrics:
                    # Handle both tuple (display_name, actual_name) and string format
                    if isinstance(item, tuple):
                        display_name, actual_name = item
                    else:
                        display_name = actual_name = item
                    
                    matching_rows = df[df['Name'] == actual_name]
                    if not matching_rows.empty:
                        # Initialize nested dict for this metric
                        if display_name not in all_data[model_col_name]:
                            all_data[model_col_name][display_name] = {}
                        
                        # Extract all requested columns
                        for col in columns:
                            if col in df.columns:
                                try:
                                    value = matching_rows[col].iloc[0]
                                except Exception:
                                    value = None
                            else:
                                value = None
                            all_data[model_col_name][display_name][col] = value
            
            print(f"  ✓ {sheet_name} -> {model_col_name}")
    
    # Create the feedstock dataframe for selected column
    print(f"\nCompiling Feedstock sheet...")
    
    # Build the dataframe with sections using sel_col
    rows = []
    for section, metrics in metrics_by_section.items():
        # Add section header row
        section_row = {'Metric': section}
        for model_name in model_mapping.values():
            section_row[model_name] = None
        rows.append(section_row)
        
        # Add metric rows
        for item in metrics:
            # Handle both tuple (display_name, actual_name) and string format
            if isinstance(item, tuple):
                display_name, actual_name = item
            else:
                display_name = actual_name = item
            
            metric_row = {'Metric': display_name}
            for model_name in model_mapping.values():
                if model_name in all_data and display_name in all_data[model_name]:
                    value = all_data[model_name][display_name].get(sel_col, None)
                    metric_row[model_name] = value
                else:
                    metric_row[model_name] = None
            rows.append(metric_row)
    
    feedstock_df = pd.DataFrame(rows)
    
    # Reorganize columns in the specified order
    column_order = [
        'Metric',
        'Q mandate',
        'Nested D2',
        'Non-nested D2',
        'Nested D2 + stricter RFS',
        'Non-nested + stricter RFS',
        'Carbon Tax',
        'Aviation intensity standard',
        'Current Policy',
        'Current Policy + No LCFS',
        'Carbon tax+no SAF',
        'No policy'
    ]
    
    # Only include columns that exist in the dataframe
    existing_columns = [col for col in column_order if col in feedstock_df.columns]
    feedstock_df = feedstock_df[existing_columns]
    print(f"  ✓ Reorganized columns in specified order")
    
    # Write to Excel file
    print(f"\nWriting Feedstock sheet to {output_file}...")
    
    try:
        if os.path.exists(output_file):
            # Load existing workbook and replace/add the sheet
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', 
                              if_sheet_exists='replace') as writer:
                feedstock_df.to_excel(writer, sheet_name='Feedstock', index=False)
            print(f"✓ Updated 'Feedstock' sheet in {output_file}")
        else:
            # Create new workbook
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                feedstock_df.to_excel(writer, sheet_name='Feedstock', index=False)
            print(f"✓ Created {output_file} with 'Feedstock' sheet")
        
                # Now add the summary tables using openpyxl
        print(f"\nAdding feedstock summary tables...")
        wb = load_workbook(output_file)
        ws = wb['Feedstock']
        
        # ============ FIRST SUMMARY TABLE: Absolute Quantities ============
        summary_start_col = 14  # Column N
        summary_start_row = 2   # Row 2 (one row below header)
        
        # Define rows for the summary table with unit conversion factors
        summary_rows = [
            ('Soybean Oil', 'Q_soyoil (pounds)', 1),
            ('Fats, oils, greases', 'Q_animal fat (pounds)', 1),
            ('Corn', 'Q_corn (bushels)', 56),
            ('Sugarcane', 'Q_sugarcane (tons)', 2000),
        ]
        
        # Define columns for the summary table (policies)
        summary_columns = [
            'Current Policy',
            'SAF Credit',  # Q mandate
            'Nested D2',
            'Non-nested D2',  # D2 with aviation obligation
            'Nested D2 + stricter RFS',
            'Carbon Tax',  # Carbon Tax + SAF credit
            'Aviation intensity standard',
        ]
        
        # Map display names to actual column names in feedstock_df
        col_mapping = {
            'Current Policy': 'Current Policy',
            'SAF Credit': 'Q mandate',
            'Nested D2': 'Nested D2',
            'Non-nested D2': 'Non-nested D2',
            'Nested D2 + stricter RFS': 'Nested D2 + stricter RFS',
            'Carbon Tax': 'Carbon Tax',
            'Aviation intensity standard': 'Aviation intensity standard',
        }
        
        # Write summary table headers
        ws.cell(row=summary_start_row, column=summary_start_col).value = 'Feedstock'
        for col_idx, col_name in enumerate(summary_columns, start=1):
            ws.cell(row=summary_start_row, column=summary_start_col + col_idx).value = col_name
        
        # Write summary table data rows
        for row_idx, (summary_row_name, metric_name, conversion_factor) in enumerate(summary_rows, start=1):
            # Write row label in column N
            ws.cell(row=summary_start_row + row_idx, column=summary_start_col).value = summary_row_name
            
            # Write data for each column
            for col_idx, display_col_name in enumerate(summary_columns, start=1):
                actual_col_name = col_mapping.get(display_col_name, display_col_name)
                
                # Find the value in feedstock_df
                metric_rows = feedstock_df[feedstock_df['Metric'] == metric_name]
                if not metric_rows.empty and actual_col_name in metric_rows.columns:
                    policy_val = metric_rows[actual_col_name].iloc[0]
                    
                    # Apply conversion factor
                    if policy_val is not None:
                        try:
                            converted_value = float(policy_val) * conversion_factor
                            ws.cell(row=summary_start_row + row_idx, column=summary_start_col + col_idx).value = round(converted_value, 4)
                        except (ValueError, TypeError):
                            ws.cell(row=summary_start_row + row_idx, column=summary_start_col + col_idx).value = None
                    else:
                        ws.cell(row=summary_start_row + row_idx, column=summary_start_col + col_idx).value = None
                else:
                    ws.cell(row=summary_start_row + row_idx, column=summary_start_col + col_idx).value = None
        
        print(f"  ✓ Feedstock quantities table added at N2")
        
        # ============ SECOND SUMMARY TABLE: Change in Feedstock Demand ============
        # This table shows the difference between each policy and Current Policy
        change_start_row = summary_start_row + len(summary_rows) + 2  # Start 2 rows after the first table
        
        # Change columns (all policies except Current Policy)
        change_columns = [
            'SAF Credit',  # Q mandate
            'Nested D2',
            'Non-nested D2',  # D2 with aviation obligation
            'Nested D2 + stricter RFS',
            'Carbon Tax',  # Carbon Tax + SAF credit
            'Aviation intensity standard',
        ]
        
        # Write change table header
        ws.cell(row=change_start_row, column=summary_start_col).value = 'Change in Feedstock Demand (billion pounds)'
        for col_idx, col_name in enumerate(change_columns, start=1):
            ws.cell(row=change_start_row, column=summary_start_col + col_idx).value = col_name
        
        # Write change table data rows
        for row_idx, (summary_row_name, metric_name, conversion_factor) in enumerate(summary_rows, start=1):
            # Write row label in column N
            ws.cell(row=change_start_row + row_idx, column=summary_start_col).value = summary_row_name
            
            # Get Current Policy value (baseline)
            metric_rows = feedstock_df[feedstock_df['Metric'] == metric_name]
            if not metric_rows.empty and 'Current Policy' in metric_rows.columns:
                current_val = metric_rows['Current Policy'].iloc[0]
                try:
                    current_val = float(current_val) * conversion_factor
                except (ValueError, TypeError):
                    current_val = None
            else:
                current_val = None
            
            # Write difference data for each column
            for col_idx, display_col_name in enumerate(change_columns, start=1):
                actual_col_name = col_mapping.get(display_col_name, display_col_name)
                
                # Find the value in feedstock_df
                if not metric_rows.empty and actual_col_name in metric_rows.columns:
                    policy_val = metric_rows[actual_col_name].iloc[0]
                    
                    # Calculate difference with conversion factor
                    if policy_val is not None and current_val is not None:
                        try:
                            converted_value = float(policy_val) * conversion_factor
                            difference = converted_value - current_val
                            ws.cell(row=change_start_row + row_idx, column=summary_start_col + col_idx).value = round(difference, 4)
                        except (ValueError, TypeError):
                            ws.cell(row=change_start_row + row_idx, column=summary_start_col + col_idx).value = None
                    else:
                        ws.cell(row=change_start_row + row_idx, column=summary_start_col + col_idx).value = None
                else:
                    ws.cell(row=change_start_row + row_idx, column=summary_start_col + col_idx).value = None
        
        print(f"  ✓ Change in feedstock demand table added at N{change_start_row}")
        
        # Save the workbook
        wb.save(output_file)
        
        print(f"\n✓ Feedstock sheet generation complete!")
        print(f"  Main sheet shape: {feedstock_df.shape}")
        print(f"  Summary table 1 (Absolute quantities): N2 - {len(summary_rows)} rows × {len(summary_columns)} columns")
        print(f"  Summary table 2 (Change in demand): N{change_start_row} - {len(summary_rows)} rows × {len(change_columns)} columns")
        print(f"  Sections: {len(metrics_by_section)}")
        print(f"  Models: {len([col for col in feedstock_df.columns if col != 'Metric'])}")
        
        return True
        
    except Exception as e:
        print(f"\nError writing to Excel file: {e}")
        print("\nTrying alternative approach...")
        
        # Try creating a backup and new file
        if os.path.exists(output_file):
            import shutil
            backup_file = output_file.replace('.xlsx', '_backup.xlsx')
            shutil.copy(output_file, backup_file)
            print(f"  Backup saved to {backup_file}")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            feedstock_df.to_excel(writer, sheet_name='Feedstock', index=False)
        print(f"✓ Created new {output_file} with 'Feedstock' sheet")
        
        return True


def main():
    """Main function to run the script"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Generate Feedstock sheet in results_plot.xlsx'
    )
    parser.add_argument(
        '--results-dir', 
        type=str, 
        default='results',
        help='Path to results directory (default: results)'
    )
    parser.add_argument(
        '--output-file',
        type=str,
        default='results_plot.xlsx',
        help='Output Excel file name (default: results_plot.xlsx)'
    )
    parser.add_argument(
        '--columns',
        default=None,
        help='Comma-separated list of columns to extract (default: Mean)'
    )
    
    args = parser.parse_args()
    
    # Generate the Feedstock sheet
    success = generate_feedstock_sheet(args.results_dir, args.output_file, args.columns)
    
    if success:
        sys.exit(0)
    else:
        sys.exit(1)


if __name__ == '__main__':
    main()
