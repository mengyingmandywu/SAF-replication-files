"""
Generate Fuel Prices Sheet for results_plot.xlsx

This script reads fuel price metrics from Fuel_price.xlsx and compiles them
into a "Fuel prices" sheet, with a summary table showing price differences
relative to Current Policy.

Source Metrics:
From Fuel_price.xlsx:
   - Blended-diesel, Blended-gasoline, Aviation
   - E100, E10
   - Blended-diesel-CA, Blended-diesel-NC
   - Blended-gasoline-CA, Blended-gasoline-NC
   - Aviation-CA, Aviation-NC
   - E100-CA, E100-NC, E10-CA, E10-NC

Summary Table (starting at N1):
Displays price differences between each policy and Current Policy for:
   - Blended diesel (CA), Blended diesel (ROUS), E10 (CA), E10 (ROUS), 
     Aviation (CA), Aviation (ROUS)

Usage:
    python generate_fuel_prices_sheet1.py [--results-dir RESULTS_DIR] [--output-file OUTPUT_FILE] [--columns COLUMNS]
"""

import pandas as pd
import os
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def generate_fuel_prices_sheet(results_dir='results', output_file='results_plot.xlsx', columns=None):
    """
    Generate the Fuel prices sheet by reading metrics from Fuel_price.xlsx
    
    Parameters:
    -----------
    results_dir : str
        Path to the results directory (default: 'results')
    output_file : str
        Path to the output Excel file (default: 'results_plot.xlsx')
    columns : str or list
        Comma-separated column names or list of columns to extract (default: ['Mean'])
    """
    
    # Determine which columns to extract (default: Mean only)
    if columns is None:
        columns = ['Mean']
    elif isinstance(columns, str):
        columns = [c.strip() for c in columns.split(',') if c.strip()]
    
    # Use first column for sheet writing
    sel_col = columns[0]
    
    # Define source file
    source_file = os.path.join(results_dir, 'Fuel_price.xlsx')
    
    # Check if file exists
    if not os.path.exists(source_file):
        print(f"Error: Missing result file: {source_file}")
        print("\nPlease run the models first to generate all results.")
        return False
    
    print("Generating Fuel prices sheet...")
    print(f"Reading from: {results_dir}")
    print(f"Output file: {output_file}")
    print(f"Extracting columns: {columns}")
    
    # Model name mapping (from sheet names to column headers)
    model_mapping = {
        'Current_policy': 'Current Policy',
        'Pure_quantity': 'Q mandate',
        'Nested_D2_RV1': 'Nested D2',
        'Nested_D2_RV2': 'Nested D2 + stricter RFS',
        'Nonnested_D2': 'Non-nested D2',
        'Nonnested_D2_RV2': 'Non-nested + stricter RFS',
        'Carbon_tax': 'Carbon Tax',
        'Aviation_intensity_standard': 'Aviation intensity standard',
        'Current_policy_no_LCFS': 'Current Policy + No LCFS',
        'Carbon_tax_noSAF_floor': 'Carbon tax+no SAF',
        'No_policy_floor': 'No policy'
    }
    
    # Models that don't differentiate between CA and NC
    no_regional_differentiation = {'Carbon Tax', 'Carbon tax+no SAF', 'No policy'}
    
    # Define fuel price metrics to extract, organized by section
    metrics_by_section = {
        'Blended fuel ($/gallon)': [
            'Blended-diesel',
            'Blended-gasoline',
            'Aviation',
            'E100',
            'E10',
            'Blended-diesel-CA',
            'Blended-diesel-NC',
            'Blended-gasoline-CA',
            'Blended-gasoline-NC',
            'Aviation-CA',
            'Aviation-NC',
            'E100-CA',
            'E100-NC',
            'E10-CA',
            'E10-NC'
        ]
    }
    
    # Read data from Fuel_price.xlsx
    print(f"\nReading Fuel_price.xlsx...")
    xl = pd.ExcelFile(source_file)
    all_data = {}
    
    for sheet_name in xl.sheet_names:
        if sheet_name in model_mapping:
            model_col_name = model_mapping[sheet_name]
            
            # Read the data
            df = pd.read_excel(xl, sheet_name=sheet_name)
            
            # Initialize the model's data dictionary
            if model_col_name not in all_data:
                all_data[model_col_name] = {}
            
            # Extract all metrics for this model
            for section, metrics in metrics_by_section.items():
                for metric in metrics:
                    matching_rows = df[df['Name'] == metric]
                    if not matching_rows.empty:
                        # Initialize nested dict for this metric
                        if metric not in all_data[model_col_name]:
                            all_data[model_col_name][metric] = {}
                        
                        # Extract all requested columns
                        for col in columns:
                            if col in df.columns:
                                try:
                                    value = matching_rows[col].iloc[0]
                                except Exception:
                                    value = None
                            else:
                                value = None
                            all_data[model_col_name][metric][col] = value
            
            # For models without CA/NC differentiation, fill in regional variants
            if model_col_name in no_regional_differentiation:
                # List of (general, ca_variant, nc_variant) tuples
                regional_fills = [
                    ('Blended-diesel', 'Blended-diesel-CA', 'Blended-diesel-NC'),
                    ('Blended-gasoline', 'Blended-gasoline-CA', 'Blended-gasoline-NC'),
                    ('Aviation', 'Aviation-CA', 'Aviation-NC'),
                    ('E100', 'E100-CA', 'E100-NC'),
                    ('E10', 'E10-CA', 'E10-NC'),
                ]
                
                for general, ca_variant, nc_variant in regional_fills:
                    # If general exists and regional variants don't, fill them for all columns
                    if general in all_data[model_col_name]:
                        for col in columns:
                            general_value = all_data[model_col_name][general].get(col)
                            if ca_variant not in all_data[model_col_name]:
                                all_data[model_col_name][ca_variant] = {}
                            if nc_variant not in all_data[model_col_name]:
                                all_data[model_col_name][nc_variant] = {}
                            all_data[model_col_name][ca_variant][col] = general_value
                            all_data[model_col_name][nc_variant][col] = general_value
            
            print(f"  ✓ {sheet_name} -> {model_col_name}")
    
    # Create the fuel prices dataframe
    print(f"\nCompiling Fuel prices sheet...")
    
    # Build the dataframe with sections using sel_col
    rows = []
    for section, metrics in metrics_by_section.items():
        # Add section header row
        section_row = {'Metric': section}
        for model_name in model_mapping.values():
            section_row[model_name] = None
        rows.append(section_row)
        
        # Add metric rows
        for metric in metrics:
            metric_row = {'Metric': metric}
            for model_name in model_mapping.values():
                if model_name in all_data and metric in all_data[model_name]:
                    value = all_data[model_name][metric].get(sel_col, None)
                    metric_row[model_name] = value
                else:
                    metric_row[model_name] = None
            rows.append(metric_row)
    
    fuel_prices_df = pd.DataFrame(rows)
    
    # Reorganize columns in the specified order
    column_order = [
        'Metric',
        'Q mandate',
        'Nested D2',
        'Non-nested D2',
        'Nested D2 + stricter RFS',
        'Non-nested + stricter RFS',
        'Carbon Tax',
        'Aviation intensity standard',
        'Current Policy',
        'Current Policy + No LCFS',
        'Carbon tax+no SAF',
        'No policy'
    ]
    
    # Only include columns that exist in the dataframe
    existing_columns = [col for col in column_order if col in fuel_prices_df.columns]
    fuel_prices_df = fuel_prices_df[existing_columns]
    print(f"  ✓ Reorganized columns in specified order")
    
    # Write to Excel file
    print(f"\nWriting Fuel prices sheet to {output_file}...")
    
    try:
        if os.path.exists(output_file):
            # Load existing workbook and replace/add the sheet
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', 
                              if_sheet_exists='replace') as writer:
                fuel_prices_df.to_excel(writer, sheet_name='Fuel prices', index=False)
            print(f"✓ Updated 'Fuel prices' sheet in {output_file}")
        else:
            # Create new workbook
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                fuel_prices_df.to_excel(writer, sheet_name='Fuel prices', index=False)
            print(f"✓ Created {output_file} with 'Fuel prices' sheet")
        
        # Now add the summary table using openpyxl
        print(f"\nAdding price difference summary table at N1...")
        wb = load_workbook(output_file)
        ws = wb['Fuel prices']
        
        # Summary table parameters
        summary_start_col = 14  # Column N
        summary_start_row = 1   # Row 1
        
        # Define rows for the summary table (ROUS = Rest of US = NC)
        summary_rows = [
            ('Blended diesel (CA)', 'Blended-diesel-CA'),
            ('Blended diesel (ROUS)', 'Blended-diesel-NC'),
            ('E10 (CA)', 'E10-CA'),
            ('E10 (ROUS)', 'E10-NC'),
            ('Aviation (CA)', 'Aviation-CA'),
            ('Aviation (ROUS)', 'Aviation-NC'),
        ]
        
        # Define columns for the summary table (policies compared to Current Policy)
        # These are in the order specified by user
        summary_columns = [
            'SAF Credit',  # This should be mapped to Q mandate
            'Nested D2',
            'Non-nested D2',  # D2 with aviation obligation
            'Nested D2 + stricter RFS',
            'Carbon Tax',  # Carbon Tax + SAF credit
            'Aviation intensity standard',
            'Current Policy',
            'Current Policy + No LCFS',
            'Carbon tax+no SAF',
            'No policy'
        ]
        
        # Map display names to actual column names in fuel_prices_df
        col_mapping = {
            'SAF Credit': 'Q mandate',
            'Nested D2': 'Nested D2',
            'Non-nested D2': 'Non-nested D2',
            'Nested D2 + stricter RFS': 'Nested D2 + stricter RFS',
            'Carbon Tax': 'Carbon Tax',
            'Aviation intensity standard': 'Aviation intensity standard',
            'Current Policy': 'Current Policy',
            'Current Policy + No LCFS': 'Current Policy + No LCFS',
            'Carbon tax+no SAF': 'Carbon tax+no SAF',
            'No policy': 'No policy'
        }
        
        # Write summary table headers
        # Column headers start at N1
        ws.cell(row=summary_start_row, column=summary_start_col).value = 'Metric'
        for col_idx, col_name in enumerate(summary_columns, start=1):
            ws.cell(row=summary_start_row, column=summary_start_col + col_idx).value = col_name
        
        # Get Current Policy values for reference
        current_policy_values = {}
        for summary_row_name, metric_name in summary_rows:
            # Find the value in fuel_prices_df
            metric_rows = fuel_prices_df[fuel_prices_df['Metric'] == metric_name]
            if not metric_rows.empty:
                current_policy_val = metric_rows['Current Policy'].iloc[0]
                current_policy_values[metric_name] = current_policy_val
        
        # Write summary table data rows
        for row_idx, (summary_row_name, metric_name) in enumerate(summary_rows, start=1):
            # Write row label in column N
            ws.cell(row=summary_start_row + row_idx, column=summary_start_col).value = summary_row_name
            
            # Get the current policy value (baseline for difference calculation)
            current_policy_val = current_policy_values.get(metric_name, None)
            
            # Write data for each column
            for col_idx, display_col_name in enumerate(summary_columns, start=1):
                actual_col_name = col_mapping.get(display_col_name, display_col_name)
                
                # Find the value in fuel_prices_df
                metric_rows = fuel_prices_df[fuel_prices_df['Metric'] == metric_name]
                if not metric_rows.empty and actual_col_name in metric_rows.columns:
                    policy_val = metric_rows[actual_col_name].iloc[0]
                    
                    # Calculate difference (policy value - current policy value)
                    if current_policy_val is not None and policy_val is not None:
                        try:
                            difference = float(policy_val) - float(current_policy_val)
                            ws.cell(row=summary_start_row + row_idx, column=summary_start_col + col_idx).value = round(difference, 4)
                        except (ValueError, TypeError):
                            ws.cell(row=summary_start_row + row_idx, column=summary_start_col + col_idx).value = None
                    else:
                        ws.cell(row=summary_start_row + row_idx, column=summary_start_col + col_idx).value = None
                else:
                    ws.cell(row=summary_start_row + row_idx, column=summary_start_col + col_idx).value = None
        
        # Save the workbook
        wb.save(output_file)
        print(f"  ✓ Summary table added at N1")
        
        print(f"\n✓ Fuel prices sheet generation complete!")
        print(f"  Main sheet shape: {fuel_prices_df.shape}")
        print(f"  Summary table: {len(summary_rows)} rows × {len(summary_columns)} columns")
        print(f"  Sections: {len(metrics_by_section)}")
        print(f"  Models: {len([col for col in fuel_prices_df.columns if col != 'Metric'])}")
        
        return True
        
    except Exception as e:
        print(f"\nError writing to Excel file: {e}")
        print("\nTrying alternative approach...")
        
        # Try creating a backup and new file
        if os.path.exists(output_file):
            import shutil
            backup_file = output_file.replace('.xlsx', '_backup.xlsx')
            shutil.copy(output_file, backup_file)
            print(f"  Backup saved to {backup_file}")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            fuel_prices_df.to_excel(writer, sheet_name='Fuel prices', index=False)
        print(f"✓ Created new {output_file} with 'Fuel prices' sheet")
        
        return True


def main():
    """Main function to run the script"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Generate Fuel prices sheet in results_plot.xlsx'
    )
    parser.add_argument(
        '--results-dir', 
        type=str, 
        default='results',
        help='Path to results directory (default: results)'
    )
    parser.add_argument(
        '--output-file',
        type=str,
        default='results_plot.xlsx',
        help='Output Excel file name (default: results_plot.xlsx)'
    )
    parser.add_argument(
        '--columns',
        default=None,
        help='Comma-separated list of columns to extract (default: Mean)'
    )
    
    args = parser.parse_args()
    
    # Generate the Fuel prices sheet
    success = generate_fuel_prices_sheet(args.results_dir, args.output_file, args.columns)
    
    if success:
        sys.exit(0)
    else:
        sys.exit(1)


if __name__ == '__main__':
    main()
