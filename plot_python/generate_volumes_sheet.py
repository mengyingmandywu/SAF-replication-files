"""
Generate Volumes Sheet for results_plot.xlsx

This script reads aggregate quantity metrics from Fitted_quantity.xlsx and compiles them
into a "Volumes" sheet with a summary table showing fuel volumes by policy.

Source Metrics:
From Fitted_quantity.xlsx:
   - Aggregate quantity (billion gal): BD_all, RD_all, BBD_all, B0_all, E100_all, E0_all,
     SAF_HEFA_all, SAF_ETJ_all, SAF_all, J0_all, Blending ratio
   - Aggregate quantity_CA (billion gal): BD_CA, RD_CA, B0_CA, E100_CA, E0_CA, SAF_CA, J0_CA

Summary Table (starting at N2):
Displays fuel volumes by policy for key fuel types:
   - Gasoline (E0_all), Ethanol (E100_all), Diesel (B0_all), Renewable Diesel (RD_all),
     Biodiesel (BD_all), Jet Fuel (J0_all), SAF (HEFA) (SAF_HEFA_all), SAF (ATJ) (SAF_ETJ_all)

Usage:
    python generate_volumes_sheet1.py [--results-dir RESULTS_DIR] [--output-file OUTPUT_FILE] [--columns COLUMNS]
"""

import pandas as pd
import os
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def generate_volumes_sheet(results_dir='results', output_file='results_plot.xlsx', columns=None):
    """
    Generate the Volumes sheet by reading metrics from Fitted_quantity.xlsx
    
    Parameters:
    -----------
    results_dir : str
        Path to the results directory (default: 'results')
    output_file : str
        Path to the output Excel file (default: 'results_plot.xlsx')
    columns : str or list
        Comma-separated column names or list of columns to extract (default: ['Mean'])
    """
    
    # Determine which columns to extract (default: Mean only)
    if columns is None:
        columns = ['Mean']
    elif isinstance(columns, str):
        columns = [c.strip() for c in columns.split(',') if c.strip()]
    
    # Use first column for sheet writing
    sel_col = columns[0]
    
    # Define source file
    source_file = os.path.join(results_dir, 'Fitted_quantity.xlsx')
    
    # Check if file exists
    if not os.path.exists(source_file):
        print(f"Error: Missing result file: {source_file}")
        print("\nPlease run the models first to generate all results.")
        return False
    
    print("Generating Volumes sheet...")
    print(f"Reading from: {results_dir}")
    print(f"Output file: {output_file}")
    print(f"Extracting columns: {columns}")
    
    # Model name mapping (from sheet names to column headers)
    model_mapping = {
        'Current_policy': 'Current Policy',
        'Pure_quantity': 'Q mandate',
        'Nested_D2_RV1': 'Nested D2',
        'Nested_D2_RV2': 'Nested D2 + stricter RFS',
        'Nonnested_D2': 'Non-nested D2',
        'Nonnested_D2_RV2': 'Non-nested + stricter RFS',
        'Carbon_tax': 'Carbon Tax',
        'Aviation_intensity_standard': 'Aviation intensity standard',
        'Current_policy_no_LCFS': 'Current Policy + No LCFS',
        'Carbon_tax_noSAF_floor': 'Carbon tax+no SAF',
        'No_policy_floor': 'No policy'
    }
    
    # Define volume metrics to extract, organized by section
    # Maps from display name to actual name in Fitted_quantity.xlsx
    metrics_by_section = {
        'Aggregate quantity (billion gal)': [
            ('Gasoline', 'E0_all'),
            ('Ethanol', 'E100_all'),
            ('Diesel', 'B0_all'),
            ('Renewable Diesel', 'RD_all'),
            ('Biodiesel', 'BD_all'),
            ('Jet Fuel', 'J0_all'),
            ('SAF (HEFA)', 'SAF_HEFA_all'),
            ('SAF (ATJ)', 'SAF_ETJ_all'),
            ('BD_all', 'BD_all'),
            ('RD_all', 'RD_all'),
            ('BBD_all', 'BBD_all'),
            ('B0_all', 'B0_all'),
            ('E100_all', 'E100_all'),
            ('E0_all', 'E0_all'),
            ('SAF_HEFA_all', 'SAF_HEFA_all'),
            ('SAF_ETJ_all', 'SAF_ETJ_all'),
            ('SAF_all', 'SAF_all'),
            ('J0_all', 'J0_all'),
            ('Blending ratio', 'Blending ratio')
        ],
        'Aggregate quantity_CA (billion gal)': [
            'BD_CA',
            'RD_CA',
            'B0_CA',
            'E100_CA',
            'E0_CA',
            'SAF_CA',
            'J0_CA'
        ]
    }
    
    # Read data from Fitted_quantity.xlsx
    print(f"\nReading Fitted_quantity.xlsx...")
    xl = pd.ExcelFile(source_file)
    all_data = {}
    
    for sheet_name in xl.sheet_names:
        if sheet_name in model_mapping:
            model_col_name = model_mapping[sheet_name]
            
            # Read the data
            df = pd.read_excel(xl, sheet_name=sheet_name)
            
            # Initialize the model's data dictionary
            if model_col_name not in all_data:
                all_data[model_col_name] = {}
            
            # Extract all metrics for this model
            for section, metrics in metrics_by_section.items():
                for item in metrics:
                    # Handle both tuple (display_name, actual_name) and string format
                    if isinstance(item, tuple):
                        display_name, actual_name = item
                    else:
                        display_name = actual_name = item
                    
                    matching_rows = df[df['Name'] == actual_name]
                    if not matching_rows.empty:
                        # Initialize nested dict for this metric
                        if display_name not in all_data[model_col_name]:
                            all_data[model_col_name][display_name] = {}
                        
                        # Extract all requested columns
                        for col in columns:
                            if col in df.columns:
                                try:
                                    value = matching_rows[col].iloc[0]
                                except Exception:
                                    value = None
                            else:
                                value = None
                            all_data[model_col_name][display_name][col] = value
            
            print(f"  ✓ {sheet_name} -> {model_col_name}")
    
    # Create the volumes dataframe
    print(f"\nCompiling Volumes sheet...")
    
    # Build the dataframe with sections using sel_col
    rows = []
    for section, metrics in metrics_by_section.items():
        # Add section header row
        section_row = {'Metric': section}
        for model_name in model_mapping.values():
            section_row[model_name] = None
        rows.append(section_row)
        
        # Add metric rows
        for item in metrics:
            # Handle both tuple (display_name, actual_name) and string format
            if isinstance(item, tuple):
                display_name, actual_name = item
            else:
                display_name = actual_name = item
            
            metric_row = {'Metric': display_name}
            for model_name in model_mapping.values():
                if model_name in all_data and display_name in all_data[model_name]:
                    value = all_data[model_name][display_name].get(sel_col, None)
                    metric_row[model_name] = value
                else:
                    metric_row[model_name] = None
            rows.append(metric_row)
    
    volumes_df = pd.DataFrame(rows)
    
    # Reorganize columns in the specified order
    column_order = [
        'Metric',
        'Q mandate',
        'Nested D2',
        'Non-nested D2',
        'Nested D2 + stricter RFS',
        'Non-nested + stricter RFS',
        'Carbon Tax',
        'Aviation intensity standard',
        'Current Policy',
        'Current Policy + No LCFS',
        'Carbon tax+no SAF',
        'No policy'
    ]
    
    # Only include columns that exist in the dataframe
    existing_columns = [col for col in column_order if col in volumes_df.columns]
    volumes_df = volumes_df[existing_columns]
    print(f"  ✓ Reorganized columns in specified order")
    
    # Write to Excel file
    print(f"\nWriting Volumes sheet to {output_file}...")
    
    try:
        if os.path.exists(output_file):
            # Load existing workbook and replace/add the sheet
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', 
                              if_sheet_exists='replace') as writer:
                volumes_df.to_excel(writer, sheet_name='Volumes', index=False)
            print(f"✓ Updated 'Volumes' sheet in {output_file}")
        else:
            # Create new workbook
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                volumes_df.to_excel(writer, sheet_name='Volumes', index=False)
            print(f"✓ Created {output_file} with 'Volumes' sheet")
        
        # Now add the summary table using openpyxl
        print(f"\nAdding volume summary table...")
        wb = load_workbook(output_file)
        ws = wb['Volumes']
        
        # Summary table parameters
        summary_start_col = 14  # Column N
        summary_start_row = 2   # Row 2
        
        # Define rows for the summary table (policies)
        summary_rows = [
            ('Aviation intensity standard', 'Aviation intensity standard'),
            ('Carbon Tax + SAF Credit', 'Carbon Tax'),
            ('Nested D2 + Stricter RFS', 'Nested D2 + stricter RFS'),
            ('D2 with Aviation Obligation', 'Non-nested D2'),
            ('Nested D2', 'Nested D2'),
            ('SAF Credit', 'Q mandate'),
            ('Current Policy', 'Current Policy'),
            ('No Policy', 'No policy'),
        ]
        
        # Define columns for the summary table (fuel types)
        summary_columns = [
            ('Gasoline', 'E0_all'),
            ('Ethanol', 'E100_all'),
            ('Diesel', 'B0_all'),
            ('Renewable Diesel', 'RD_all'),
            ('Biodiesel', 'BD_all'),
            ('Jet Fuel', 'J0_all'),
            ('SAF (HEFA)', 'SAF_HEFA_all'),
            ('SAF (ATJ)', 'SAF_ETJ_all'),
        ]
        
        # Write summary table headers
        ws.cell(row=summary_start_row, column=summary_start_col).value = 'Policy'
        for col_idx, (display_col_name, actual_col_name) in enumerate(summary_columns, start=1):
            ws.cell(row=summary_start_row, column=summary_start_col + col_idx).value = display_col_name
        
        # Write summary table data rows
        for row_idx, (display_row_name, actual_col_name) in enumerate(summary_rows, start=1):
            # Write row label in column N
            ws.cell(row=summary_start_row + row_idx, column=summary_start_col).value = display_row_name
            
            # Write data for each fuel type column
            for col_idx, (display_col_name, fuel_metric_name) in enumerate(summary_columns, start=1):
                # Find the value in volumes_df for this policy and fuel type
                # Look for the row with the display name (e.g., "Gasoline")
                metric_rows = volumes_df[volumes_df['Metric'] == display_col_name]
                if not metric_rows.empty and actual_col_name in metric_rows.columns:
                    policy_val = metric_rows[actual_col_name].iloc[0]
                    
                    if policy_val is not None:
                        try:
                            ws.cell(row=summary_start_row + row_idx, column=summary_start_col + col_idx).value = round(float(policy_val), 4)
                        except (ValueError, TypeError):
                            ws.cell(row=summary_start_row + row_idx, column=summary_start_col + col_idx).value = None
                    else:
                        ws.cell(row=summary_start_row + row_idx, column=summary_start_col + col_idx).value = None
                else:
                    ws.cell(row=summary_start_row + row_idx, column=summary_start_col + col_idx).value = None
        
        # Save the workbook
        wb.save(output_file)
        print(f"  ✓ Summary table added at N{summary_start_row}")
        
        print(f"\n✓ Volumes sheet generation complete!")
        print(f"  Shape: {volumes_df.shape}")
        print(f"  Sections: {len(metrics_by_section)}")
        print(f"  Models: {len([col for col in volumes_df.columns if col != 'Metric'])}")
        print(f"  Summary table: {len(summary_rows)} policies × {len(summary_columns)} fuel types")
        
        return True
        
    except Exception as e:
        print(f"\nError writing to Excel file: {e}")
        print("\nTrying alternative approach...")
        
        # Try creating a backup and new file
        if os.path.exists(output_file):
            import shutil
            backup_file = output_file.replace('.xlsx', '_backup.xlsx')
            shutil.copy(output_file, backup_file)
            print(f"  Backup saved to {backup_file}")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            volumes_df.to_excel(writer, sheet_name='Volumes', index=False)
        print(f"✓ Created new {output_file} with 'Volumes' sheet")
        
        return True


def main():
    """Main function to run the script"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Generate Volumes sheet in results_plot.xlsx'
    )
    parser.add_argument(
        '--results-dir', 
        type=str, 
        default='results',
        help='Path to results directory (default: results)'
    )
    parser.add_argument(
        '--output-file',
        type=str,
        default='results_plot.xlsx',
        help='Output Excel file name (default: results_plot.xlsx)'
    )
    parser.add_argument(
        '--columns',
        default=None,
        help='Comma-separated list of columns to extract (default: Mean)'
    )
    
    args = parser.parse_args()
    
    # Generate the Volumes sheet
    success = generate_volumes_sheet(args.results_dir, args.output_file, args.columns)
    
    if success:
        sys.exit(0)
    else:
        sys.exit(1)


if __name__ == '__main__':
    main()
