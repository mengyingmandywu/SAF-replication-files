"""
Generate Subsidy Breakup Sheet for results_plot.xlsx

This script reads subsidy-related metrics from Total.xlsx and compiles them
into a "Subsidy breakup" sheet for detailed subsidy analysis.

Usage:
    python generate_subsidy_sheet1.py [--results-dir RESULTS_DIR] [--output-file OUTPUT_FILE] [--columns COLUMNS]
"""

import pandas as pd
import os
import sys

def generate_subsidy_sheet(results_dir='results', output_file='results_plot.xlsx', columns=None):
    """
    Generate the Subsidy breakup sheet by reading metrics from Total.xlsx
    
    Parameters:
    -----------
    results_dir : str
        Path to the results directory (default: 'results')
    output_file : str
        Path to the output Excel file (default: 'results_plot.xlsx')
    columns : str or list
        Comma-separated column names or list of columns to extract (default: ['Mean'])
    """
    
    # Determine which columns to extract (default: Mean only)
    if columns is None:
        columns = ['Mean']
    elif isinstance(columns, str):
        columns = [c.strip() for c in columns.split(',') if c.strip()]
    
    # Use first column for sheet writing
    sel_col = columns[0]
    
    # Define source file
    source_file = os.path.join(results_dir, 'Total.xlsx')
    
    # Check if file exists
    if not os.path.exists(source_file):
        print(f"Error: Missing result file: {source_file}")
        print("\nPlease run the models first to generate all results.")
        return False
    
    print("Generating Subsidy breakup sheet...")
    print(f"Reading from: {results_dir}")
    print(f"Output file: {output_file}")
    print(f"Extracting columns: {columns}")
    
    # Model name mapping (from sheet names to column headers)
    model_mapping = {
        'Current_policy': 'Current Policy',
        'Pure_quantity': 'Q mandate',
        'Nested_D2_RV1': 'Nested D2',
        'Nested_D2_RV2': 'Nested D2 + stricter RFS',
        'Nonnested_D2': 'Non-nested D2',
        'Nonnested_D2_RV2': 'Non-nested + stricter RFS',
        'Carbon_tax': 'Carbon Tax',
        'Aviation_intensity_standard': 'Aviation intensity standard',
        'Current_policy_no_LCFS': 'Current Policy + No LCFS',
        'Carbon_tax_noSAF_floor': 'Carbon tax+no SAF',
        'No_policy_floor': 'No policy'
    }
    
    # Define all metrics to extract from Total.xlsx, organized by section
    metrics_by_section = {
        'Subsidy (billion $)': [
            'Taxpayer',
            'IRA',
            '45Z',
            '45Q',
            'State_tax_credit',
            'Additional_SAF_tax'
        ],
        'RFS breakup credits (billion $)': [
            'RFS_credits_road_CA',
            'RFS_credits_road_NC',
            'RFS_credits_jet_CA',
            'RFS_credits_jet_NC',
            'RFS_credits_D_CA',
            'RFS_credits_D_NC',
            'RFS_credits_G_CA',
            'RFS_credits_G_NC',
            'RFS_credits_J_CA',
            'RFS_credits_J_NC'
        ],
        'RFS breakup deficits (billion $)': [
            'RFS_deficits_road_CA',
            'RFS_deficits_road_NC',
            'RFS_deficits_jet_CA',
            'RFS_deficits_jet_NC',
            'RFS_deficits_D_CA',
            'RFS_deficits_D_NC',
            'RFS_deficits_G_CA',
            'RFS_deficits_G_NC',
            'RFS_deficits_J_CA',
            'RFS_deficits_J_NC'
        ],
        'CA LCFS breakup (billion $)': [
            'LCFS_CA_biofuel_credits',
            'LCFS_CA_other_credits',
            'LCFS_CA_all_credits',
            'LCFS_CA_biofuel_credits_D',
            'LCFS_CA_biofuel_credits_G',
            'LCFS_CA_biofuel_credits_jet',
            'LCFS_CA_deficits_D',
            'LCFS_CA_deficits_G'
        ],
        'SAF intensity breakup (billion $)': [
            'LCFS_all_jet',
            'LCFS_all_credits_jet_CA',
            'LCFS_all_credits_jet_NC',
            'LCFS_all_deficits_jet_CA',
            'LCFS_all_deficits_jet_NC',
            'Total'
        ],
        'Taxpayer breakup ($ billion)': [
            'Taxpayer_D',
            'Taxpayer_G',
            'Taxpayer_J'
        ],
        'Carbon tax revenue ($ billion)': [
            'Carbon_tax_D',
            'Carbon_tax_G',
            'Carbon_tax_J'
        ]
    }
    
    # Read data from Total.xlsx
    print(f"\nReading Total.xlsx...")
    xl = pd.ExcelFile(source_file)
    all_data = {}
    
    for sheet_name in xl.sheet_names:
        if sheet_name in model_mapping:
            model_col_name = model_mapping[sheet_name]
            
            # Read the data
            df = pd.read_excel(xl, sheet_name=sheet_name)
            
            # Initialize the model's data dictionary
            if model_col_name not in all_data:
                all_data[model_col_name] = {}
            
            # Extract all metrics for this model
            for section, metrics in metrics_by_section.items():
                for metric in metrics:
                    matching_rows = df[df['Name'] == metric]
                    if not matching_rows.empty:
                        # Initialize nested dict for this metric
                        if metric not in all_data[model_col_name]:
                            all_data[model_col_name][metric] = {}
                        
                        # Extract all requested columns
                        for col in columns:
                            if col in df.columns:
                                try:
                                    value = matching_rows[col].iloc[0]
                                except Exception:
                                    value = None
                            else:
                                value = None
                            all_data[model_col_name][metric][col] = value
            
            print(f"  ✓ {sheet_name} -> {model_col_name}")
    
    # Create the subsidy breakup dataframe
    print(f"\nCompiling Subsidy breakup sheet...")
    
    # Build the dataframe with calculated rows first using sel_col
    rows = []
    
    # Row 1: Taxpayers (equals Taxpayer)
    taxpayers_row = {'Metric': 'Taxpayers'}
    for model_name in model_mapping.values():
        if model_name in all_data:
            taxpayers_row[model_name] = all_data[model_name].get('Taxpayer', {}).get(sel_col, None)
        else:
            taxpayers_row[model_name] = None
    rows.append(taxpayers_row)

    # Add header row for the calculated section (moved after Taxpayers)
    header_row = {'Metric': 'RFS, LCFS, Aviation intensity standards obligation'}
    for model_name in model_mapping.values():
        header_row[model_name] = None
    rows.append(header_row)
    
    # Row 2: Diesel (CA) = -RFS_deficits_D_CA - LCFS_CA_deficits_D
    diesel_ca_row = {'Metric': 'Diesel (CA)'}
    for model_name in model_mapping.values():
        if model_name in all_data:
            rfs_deficit = all_data[model_name].get('RFS_deficits_D_CA', {}).get(sel_col, 0)
            lcfs_deficit = all_data[model_name].get('LCFS_CA_deficits_D', {}).get(sel_col, 0)
            rfs_val = rfs_deficit if (rfs_deficit is not None and pd.notna(rfs_deficit)) else 0
            lcfs_val = lcfs_deficit if (lcfs_deficit is not None and pd.notna(lcfs_deficit)) else 0
            diesel_ca_row[model_name] = -rfs_val - lcfs_val
        else:
            diesel_ca_row[model_name] = None
    rows.append(diesel_ca_row)
    
    # Row 3: Diesel (ROUS) = -RFS_deficits_D_NC
    diesel_rous_row = {'Metric': 'Diesel (ROUS)'}
    for model_name in model_mapping.values():
        if model_name in all_data:
            rfs_deficit = all_data[model_name].get('RFS_deficits_D_NC', {}).get(sel_col, 0)
            rfs_val = rfs_deficit if (rfs_deficit is not None and pd.notna(rfs_deficit)) else 0
            diesel_rous_row[model_name] = -rfs_val
        else:
            diesel_rous_row[model_name] = None
    rows.append(diesel_rous_row)
    
    # Row 4: Gasoline (CA) = -RFS_deficits_G_CA - LCFS_CA_deficits_G
    gasoline_ca_row = {'Metric': 'Gasoline (CA)'}
    for model_name in model_mapping.values():
        if model_name in all_data:
            rfs_deficit = all_data[model_name].get('RFS_deficits_G_CA', {}).get(sel_col, 0)
            lcfs_deficit = all_data[model_name].get('LCFS_CA_deficits_G', {}).get(sel_col, 0)
            rfs_val = rfs_deficit if (rfs_deficit is not None and pd.notna(rfs_deficit)) else 0
            lcfs_val = lcfs_deficit if (lcfs_deficit is not None and pd.notna(lcfs_deficit)) else 0
            gasoline_ca_row[model_name] = -rfs_val - lcfs_val
        else:
            gasoline_ca_row[model_name] = None
    rows.append(gasoline_ca_row)
    
    # Row 5: Gasoline (ROUS) = -RFS_deficits_G_NC
    gasoline_rous_row = {'Metric': 'Gasoline (ROUS)'}
    for model_name in model_mapping.values():
        if model_name in all_data:
            rfs_deficit = all_data[model_name].get('RFS_deficits_G_NC', {}).get(sel_col, 0)
            rfs_val = rfs_deficit if (rfs_deficit is not None and pd.notna(rfs_deficit)) else 0
            gasoline_rous_row[model_name] = -rfs_val
        else:
            gasoline_rous_row[model_name] = None
    rows.append(gasoline_rous_row)
    
    # Row 6: Air passengers (CA) = -RFS_deficits_jet_CA - LCFS_all_deficits_jet_CA
    air_ca_row = {'Metric': 'Air passengers (CA)'}
    for model_name in model_mapping.values():
        if model_name in all_data:
            rfs_deficit = all_data[model_name].get('RFS_deficits_jet_CA', {}).get(sel_col, 0)
            lcfs_deficit = all_data[model_name].get('LCFS_all_deficits_jet_CA', {}).get(sel_col, 0)
            rfs_val = rfs_deficit if (rfs_deficit is not None and pd.notna(rfs_deficit)) else 0
            lcfs_val = lcfs_deficit if (lcfs_deficit is not None and pd.notna(lcfs_deficit)) else 0
            air_ca_row[model_name] = -rfs_val - lcfs_val
        else:
            air_ca_row[model_name] = None
    rows.append(air_ca_row)
    
    # Row 7: Air passengers (NC) = -RFS_deficits_jet_NC - LCFS_all_deficits_jet_NC
    air_nc_row = {'Metric': 'Air passengers (NC)'}
    for model_name in model_mapping.values():
        if model_name in all_data:
            rfs_deficit = all_data[model_name].get('RFS_deficits_jet_NC', {}).get(sel_col, 0)
            lcfs_deficit = all_data[model_name].get('LCFS_all_deficits_jet_NC', {}).get(sel_col, 0)
            rfs_val = rfs_deficit if (rfs_deficit is not None and pd.notna(rfs_deficit)) else 0
            lcfs_val = lcfs_deficit if (lcfs_deficit is not None and pd.notna(lcfs_deficit)) else 0
            air_nc_row[model_name] = -rfs_val - lcfs_val
        else:
            air_nc_row[model_name] = None
    rows.append(air_nc_row)
    
    # After calculated rows, insert Carbon tax revenue rows (so they appear directly below Air passengers (NC))
    carbon_tax_d_row = {'Metric': 'Carbon_tax_D'}
    carbon_tax_g_row = {'Metric': 'Carbon_tax_G'}
    carbon_tax_j_row = {'Metric': 'Carbon_tax_J'}
    for model_name in model_mapping.values():
        if model_name in all_data:
            carbon_tax_d_row[model_name] = all_data[model_name].get('Carbon_tax_D', {}).get(sel_col, None)
            carbon_tax_g_row[model_name] = all_data[model_name].get('Carbon_tax_G', {}).get(sel_col, None)
            carbon_tax_j_row[model_name] = all_data[model_name].get('Carbon_tax_J', {}).get(sel_col, None)
        else:
            carbon_tax_d_row[model_name] = None
            carbon_tax_g_row[model_name] = None
            carbon_tax_j_row[model_name] = None

    rows.append(carbon_tax_d_row)
    rows.append(carbon_tax_g_row)
    rows.append(carbon_tax_j_row)

    # Add a blank row after calculated/carbon-tax section
    blank_row = {'Metric': ''}
    for model_name in model_mapping.values():
        blank_row[model_name] = None
    rows.append(blank_row)

    # Now add the existing sections, but skip the Carbon tax revenue section (we already included its rows)
    for section, metrics in metrics_by_section.items():
        if section == 'Carbon tax revenue ($ billion)':
            continue
        # Add section header row
        section_row = {'Metric': section}
        for model_name in model_mapping.values():
            section_row[model_name] = None
        rows.append(section_row)
        
        # Add metric rows
        for metric in metrics:
            metric_row = {'Metric': metric}
            for model_name in model_mapping.values():
                if model_name in all_data and metric in all_data[model_name]:
                    value = all_data[model_name][metric].get(sel_col, None)
                    metric_row[model_name] = value
                else:
                    metric_row[model_name] = None
            rows.append(metric_row)
    
    subsidy_df = pd.DataFrame(rows)
    
    # Reorganize columns in the specified order
    column_order = [
        'Metric',
        'Q mandate',
        'Nested D2',
        'Non-nested D2',
        'Nested D2 + stricter RFS',
        'Non-nested + stricter RFS',
        'Carbon Tax',
        'Aviation intensity standard',
        'Current Policy',
        'Current Policy + No LCFS',
        'Carbon tax+no SAF',
        'No policy'
    ]
    
    # Only include columns that exist in the dataframe
    existing_columns = [col for col in column_order if col in subsidy_df.columns]
    subsidy_df = subsidy_df[existing_columns]
    print(f"  ✓ Reorganized columns in specified order")
    
    # Write to Excel file
    print(f"\nWriting Subsidy breakup sheet to {output_file}...")
    
    try:
        if os.path.exists(output_file):
            # Load existing workbook and replace/add the sheet
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', 
                              if_sheet_exists='replace') as writer:
                subsidy_df.to_excel(writer, sheet_name='Subsidy breakup', index=False)
            print(f"✓ Updated 'Subsidy breakup' sheet in {output_file}")
        else:
            # Create new workbook
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                subsidy_df.to_excel(writer, sheet_name='Subsidy breakup', index=False)
            print(f"✓ Created {output_file} with 'Subsidy breakup' sheet")
        
        # Also add a compact summary table starting at N1 in the 'Subsidy breakup' sheet
        try:
            from openpyxl import load_workbook
            wb = load_workbook(output_file)
            ws = wb['Subsidy breakup']

            # Define the column order requested (starting in column O)
            summary_headers = [
                'Aviation Intensity Std',
                'Carbon Tax + SAF credit',
                'Nested D2 + stricter RFS',
                'D2 with aviation obligation',
                'Nested D2',
                'SAF Credit',
                'Current Policy'
            ]

            # Map these display headers to subsidy_df column names (if different)
            col_map = {
                'Aviation Intensity Std': 'Aviation intensity standard',
                'Carbon Tax + SAF credit': 'Carbon Tax',
                'Nested D2 + stricter RFS': 'Nested D2 + stricter RFS',
                'D2 with aviation obligation': 'Non-nested D2',
                'Nested D2': 'Nested D2',
                'SAF Credit': 'Q mandate',
                'Current Policy': 'Current Policy'
            }

            # Write headers in row 1, starting at column O (15)
            for i, h in enumerate(summary_headers, start=15):
                ws.cell(row=1, column=i, value=h)

            # Rows to write (row labels in column N)
            row_labels = [
                'Taxpayers',
                'Diesel (CA)',
                'Diesel (ROUS)',
                'Gasoline (CA)',
                'Gasoline (ROUS)',
                'Air passengers (CA)',
                'Air passengers (NC)'
            ]

            # Place row labels in column N (14) at rows 2..8
            for idx, label in enumerate(row_labels, start=2):
                ws.cell(row=idx, column=14, value=label)

            # Build a quick lookup from subsidy_df (Metric -> row index)
            metric_to_row = {r['Metric']: i+2 for i, r in subsidy_df.iterrows()}  # +2 for header and 0-index

            # Fill values
            for col_idx, header in enumerate(summary_headers, start=15):
                src_col = col_map.get(header)
                for row_idx, label in enumerate(row_labels, start=2):
                    # find metric row in subsidy_df
                    match = subsidy_df[subsidy_df['Metric'] == label]
                    val = None
                    if not match.empty and src_col in subsidy_df.columns:
                        val = match.iloc[0][src_col]

                    # If column is Carbon Tax + SAF credit, for certain rows add Carbon_tax_*
                    if header == 'Carbon Tax + SAF credit' and label in ('Diesel (ROUS)', 'Gasoline (ROUS)', 'Air passengers (NC)'):
                        # determine which carbon tax component to add
                        tax_val = None
                        if label == 'Diesel (ROUS)':
                            # add Carbon_tax_D from the same model column (src_col)
                            ct_match = subsidy_df[subsidy_df['Metric'] == 'Carbon_tax_D']
                            if not ct_match.empty and src_col in ct_match.columns:
                                tax_val = ct_match.iloc[0][src_col]
                        elif label == 'Gasoline (ROUS)':
                            ct_match = subsidy_df[subsidy_df['Metric'] == 'Carbon_tax_G']
                            if not ct_match.empty and src_col in ct_match.columns:
                                tax_val = ct_match.iloc[0][src_col]
                        elif label == 'Air passengers (NC)':
                            ct_match = subsidy_df[subsidy_df['Metric'] == 'Carbon_tax_J']
                            if not ct_match.empty and src_col in ct_match.columns:
                                tax_val = ct_match.iloc[0][src_col]

                        # if tax_val found, add to val (treat None as 0)
                        if tax_val is not None and pd.notna(tax_val):
                            base = val if (val is not None and pd.notna(val)) else 0
                            val = base + tax_val

                    # if value is numeric, write it
                    if val is not None and pd.notna(val):
                        ws.cell(row=row_idx, column=col_idx, value=val)
                        # format numbers with two decimals
                        ws.cell(row=row_idx, column=col_idx).number_format = '#,##0.00'

            wb.save(output_file)
            print(f"✓ Added compact summary table at N1 in 'Subsidy breakup' sheet")
        except Exception as e:
            print(f"Warning: could not write compact summary table: {e}")
        
        print(f"\n✓ Subsidy breakup sheet generation complete!")
        print(f"  Shape: {subsidy_df.shape}")
        print(f"  Sections: {len(metrics_by_section)}")
        print(f"  Models: {len([col for col in subsidy_df.columns if col != 'Metric'])}")
        
        return True
        
    except Exception as e:
        print(f"\nError writing to Excel file: {e}")
        print("\nTrying alternative approach...")
        
        # Try creating a backup and new file
        if os.path.exists(output_file):
            import shutil
            backup_file = output_file.replace('.xlsx', '_backup.xlsx')
            shutil.copy(output_file, backup_file)
            print(f"  Backup saved to {backup_file}")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            subsidy_df.to_excel(writer, sheet_name='Subsidy breakup', index=False)
        print(f"✓ Created new {output_file} with 'Subsidy breakup' sheet")
        
        return True


def main():
    """Main function to run the script"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Generate Subsidy breakup sheet in results_plot.xlsx'
    )
    parser.add_argument(
        '--results-dir', 
        type=str, 
        default='results',
        help='Path to results directory (default: results)'
    )
    parser.add_argument(
        '--output-file',
        type=str,
        default='results_plot.xlsx',
        help='Output Excel file name (default: results_plot.xlsx)'
    )
    parser.add_argument(
        '--columns',
        default=None,
        help='Comma-separated list of columns to extract (default: Mean)'
    )
    
    args = parser.parse_args()
    
    # Generate the Subsidy breakup sheet
    success = generate_subsidy_sheet(args.results_dir, args.output_file, args.columns)
    
    if success:
        sys.exit(0)
    else:
        sys.exit(1)


if __name__ == '__main__':
    main()
