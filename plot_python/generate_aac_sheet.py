"""
Generate AAC (Abatement Cost Curve) Sheet for results_plot.xlsx

This script reads specific metrics from multiple result files and compiles them
into an "AAC" sheet for abatement cost curve analysis.

Source Metrics:
1. From Total.xlsx:
   - Emissions_D, Emissions_G, Emissions_J, Emissions_total
   - Taxpayer, Taxpayer_D, Taxpayer_G, Taxpayer_J
   - Carbon_tax_total, Carbon_tax_D, Carbon_tax_G, Carbon_tax_J

2. From Fuel_price.xlsx:
   - Blended-diesel-CA, Blended-diesel-NC
   - Blended-gasoline-CA, Blended-gasoline-NC
   - Aviation-CA, Aviation-NC

3. From Fitted_quantity.xlsx:
   - D_CA, G_CA, J_CA, D_NC, G_NC, J_NC

Usage:
    python generate_aac_sheet.py [--results-dir RESULTS_DIR] [--output-file OUTPUT_FILE]
"""

import pandas as pd
import os
import sys

def generate_aac_sheet(results_dir='results', output_file='results_plot.xlsx', columns=None):
    """
    Generate the AAC sheet by reading specific metrics from result files
    
    Parameters:
    -----------
    results_dir : str
        Path to the results directory (default: 'results')
    output_file : str
        Path to the output Excel file (default: 'results_plot.xlsx')
    """
    
    # Define source files
    source_files = {
        'Total': os.path.join(results_dir, 'Total.xlsx'),
        'Fuel_price': os.path.join(results_dir, 'Fuel_price.xlsx'),
        'Fitted_quantity': os.path.join(results_dir, 'Fitted_quantity.xlsx')
    }
    
    # Check if all files exist
    missing_files = []
    for name, path in source_files.items():
        if not os.path.exists(path):
            missing_files.append(path)
    
    if missing_files:
        print(f"Error: Missing result files:")
        for f in missing_files:
            print(f"  - {f}")
        print("\nPlease run the models first to generate all results.")
        return False
    
    print("Generating AAC sheet...")
    print(f"Reading from: {results_dir}")
    print(f"Output file: {output_file}")
    
    # Model name mapping (from sheet names to column headers)
    model_mapping = {
        'Current_policy': 'Current Policy',
        'Pure_quantity': 'Q mandate',
        'Nested_D2_RV1': 'Nested D2',
        'Nested_D2_RV2': 'Nested D2 + stricter RFS',
        'Nonnested_D2': 'Non-nested D2',
        'Nonnested_D2_RV2': 'Non-nested + stricter RFS',
        'Carbon_tax': 'Carbon Tax',
        'Aviation_intensity_standard': 'Aviation intensity standard',
        'Current_policy_no_LCFS': 'Current Policy + No LCFS',
        'Carbon_tax_noSAF_floor': 'Carbon tax+no SAF',
        'No_policy_floor': 'No policy'
    }
    
    # Define which metrics to extract from each file
    metrics_to_extract = {
        'Total': [
            'Emissions_D', 'Emissions_G', 'Emissions_J', 'Emissions_total',
            'Taxpayer', 'Taxpayer_D', 'Taxpayer_G', 'Taxpayer_J',
            'Carbon_tax_total', 'Carbon_tax_D', 'Carbon_tax_G', 'Carbon_tax_J'
        ],
        'Fuel_price': [
            'Blended fuel price',  # Header row with no values
            'Blended-diesel','Blended-gasoline','Aviation',
            'Blended-diesel-CA', 'Blended-diesel-NC',
            'Blended-gasoline-CA', 'Blended-gasoline-NC',
            'Aviation-CA', 'Aviation-NC'
        ],
        'Fitted_quantity': [
            'Quantity',  # Header row with no values
            'D','G','J',
            'D_CA', 'G_CA', 'J_CA', 'D_NC', 'G_NC', 'J_NC'
        ]
    }
    # Determine which columns to extract (default: Mean only)
    if columns is None:
        columns = ['Mean']
    elif isinstance(columns, str):
        columns = [c.strip() for c in columns.split(',') if c.strip()]

    # Read all data from source files
    # Structure: all_data[model_name][metric][col] = value
    all_data = {}
    
    # Define header rows that should have no values
    header_rows = {'Blended fuel price', 'Quantity'}
    
    for source_name, file_path in source_files.items():
        print(f"\nReading {source_name}...")
        xl = pd.ExcelFile(file_path)
        
        for sheet_name in xl.sheet_names:
            if sheet_name in model_mapping:
                model_col_name = model_mapping[sheet_name]
                
                # Read the data
                df = pd.read_excel(xl, sheet_name=sheet_name)
                
                # Initialize the model's data dictionary if not exists
                if model_col_name not in all_data:
                    all_data[model_col_name] = {}
                
                # Extract only the metrics we need from this source
                if source_name in metrics_to_extract:
                    metrics = metrics_to_extract[source_name]
                    for metric in metrics:
                        # Skip header rows - they'll be added later with no values
                        if metric in header_rows:
                            continue
                        matching_rows = df[df['Name'] == metric]
                        # Initialize metric dict
                        if metric not in all_data[model_col_name]:
                            all_data[model_col_name][metric] = {}

                        if not matching_rows.empty:
                            for col in columns:
                                if col in df.columns:
                                    try:
                                        val = matching_rows[col].iloc[0]
                                    except Exception:
                                        val = None
                                else:
                                    val = None
                                all_data[model_col_name][metric][col] = val
                        else:
                            # No row found: set None for all requested columns
                            for col in columns:
                                all_data[model_col_name][metric][col] = None
                
                # Fill in regional fuel prices from general fuel prices if missing
                if source_name == 'Fuel_price':
                    # If regional prices are missing, use the general prices
                    fuel_mappings = {
                        ('Blended-diesel', 'Blended-diesel-CA', 'Blended-diesel-NC'),
                        ('Blended-gasoline', 'Blended-gasoline-CA', 'Blended-gasoline-NC'),
                        ('Aviation', 'Aviation-CA', 'Aviation-NC')
                    }
                    
                    for general, ca_metric, nc_metric in fuel_mappings:
                        if general in all_data[model_col_name]:
                            for col in columns:
                                general_value = all_data[model_col_name][general].get(col, None)
                                # Fill CA if missing
                                if ca_metric not in all_data[model_col_name] or pd.isna(all_data[model_col_name].get(ca_metric, {}).get(col)):
                                    # ensure dicts exist
                                    if ca_metric not in all_data[model_col_name]:
                                        all_data[model_col_name][ca_metric] = {}
                                    all_data[model_col_name][ca_metric][col] = general_value
                                # Fill NC if missing
                                if nc_metric not in all_data[model_col_name] or pd.isna(all_data[model_col_name].get(nc_metric, {}).get(col)):
                                    if nc_metric not in all_data[model_col_name]:
                                        all_data[model_col_name][nc_metric] = {}
                                    all_data[model_col_name][nc_metric][col] = general_value
                
                # Split total fuel quantities into CA/NC based on CA fuel demand percentages
                if source_name == 'Fitted_quantity':
                    ca_fuel_demand_pct = {
                        'D': 0.0990027,  # Diesel: 9.90027%
                        'G': 0.0904482,  # Gasoline: 9.04482%
                        'J': 0.1700000   # Jet: 17.00000%
                    }
                    
                    for fuel_type, ca_pct in ca_fuel_demand_pct.items():
                        ca_metric = f'{fuel_type}_CA'
                        nc_metric = f'{fuel_type}_NC'

                        # If we have total but missing regional breakdown
                        if fuel_type in all_data[model_col_name]:
                            for col in columns:
                                total_value = all_data[model_col_name][fuel_type].get(col, None)
                                # initialize dicts if needed
                                if ca_metric not in all_data[model_col_name]:
                                    all_data[model_col_name][ca_metric] = {}
                                if nc_metric not in all_data[model_col_name]:
                                    all_data[model_col_name][nc_metric] = {}

                                # Fill CA if missing
                                if pd.isna(all_data[model_col_name][ca_metric].get(col)):
                                    try:
                                        all_data[model_col_name][ca_metric][col] = (total_value * ca_pct) if total_value is not None else None
                                    except Exception:
                                        all_data[model_col_name][ca_metric][col] = None

                                # Fill NC if missing (NC = Total - CA)
                                if pd.isna(all_data[model_col_name][nc_metric].get(col)):
                                    ca_value = all_data[model_col_name][ca_metric].get(col, (total_value * ca_pct) if total_value is not None else None)
                                    try:
                                        all_data[model_col_name][nc_metric][col] = (total_value - ca_value) if (total_value is not None and ca_value is not None) else None
                                    except Exception:
                                        all_data[model_col_name][nc_metric][col] = None
                
                print(f"  ✓ {sheet_name} -> {model_col_name}")
    
    # Compile all metrics in order
    all_metrics = []
    for source_name in ['Total', 'Fuel_price', 'Fitted_quantity']:
        if source_name in metrics_to_extract:
            all_metrics.extend(metrics_to_extract[source_name])
    
    print(f"\nCompiling AAC sheet with {len(all_metrics)} metrics...")
    
    # For each requested column, build a separate AAC dataframe and write as its own sheet
    aac_dfs = {}
    for col in columns:
        aac_data = {'Metric': all_metrics}
        for model_name in model_mapping.values():
            model_values = []
            if model_name in all_data:
                for metric in all_metrics:
                    metric_dict = all_data[model_name].get(metric, {})
                    value = metric_dict.get(col, None)
                    model_values.append(value)
            else:
                model_values = [None] * len(all_metrics)
            aac_data[model_name] = model_values
        aac_dfs[col] = pd.DataFrame(aac_data)
        print(f"  ✓ Built AAC dataframe for column: {col}")
    
    # Add calculated rows at the top for each column's dataframe
    print(f"\nAdding calculated rows (for each requested column)...")
    # Helper to insert calculated rows into aac dataframe for a specific col
    def insert_calculated_rows(aac_df_col, col_name):
        # Row 1: "Transportation sector" (no values, just label)
        transportation_row = {'Metric': 'Transportation sector'}
        for model_name in model_mapping.values():
            transportation_row[model_name] = None

        # Get Emissions_total index
        emissions_total_idx = all_metrics.index('Emissions_total') if 'Emissions_total' in all_metrics else None

        co2_reduction_row = {'Metric': 'MTs of CO2 reduction'}
        pct_reduction_row = {'Metric': '% of CO2 reduction'}

        no_policy_emissions = None
        if emissions_total_idx is not None and 'No policy' in aac_df_col.columns:
            no_policy_emissions = aac_df_col.loc[emissions_total_idx, 'No policy']

        for model_name in model_mapping.values():
            policy_emissions = aac_df_col.loc[emissions_total_idx, model_name] if emissions_total_idx is not None and model_name in aac_df_col.columns else None
            if no_policy_emissions is not None and pd.notna(policy_emissions):
                reduction = no_policy_emissions - policy_emissions
                co2_reduction_row[model_name] = reduction
                pct_reduction_row[model_name] = (reduction / no_policy_emissions) if no_policy_emissions != 0 else None
            else:
                co2_reduction_row[model_name] = None
                pct_reduction_row[model_name] = None

        # Row 4: Equivalent Variation (no values)
        ev_row = {'Metric': 'Equivalent Variation'}
        for model_name in model_mapping.values():
            ev_row[model_name] = None

        # Row 5: Fuel Expenditure at current prices, no policy quantities
        fuel_exp_row = {'Metric': 'Fuel Expenditure at current prices, no policy quantities'}

        # Get no policy quantities for calculation from all_data
        no_policy_quantities = {}
        if 'No policy' in all_data:
            for fuel_region in ['D_CA', 'D_NC', 'G_CA', 'G_NC', 'J_CA', 'J_NC']:
                no_policy_quantities[fuel_region] = all_data['No policy'].get(fuel_region, {}).get(col_name, 0)

        fuel_price_mapping = {
            'D_CA': 'Blended-diesel-CA',
            'D_NC': 'Blended-diesel-NC',
            'G_CA': 'Blended-gasoline-CA',
            'G_NC': 'Blended-gasoline-NC',
            'J_CA': 'Aviation-CA',
            'J_NC': 'Aviation-NC'
        }

        for model_name in model_mapping.values():
            total_fuel_exp = 0
            for fuel_region, price_metric in fuel_price_mapping.items():
                price = None
                if model_name in all_data and price_metric in all_data[model_name]:
                    price = all_data[model_name][price_metric].get(col_name, None)
                quantity = no_policy_quantities.get(fuel_region, 0)
                if price is not None and quantity is not None:
                    try:
                        total_fuel_exp += price * quantity
                    except Exception:
                        pass
            fuel_exp_row[model_name] = total_fuel_exp if total_fuel_exp != 0 else None

        # Row 6: Government spending (Tax subsidies) = Taxpayer
        gov_spending_row = {'Metric': 'Government spending (Tax subsidies)'}
        taxpayer_idx = all_metrics.index('Taxpayer') if 'Taxpayer' in all_metrics else None
        for model_name in model_mapping.values():
            if taxpayer_idx is not None and model_name in aac_df_col.columns:
                gov_spending_row[model_name] = aac_df_col.loc[taxpayer_idx, model_name]
            else:
                gov_spending_row[model_name] = None

        # Row 7: Carbon tax revenue = Carbon_tax_total
        ctax_revenue_row = {'Metric': 'Carbon tax revenue'}
        ctax_idx = all_metrics.index('Carbon_tax_total') if 'Carbon_tax_total' in all_metrics else None
        for model_name in model_mapping.values():
            if ctax_idx is not None and model_name in aac_df_col.columns:
                value = aac_df_col.loc[ctax_idx, model_name]
                ctax_revenue_row[model_name] = value if (value is not None and pd.notna(value)) else 0
            else:
                ctax_revenue_row[model_name] = 0

        # Row 8: Total expenditure
        total_exp_row = {'Metric': 'Total expenditure: Fuel exp. + government spending - ctax revenue'}
        for model_name in model_mapping.values():
            fuel_exp = fuel_exp_row.get(model_name, 0)
            gov_spending = gov_spending_row.get(model_name, 0)
            ctax_revenue = ctax_revenue_row.get(model_name, 0)
            fuel_exp_val = fuel_exp if (fuel_exp is not None and pd.notna(fuel_exp)) else 0
            gov_spending_val = gov_spending if (gov_spending is not None and pd.notna(gov_spending)) else 0
            ctax_revenue_val = ctax_revenue if (ctax_revenue is not None and pd.notna(ctax_revenue)) else 0
            total_exp = fuel_exp_val + gov_spending_val - ctax_revenue_val
            total_exp_row[model_name] = total_exp

        # Row 9: Average Abatement Cost
        aac_cost_row = {'Metric': 'Average Abatement Cost (EV/emissions reduction)'}
        no_policy_total_exp = total_exp_row.get('No policy', 0)
        no_policy_total_exp_val = no_policy_total_exp if (no_policy_total_exp is not None and pd.notna(no_policy_total_exp)) else 0

        for model_name in model_mapping.values():
            policy_total_exp = total_exp_row.get(model_name, 0)
            policy_total_exp_val = policy_total_exp if (policy_total_exp is not None and pd.notna(policy_total_exp)) else 0
            co2_reduction = co2_reduction_row.get(model_name, 0)
            co2_reduction_val = co2_reduction if (co2_reduction is not None and pd.notna(co2_reduction)) else 0
            if co2_reduction_val != 0:
                try:
                    aac_cost_row[model_name] = (policy_total_exp_val - no_policy_total_exp_val) / co2_reduction_val
                except Exception:
                    aac_cost_row[model_name] = None
            else:
                aac_cost_row[model_name] = None

        # Insert the new rows at the beginning
        transportation_df = pd.DataFrame([transportation_row])
        co2_reduction_df = pd.DataFrame([co2_reduction_row])
        pct_reduction_df = pd.DataFrame([pct_reduction_row])
        ev_df = pd.DataFrame([ev_row])
        fuel_exp_df = pd.DataFrame([fuel_exp_row])
        gov_spending_df = pd.DataFrame([gov_spending_row])
        ctax_revenue_df = pd.DataFrame([ctax_revenue_row])
        total_exp_df = pd.DataFrame([total_exp_row])
        aac_cost_df = pd.DataFrame([aac_cost_row])

        new_aac = pd.concat([
            transportation_df,
            co2_reduction_df,
            pct_reduction_df,
            ev_df,
            fuel_exp_df,
            gov_spending_df,
            ctax_revenue_df,
            total_exp_df,
            aac_cost_df,
            aac_df_col
        ], ignore_index=True)
        return new_aac

    # Process each built AAC dataframe: insert calculated rows, add blank row, and store processed dfs
    processed_aac_dfs = {}
    for col, aac_df in aac_dfs.items():
        processed = insert_calculated_rows(aac_df, col)

        # Add a blank row between calculated values and appended values
        blank_row = {'Metric': ''}
        for model_name in model_mapping.values():
            if model_name in processed.columns:
                blank_row[model_name] = None
        blank_df = pd.DataFrame([blank_row])

        # Insert blank row after the first 9 rows (calculated rows)
        try:
            processed = pd.concat([processed.iloc[:9], blank_df, processed.iloc[9:]], ignore_index=True)
        except Exception:
            # If processed has fewer than 9 rows, just prepend the blank row
            processed = pd.concat([blank_df, processed], ignore_index=True)

        processed_aac_dfs[col] = processed
        print(f"  ✓ Added blank row and processed AAC for column: {col}")

    # Reorganize columns in the specified order
    column_order = [
        'Metric',
        'Q mandate',
        'Nested D2',
        'Non-nested D2',
        'Nested D2 + stricter RFS',
        'Non-nested + stricter RFS',
        'Carbon Tax',
        'Aviation intensity standard',
        'Current Policy',
        'Current Policy + No LCFS',
        'Carbon tax+no SAF',
        'No policy'
    ]

    # Write to Excel file
    print(f"\nWriting AAC sheet(s) to {output_file}...")
    try:
        # Decide writer mode
        writer_mode = 'a' if os.path.exists(output_file) else 'w'
        if writer_mode == 'a':
            writer = pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace')
        else:
            writer = pd.ExcelWriter(output_file, engine='openpyxl')

        with writer:
            # Write each processed dataframe to its own sheet
            for col, df in processed_aac_dfs.items():
                # Reorder columns to available subset
                existing_columns = [c for c in column_order if c in df.columns]
                if 'Metric' not in existing_columns:
                    existing_columns = ['Metric'] + [c for c in df.columns if c != 'Metric']
                df_to_write = df[existing_columns]
                # Use uniform sheet name 'AAC' (meta-script writes one workbook per column)
                sheet_name = 'AAC'
                df_to_write.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"  ✓ Wrote sheet: {sheet_name}")

        # Now add the summary table for each written sheet starting at N1
        print(f"\nAdding summary tables at N1 for each AAC sheet...")
        from openpyxl import load_workbook
        wb = load_workbook(output_file)

        summary_policies = [
            'Q mandate',  # SAF Credit
            'Nested D2',
            'Non-nested D2',  # D2 with aviation obligation
            'Nested D2 + stricter RFS',
            'Non-nested + stricter RFS',  # D2 + stricter RFS
            'Carbon Tax',  # Carbon Tax + SAF credit
            'Aviation intensity standard',  # Aviation Intensity Std
            'Current Policy'
        ]
        
        summary_headers = [
            'SAF Credit',
            'Nested D2',
            'D2 with aviation obligation',
            'Nested D2 + stricter RFS',
            'D2 + stricter RFS',
            'Carbon Tax + SAF credit',
            'Aviation Intensity Std',
            'Current Policy'
        ]

        for col, df in processed_aac_dfs.items():
            # Summary table always targets sheet named 'AAC'
            sheet_name = 'AAC'
            ws = wb[sheet_name]

            # Write headers in row 1, starting at column O (column 15)
            for col_idx, header in enumerate(summary_headers, start=15):
                ws.cell(row=1, column=col_idx, value=header)

            # Find the rows for "% of CO2 reduction" and "Average Abatement Cost"
            pct_reduction_row = None
            aac_row = None
            for idx, row in enumerate(df.itertuples(), start=2):  # Start at row 2 (row 1 is headers)
                if row.Metric == '% of CO2 reduction':
                    pct_reduction_row = idx
                elif row.Metric == 'Average Abatement Cost (EV/emissions reduction)':
                    aac_row = idx

            # Write "Emissions Reduction" label in row 2, column N
            ws.cell(row=2, column=14, value='Emissions Reduction')

            # Write emissions reduction values in row 2
            if pct_reduction_row:
                for col_idx, policy in enumerate(summary_policies, start=15):
                    if policy in df.columns:
                        value = df.loc[pct_reduction_row - 2, policy]  # -2 because itertuples starts at 2
                        if value is not None and pd.notna(value):
                            # Format as percentage
                            ws.cell(row=2, column=col_idx, value=value)
                            ws.cell(row=2, column=col_idx).number_format = '0.00%'

            # Write "Average Abatement Cost (EV/emissions reduction)" label in row 3, column N
            ws.cell(row=3, column=14, value='Average Abatement Cost (EV/emissions reduction)')

            # Write AAC values in row 3
            if aac_row:
                for col_idx, policy in enumerate(summary_policies, start=15):
                    if policy in df.columns:
                        value = df.loc[aac_row - 2, policy]  # -2 because itertuples starts at 2
                        if value is not None and pd.notna(value):
                            ws.cell(row=3, column=col_idx, value=value)
                            ws.cell(row=3, column=col_idx).number_format = '$#,##0.00'

        wb.save(output_file)
        print(f"✓ Added summary tables at N1 for all AAC sheets")

        print(f"\n✓ AAC sheet generation complete!")
        print(f"  Columns processed: {list(processed_aac_dfs.keys())}")
        print(f"  Metrics: {len(all_metrics)}")
        print(f"  Models: {len(model_mapping)}")

        return True

    except Exception as e:
        print(f"\nError writing to Excel file: {e}")
        print("\nTrying alternative approach...")

        # Try creating a backup and new file
        if os.path.exists(output_file):
            import shutil
            backup_file = output_file.replace('.xlsx', '_backup.xlsx')
            shutil.copy(output_file, backup_file)
            print(f"  Backup saved to {backup_file}")

        # Write first processed dataframe as fallback
        first_col = next(iter(processed_aac_dfs))
        first_df = processed_aac_dfs[first_col]
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Fallback: write first processed dataframe to sheet 'AAC'
            sheet_name = 'AAC'
            first_df.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"✓ Created new {output_file} with '{sheet_name}' sheet")

        return True


def main():
    """Main function to run the script"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Generate AAC sheet in results_plot.xlsx'
    )
    parser.add_argument(
        '--results-dir', 
        type=str, 
        default='results',
        help='Path to results directory (default: results)'
    )
    parser.add_argument(
        '--output-file',
        type=str,
        default='results_plot.xlsx',
        help='Output Excel file name (default: results_plot.xlsx)'
    )
    parser.add_argument(
        '--columns',
        type=str,
        default=None,
        help='Column name or comma-separated list of columns/quantiles to process (optional)'
    )
    
    args = parser.parse_args()
    
    # Generate the AAC sheet (pass columns through if provided)
    success = generate_aac_sheet(args.results_dir, args.output_file, columns=args.columns)
    
    if success:
        sys.exit(0)
    else:
        sys.exit(1)


if __name__ == '__main__':
    main()
